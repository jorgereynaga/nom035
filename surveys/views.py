from surveys.services.credits import assign_nom035_credits
from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, HttpResponse, JsonResponse
from django.db.models import Q
from django.views.generic import View
from django.utils import timezone
import json
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse,Http404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.forms import UserCreationForm,SetPasswordForm,PasswordChangeForm
from rest_framework import generics,mixins,status
from rest_framework.views import APIView
from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication,SessionAuthentication
from rest_framework.authtoken.models import Token
from django.contrib.postgres.search import SearchVector
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.files.base import ContentFile
from django.contrib.auth import update_session_auth_hash
from django.conf import settings
from celery.result import AsyncResult
from django.views.decorators.csrf import csrf_exempt
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad,unpad
from datetime import datetime,timedelta
from .forms import *
from .models import *
from .serializers import *
from .tasks import pdf_reports_task
from math import ceil
from io import BytesIO
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from django.template.loader import get_template
from django.core.mail import EmailMultiAlternatives

import logging, requests,uuid,time,os,base64,json

import stripe

p_= logging.getLogger(__name__)
#https://n035.page.link/?link=https://normaia.ihes.mx/app/access?workplace_id=280053610240503210429881290965155008225-[1580245165.810466&apn=ihes.com.mx.n035

def download_file(request, user_id,file_name):
	if request.user.is_authenticated:
		userapp = getattr(request.user, "userapp", None)
		if userapp and not getattr(userapp, "stripe_plan_key", ""):
			from django.http import HttpResponse
			return HttpResponse('<h2 style="font-family:sans-serif;text-align:center;margin-top:80px">&#128274; Descarga no disponible en modo demo.<br><a href="/stripe/planes/" style="color:#2563eb">Adquiere un plan para descargar tus resultados</a></h2>', status=403)
	token=None
	if "_-_Token " in file_name:
		token=file_name.split("_-_Token ")[1]
		filename=file_name.split("_-_Token ")[0]
		user = Token.objects.filter(key=token).last().user
		if not user:
			print("invalid user")
			raise Http404
		if user.id != user_id :
			print("invalid user")
			raise Http404
	else:
		filename=file_name
		if request.user.id != user_id:
			print("invalid user")
			raise Http404
	file_path = os.path.join(settings.PROTECTED_MEDIA_ROOT, f"{user_id}/{filename}")
	if os.path.exists(f"{settings.BASE_DIR}/files/tmp/{user_id}"):
		with open(f"{settings.BASE_DIR}/files/tmp/{user_id}/{filename}", 'rb') as fh:
			response = HttpResponse(fh.read(), content_type="application/pdf")
			response['Content-Disposition'] = 'inline; filename=resultados_eventos_traumaticos.pdf' 
			return response
	print("path not found")
	raise Http404
def download_file2(request, workplace_id,evaluation,file_name):
	if request.user.is_authenticated:
		userapp = getattr(request.user, "userapp", None)
		if userapp and not getattr(userapp, "stripe_plan_key", ""):
			return HttpResponse('<h2 style="font-family:sans-serif;text-align:center;margin-top:80px">&#128274; Descarga no disponible en modo demo.<br><a href="/stripe/planes/" style="color:#2563eb">Adquiere un plan para descargar tus resultados</a></h2>', status=403)
	token=None
	if "_-_Token " in file_name:
		token=file_name.split("_-_Token ")[1]
		filename=file_name.split("_-_Token ")[0]
		user = Token.objects.filter(key=token).last().user
		if not user:
			print("invalid user")
			raise Http404
	else:
		filename=file_name
		if not request.user.is_authenticated:
			print("invalid user")
			raise Http404
		if not Workplace.objects.filter(id=workplace_id, user=request.user).exists():
			print("invalid user")
			raise Http404
	file_path = os.path.join(settings.PROTECTED_MEDIA_ROOT, f"charts/{workplace_id}/{evaluation}/{filename}")
	if os.path.exists(f"{settings.BASE_DIR}/files/charts/{workplace_id}/{evaluation}"):
		with open(f"{settings.BASE_DIR}/files/charts/{workplace_id}/{evaluation}/{filename}", 'rb') as fh:
			response = HttpResponse(fh.read(), content_type="application/pdf")
			response['Content-Disposition'] = 'inline; filename=resultados_generales.pdf' 
			return response
	print("path not found")
	raise Http404

@login_required
def download_logo(request):
	userapp = getattr(request.user, "userapp", None)
	if not userapp or not userapp.image:
		raise Http404
	return FileResponse(open(userapp.image.path, 'rb'))

@login_required
def download_result_image(request, workplace_id, result_id):
	if not Workplace.objects.filter(id=workplace_id, user=request.user).exists():
		raise Http404
	result = get_object_or_404(ResultFiles, id=result_id, workplace_id=workplace_id)
	if not result.image:
		raise Http404
	return FileResponse(open(result.image.path, 'rb'))

# Reemplazado en Fase 2-B (2026-07) por checklist de estado. Descomentar si se revierte la eliminacion de subida de archivos.
# @login_required
# def download_evidencia_fase_c(request, evidencia_id):
# 	evidencia = get_object_or_404(EvidenciaFaseC, id=evidencia_id)
# 	if evidencia.workplace.user_id != request.user.id:
# 		raise Http404
# 	if not evidencia.archivo:
# 		raise Http404
# 	return FileResponse(open(evidencia.archivo.path, 'rb'))

def send_mail(to_emails,ctx,template='email-template.html',subject='Tu registro a nuestra plataforma IHES 035 está completo',text_content='Gracias por tu registro con nosotros.'):
	from_email=f'NormaIA <{settings.DEFAULT_FROM_EMAIL}>'
	htmly=get_template(template)
	html_content=htmly.render(ctx)
	msg=EmailMultiAlternatives(subject, text_content, from_email, to_emails)
	msg.attach_alternative(html_content, "text/html")
	return msg.send()

class TACView(View):
	def get(self, request, *args, **kwargs):
		return render(request, 'tyc.html')
class Index(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		wk=[]
		for item in Workplace.objects.filter(user_id=self.request.user.id):
			employees=item.employees.all()
			eval_to_check = item.last_completed_evaluation()
			from django.test import RequestFactory
			factory = RequestFactory()
			fake_req_riesgo = factory.get('/get_riesgo_general/', {'workplace_id': str(item.id), 'evaluation': str(eval_to_check)})
			fake_req_riesgo.user = request.user
			riesgo_data = json.loads(get_riesgo_general(fake_req_riesgo).content)
			if riesgo_data.get('status') == 'ok':
				riesgo_nivel = riesgo_data['riesgo_general']['nivel']
				riesgo_nivel_nombre = riesgo_data['riesgo_general']['nivel_nombre']
			else:
				riesgo_nivel = None
				riesgo_nivel_nombre = 'Sin datos'
			fake_req_portafolio = factory.get('/get_portafolio_status/', {'workplace_id': str(item.id)})
			fake_req_portafolio.user = request.user
			portafolio_data = json.loads(get_portafolio_status(fake_req_portafolio).content)
			cumplimiento_pct_item = portafolio_data.get('porcentaje_cumplimiento', 0)
			print(employees.count())
			print(list(employees.values_list('id',flat=True)))
			if item.survey_type() != 3:
				print("A")
				survey_completed=RiskSurveyA.objects.filter(evaluation=eval_to_check,employee_id__in=list(employees.values_list('id',flat=True))).count()
			elif item.survey_type() == 3:
				print("B")
				survey_completed=RiskSurveyB.objects.filter(evaluation=eval_to_check,employee_id__in=list(employees.values_list('id',flat=True))).count()
			else:
				print("N")
				survey_completed=0
			# Calcular dominios para dashboard (mismos umbrales oficiales de la Guia II usados en get_chart_data)
			dimensions_preview = []
			col_map={0:"#9be5f7",1:"#6bf56e",2:"#ffff00",3:"#ffc000",4:"#ff7070"}
			name_map={0:"Nulo",1:"Bajo",2:"Medio",3:"Alto",4:"Muy alto"}
			domainsA_dash={"Condiciones en el ambiente de trabajo":["r2_p1","r2_p2","r2_p3"],"Carga de trabajo":["r2_p4","r2_p5","r2_p6","r2_p7","r2_p8","r2_p9","r2_p10","r2_p11","r2_p12","r2_p13","r2_p41","r2_p42","r2_p43"],"Falta de control":["r2_p18","r2_p19","r2_p20","r2_p21","r2_p22","r2_p26","r2_p27"],"Jornada de trabajo":["r2_p14","r2_p15"],"Interferencia trabajo-familia":["r2_p16","r2_p17"],"Liderazgo":["r2_p23","r2_p24","r2_p25","r2_p28","r2_p29"],"Relaciones en el trabajo":["r2_p30","r2_p31","r2_p32","r2_p44","r2_p45","r2_p46"],"Violencia":["r2_p33","r2_p34","r2_p35","r2_p36","r2_p37","r2_p38","r2_p39","r2_p40"]}
			domainsA_dash_thresholds={"Condiciones en el ambiente de trabajo":[3,5,7,9],"Carga de trabajo":[12,16,20,24],"Falta de control":[5,8,11,14],"Jornada de trabajo":[1,2,4,6],"Interferencia trabajo-familia":[1,2,4,6],"Liderazgo":[3,5,8,11],"Relaciones en el trabajo":[5,8,11,14],"Violencia":[7,10,13,16]}
			if item.survey_type() != 3 and survey_completed > 0:
				survey_all=RiskSurveyA.objects.filter(evaluation=eval_to_check,employee_id__in=list(employees.values_list("id",flat=True)))
				if survey_all.exists():
					for domain,fields in domainsA_dash.items():
						max_score=len(fields)*4
						sums=[]
						for s in survey_all:
							_sum=sum((getattr(s,f) or 0) for f in fields if hasattr(s,f))
							sums.append(_sum)
						avg_sum=sum(sums)/len(sums) if sums else 0
						pct=round((avg_sum/max_score)*100) if max_score else 0
						t=domainsA_dash_thresholds[domain]
						color_idx=0 if avg_sum<t[0] else (1 if avg_sum<t[1] else (2 if avg_sum<t[2] else (3 if avg_sum<t[3] else 4)))
						dimensions_preview.append({"name":domain,"pct":pct,"color":col_map[color_idx],"nivel":name_map[color_idx]})

			# Dimensiones clima laboral
			climate_dims=[]
			CLIMA_DIM_FIELDS={"Liderazgo y supervision":[1,2,3,4,5],"Comunicacion interna":[6,7,8,9,10],"Trabajo en equipo":[11,12,13,14,15],"Reconocimiento y motivacion":[16,17,18,19,20],"Condiciones de trabajo":[21,22,23,24,25],"Carga laboral y equilibrio":[26,27,28,29,30],"Desarrollo y crecimiento":[31,32,33,34,35],"Sentido de pertenencia":[36,37,38,39,40]}
			climate_surveys=item.climate_surveys.all()
			if climate_surveys.exists():
				for dim_name,fields in CLIMA_DIM_FIELDS.items():
					promedios=[]
					for s in climate_surveys:
						vals=[getattr(s,f"cl_p{i}") for i in fields if getattr(s,f"cl_p{i}") is not None]
						if vals: promedios.append(sum(vals)/len(vals))
					if promedios:
						prom=round(sum(promedios)/len(promedios),2)
						if prom<2.5: nivel,color="Critico","#ff7070"
						elif prom<3.5: nivel,color="En riesgo","#ffc000"
						elif prom<4.25: nivel,color="Adecuado","#ffff00"
						else: nivel,color="Favorable","#6bf56e"
						climate_dims.append({"name":dim_name,"prom":prom,"nivel":nivel,"color":color})
			wk.append({"id":item.id,"name":item.name,"address":item.address,"employee_count":item.employee_num,
			"access_code":f"https://normaia.ihes.mx/app/access/?d={item.access_code}",
			"employee_completion":ceil((employees.count()/item.employee_num)*100),
			"employee_total":employees.count(),"cat":item.survey_type(),
			"survey_completed":survey_completed,
			"survey_completion":ceil((survey_completed/item.employee_num)*100),
			"climate_surveys_count":item.climate_surveys.count(),
			"raw_access_code":item.access_code,
			"eval_to_check":eval_to_check,
			"dimensions_preview":dimensions_preview,
			"climate_dimensions_preview":climate_dims,
			"riesgo_nivel":riesgo_nivel,
			"riesgo_nivel_nombre":riesgo_nivel_nombre,
			"cumplimiento_pct":cumplimiento_pct_item,
			"empleados_registrados":employees.count(),
			"evaluaciones_aplicadas":max(0, item.evaluation - 1),
			})
		ctx={"workplaces":wk}
		ctx['workplaces_available']=request.user.userapp.workplaces_available
		ctx['has_workplaces']=Workplace.objects.filter(user_id=self.request.user.id).exists()
		ctx['name']=request.user.userapp.name
		ctx['dashboard_msg']=request.GET.get('msg','')
		ctx['dashboard_msg']=request.GET.get('msg','')
		ctx['phone']=request.user.userapp.phone
		ctx['email']=request.user.email
		ctx['len_workplaces']=len(ctx["workplaces"])
		ctx['total_employees']=sum(item['employee_count'] for item in wk)
		ctx['total_surveys']=sum(item['survey_completed'] for item in wk)
		ctx['candidates']=[{'id':c.id,'nombre':c.nombre,'email':c.email,'puesto':c.puesto,'tipo':c.tipo,'completed':c.sessions.filter(status='completada').exists()} for c in Candidate.objects.filter(user=request.user).order_by('-record_create')[:3]]
		userapp=request.user.userapp
		ctx['psico_disponibles']=getattr(userapp,'psico_evaluaciones_disponibles',0) + getattr(userapp,'psico_demo',0)
		ctx['nom035_demo']=getattr(userapp,'nom035_demo',0)
		ctx['tiene_datos_demo']=Workplace.objects.filter(user=request.user, es_demo=True).exists()
		ctx['psico_demo']=getattr(userapp,'psico_demo',0)
		ctx['nom035_disponibles']=getattr(userapp,'nom035_creditos',0) + getattr(userapp,'nom035_demo',0)
		NIVEL_NOMBRE = {0:"Nulo",1:"Bajo",2:"Medio",3:"Alto",4:"Muy alto"}
		total_centros = len(wk)
		niveles_presentes = [w['riesgo_nivel'] for w in wk if w['riesgo_nivel'] is not None]
		riesgo_predominante_nivel = max(niveles_presentes) if niveles_presentes else None
		ctx['kpi_total_centros'] = total_centros
		ctx['kpi_empleados_totales'] = sum(w['empleados_registrados'] for w in wk)
		ctx['kpi_evaluaciones_totales'] = sum(w['evaluaciones_aplicadas'] for w in wk)
		ctx['kpi_cumplimiento_promedio'] = round(sum(w['cumplimiento_pct'] for w in wk) / total_centros) if total_centros else 0
		ctx['kpi_riesgo_predominante_nivel'] = riesgo_predominante_nivel
		ctx['kpi_riesgo_predominante_nombre'] = NIVEL_NOMBRE[riesgo_predominante_nivel] if riesgo_predominante_nivel is not None else 'Sin datos'
		centros_ordenados = sorted(wk, key=lambda w: w['riesgo_nivel'] if w['riesgo_nivel'] is not None else -1, reverse=True)
		ctx['centros_atencion'] = [w for w in centros_ordenados if w['riesgo_nivel'] is not None][:3]
		todas_las_dimensiones_clima = []
		suma_climate_surveys = 0
		suma_employee_num = 0
		for w in wk:
			todas_las_dimensiones_clima.extend(w['climate_dimensions_preview'])
			suma_climate_surveys += w['climate_surveys_count']
			suma_employee_num += w['employee_count']
		if todas_las_dimensiones_clima:
			promedio_general_clima = round(sum(d['prom'] for d in todas_las_dimensiones_clima) / len(todas_las_dimensiones_clima), 2)
			peor_dimension = min(todas_las_dimensiones_clima, key=lambda d: d['prom'])
		else:
			promedio_general_clima = None
			peor_dimension = None
		ctx['clima_resumen'] = {
			'promedio_general': promedio_general_clima,
			'tasa_respuesta_pct': round((suma_climate_surveys / suma_employee_num) * 100) if suma_employee_num else 0,
			'peor_dimension': peor_dimension,
		}
		hoy = timezone.now().date()
		acciones_qs = PlanAccionItem.objects.filter(workplace__user=request.user).exclude(estado='completado').order_by('fecha_programada')[:5]
		ctx['acciones_pendientes'] = [{
			'id': a.id,
			'tipo_accion': a.tipo_accion,
			'workplace_nombre': a.workplace.name,
			'responsable': a.responsable,
			'fecha_programada': a.fecha_programada.strftime('%d/%m/%Y'),
			'es_vencida': a.fecha_programada < hoy,
		} for a in acciones_qs]
		plan_key=getattr(userapp,'stripe_plan_key','')
		if plan_key:
			try:
				from surveys.stripe_plans import PLANS
				plan=PLANS.get(plan_key,{})
				ctx['plan_activo']=plan.get('name','')
				ctx['plan_precio']='{:,}'.format(plan.get('precio',0))
				ctx['plan_periodo']=plan.get('periodo','')
				ctx['plan_empleados_max']=plan.get('empleados_max')
				ctx['plan_evaluaciones']=plan.get('evaluaciones_mes')
			except Exception:
				ctx['plan_activo']=None
		else:
			ctx['plan_activo']=None
		psico_key=getattr(userapp,'psico_plan_key','')
		if psico_key:
			try:
				from surveys.stripe_plans import PLANS
				psico_plan=PLANS.get(psico_key,{})
				ctx['psico_plan_activo']=psico_plan.get('name','')
				ctx['psico_plan_precio']='{:,}'.format(psico_plan.get('precio',0))
				ctx['psico_plan_periodo']=psico_plan.get('periodo','')
				ctx['psico_plan_evaluaciones']=psico_plan.get('evaluaciones_mes') or psico_plan.get('evaluaciones_total')
			except Exception:
				ctx['psico_plan_activo']=None
		else:
			ctx['psico_plan_activo']=None
		return render(request, 'index.html',ctx)
		
class SaveAnswers(generics.GenericAPIView):
	serializer_class = WorkplaceSerializer
	def post(self, request, *args, **kwargs):
		#[null, name[omar.mendoza], gender[0], age[6], civil_state[3], study_level[4], ocupation[omar.mendoza], department[omar.mendoza], charge_type[2], contract_type[2], employee_type[2], shift_type[1], shift_rotation[0], time_in_charge[4]]
		#[{'name': 'asd'}, {'gender': '1'}, {'age': '10'}, {'civil_state': '2'}, {'study_level': '2'}, {'ocupation': 'eeee'}, {'department': 'wwww'}, {'charge_type': '2'}, {'contract_type': '3'}, {'employee_type': '1'}, {'shift_type': '0'}, {'shift_rotation': '0'}, {'time_in_charge': '5'}, {'exp': '5'}]
		if 'data' in self.request.query_params:
			data=self.request.query_params.getlist('data')
			answers=[]
			for item in data:
				name=item.split('[')[0]
				value=item.split('[')[1]
				value=value[:-1]
				value=int(value)+1 if name =='gender'else value
				answers.append({f"{name}":f"{value}" })
		else:
			answers=request.data['answers']
		if 'employee_id' in self.request.query_params:
			employee_id=self.request.query_params['employee_id']
		else:
			employee_id=request.data['employee_id']
		if 'workplace_id' in self.request.query_params:
			workplace_id=self.request.query_params['workplace_id']
		else:
			workplace_id=request.data['workplace_id']
		if 'form' in self.request.query_params:
			form=self.request.query_params['form']
		else:
			form=request.data['form']
		workplace=Workplace.objects.get(id=workplace_id)
		dic={'workplace':workplace_id,'evaluation':workplace.evaluation}
		for item in answers:
			dic.update(item)
		if form=='employee':
			if workplace.employee_num>workplace.employees.all().count():
				serializer=EmployeeSerializer(data=dic)
				next_form=0
			else:
				return Response("No quedan empleados para registrar", status=status.HTTP_400_BAD_REQUEST)
		elif form=='traumasurvey':
			dic['employee']=employee_id
			serializer=TraumaSurveySerializer(data=dic)
			if workplace.survey_type()==3:
				next_form=2
			elif workplace.survey_type()!=3:
				next_form=1
		elif form=='risksurveya':
			dic['employee']=employee_id
			serializer=RiskSurveyASerializer(data=dic)
			next_form=3
		elif form=='risksurveyb':
			dic['employee']=employee_id
			serializer=RiskSurveyBSerializer(data=dic)
			next_form=3
		else:
			return Response('invalid form', status=status.HTTP_400_BAD_REQUEST)
		if serializer.is_valid():
			if form in ("risksurveya","risksurveyb","traumasurvey"):
				userapp = workplace.user.userapp
				if workplace.es_demo:
					return Response('Modo demo: adquiere un plan para registrar encuestas reales.', status=status.HTTP_403_FORBIDDEN)
				if userapp.nom035_creditos <= 0:
					return Response('Sin creditos disponibles. Adquiere un plan.', status=status.HTTP_403_FORBIDDEN)
				try:
					with transaction.atomic():
						data = serializer.save()
						userapp.nom035_creditos -= 1
						userapp.save()
				except IntegrityError:
					return Response('Ya existe una respuesta registrada para este empleado en esta evaluacion.', status=status.HTTP_409_CONFLICT)
			else:
				data = serializer.save()
			if form=="employee":
				code=data.get_code()
			else:
				code=data.employee.get_code()
			return Response({"status":"ok","employee_url":code,"next_form":next_form}, status=status.HTTP_201_CREATED)

		if form in ("risksurveya","risksurveyb","traumasurvey") and serializer.Meta.model.objects.filter(employee_id=employee_id, evaluation=workplace.evaluation).exists():
			return Response('Ya existe una respuesta registrada para este empleado en esta evaluacion.', status=status.HTTP_409_CONFLICT)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class SaveCharts(APIView):
	http_method_names = ['post']
	permission_classes = (IsAuthenticated,)
	authentication_classes = (TokenAuthentication,SessionAuthentication)
	def post(self, request):
		chart1=request.data["chart1"] or None
		chart2=request.data["chart2"] or None
		chart3=request.data["chart3"] or None
		chart4=request.data["chart4"] or None
		wp=request.data["wp"] or None
		ev=request.data["ev"] or None
		if chart1 is not None:
			form, imgstr = chart1.split(';base64,')
			ext = form.split('/')[-1] 
			b64=base64.b64decode(imgstr)
			filename1='chart1.' + ext
			data1 = ContentFile(b64, name=filename1)
		if chart2 is not None:
			form, imgstr = chart2.split(';base64,') 
			ext = form.split('/')[-1] 
			b64=base64.b64decode(imgstr)
			filename2='chart2.' + ext
			data2 = ContentFile(b64, name=filename2)
		if chart3 is not None:
			form, imgstr = chart3.split(';base64,') 
			ext = form.split('/')[-1] 
			b64=base64.b64decode(imgstr)
			filename3='chart3.' + ext
			data3 = ContentFile(b64, name=filename3)
		if chart4 is not None:
			form, imgstr = chart4.split(';base64,') 
			ext = form.split('/')[-1] 
			b64=base64.b64decode(imgstr)
			filename4='chart4.' + ext
			data4 = ContentFile(b64, name=filename4)
		print(request.data)
		if chart1 or chart2 or chart3 or chart4:
			if not os.path.exists(f"{settings.BASE_DIR}/files/charts/{wp}"):
				os.mkdir(f"{settings.BASE_DIR}/files/charts/{wp}")
				if not os.path.exists(f"{settings.BASE_DIR}/files/charts/{wp}/{ev}"):
					os.mkdir(f"{settings.BASE_DIR}/files/charts/{wp}/{ev}")
			image = open(f"{settings.BASE_DIR}/files/charts/{wp}/{ev}/{filename1}", "wb")
			for chunk in data1.chunks():
				image.write(chunk)
			image.close()
			image = open(f"{settings.BASE_DIR}/files/charts/{wp}/{ev}/{filename2}", "wb")
			for chunk in data2.chunks():
				image.write(chunk)
			image.close()
			image = open(f"{settings.BASE_DIR}/files/charts/{wp}/{ev}/{filename3}", "wb")
			for chunk in data3.chunks():
				image.write(chunk)
			image.close()
			image = open(f"{settings.BASE_DIR}/files/charts/{wp}/{ev}/{filename4}", "wb")
			for chunk in data4.chunks():
				image.write(chunk)
			image.close()
			template="results"
			workplace=Workplace.objects.filter(id=wp).last()
			token = Token.objects.filter(user_id=self.request.user.id).last()
			if not token:
				return Response("Vuelve a iniciar sesión",status=400)
			token = token.key
			ctx={"user":self.request.user.id,
				"name":workplace.name,
				"address":workplace.address,
				"address_locality":workplace.address_locality,
				"address_state":workplace.address_state,
				"address_postal_code":workplace.address_postal_code,
				"phone":self.request.user.userapp.phone,
				"chart1":f"https://normaia.ihes.mx/files/charts/{wp}/{ev}/{filename1}_-_Token {token}",
				"chart2":f"https://normaia.ihes.mx/files/charts/{wp}/{ev}/{filename2}_-_Token {token}",
				"chart3":f"https://normaia.ihes.mx/files/charts/{wp}/{ev}/{filename3}_-_Token {token}",
				"chart4":f"https://normaia.ihes.mx/files/charts/{wp}/{ev}/{filename4}_-_Token {token}",
			}
			task=pdf_reports_task.delay(self.request.user.id,template,ctx)
			time.sleep(7)
			t=AsyncResult(task.id)
			if t.successful():
				p_.error("task succeded")
				# else:
				# 	response=f"https://normaia.ihes.mx/files/tmp/{self.request.user.id}/{template}.pdf"
			return Response({"pdf":f"https://normaia.ihes.mx/files/tmp/{self.request.user.id}/{template}.pdf"})
		else:
			return Response("error",status=400)

class TokenCreation(generics.GenericAPIView):
	serializer_class = WorkplaceSerializer
	def post(self, request, *args, **kwargs):
		p_.error(request.data)
		print(request.data)
		username=request.data['username']
		password=request.data['password']
		if username=='':
			return Response("Debes indicar tu correo electrónico", status=status.HTTP_401_UNAUTHORIZED)
		if password=='':
			return Response("Debes ingresar tu contraseña", status=status.HTTP_401_UNAUTHORIZED)
		if username!='' and password!='':
			user = authenticate(username=username, password=password)
			if user is not None and user.is_active:
				if user.userapp.validated_email:
					print(user.id)
					try:
						try:
							token=Token.objects.create(user=user)
						except IntegrityError as e:
							token=Token.objects.get(user=user)
							return Response({
								"response": f'The user {user.get_username()} already has a token.',
								"id": user.id,
								"name": user.userapp.name,
								"mail": user.get_username(),
								"token": token.key
							})
						return Response({
							"response": f'The token was generated for user: {user.get_username()}',
							"id": user.id,
							"name": user.userapp.name,
							"mail": user.get_username(),
							"token": token.key
						})
					except Exception as e:
						print(e)
				else:
					logout(request)
					return Response("Verifica tu correo electrónico antes de iniciar sesión.", status=status.HTTP_401_UNAUTHORIZED)
		return Response("Verifica tus datos, alguno está mal", status=status.HTTP_401_UNAUTHORIZED)

class LoginView(View):
	def get(self, request, *args, **kwargs):
		if "logout" in request.path:			
			logout(request)
		ctx = {'form':LoginForm(),'msg':""}
		return render(request, 'auth-login.html', ctx)

	def post(self, request):
		msg = ""
		form = LoginForm(request.POST)
		if form.is_valid():
			us = form.cleaned_data['username']
			pw = form.cleaned_data['password']
			user = authenticate(username=us, password=pw)
			if user is not None and user.is_active:
				if user.userapp.validated_email:
					login(request, user)
					if 'redirect' in request.POST:
						return HttpResponseRedirect(request.POST['redirect'])
					elif not user.userapp.stripe_plan_key:
						return HttpResponseRedirect(reverse_lazy('index'))
					else:
						return HttpResponseRedirect(reverse_lazy('index'))
				else:
					logout(request)
					messages.success(request, "not_verified_email")
			else:
				messages.success(request, "wrong_credentials")
		else:
			form = LoginForm()
		ctx = {}
		ctx['form'] = form
		ctx['msg'] = msg
		return render(request, 'auth-login.html', ctx)
class ApiLoginView(View):
    def post(self, request):
        import json
        try:
            data = json.loads(request.body)
            username = data.get('username')
            password = data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None and user.is_active and user.userapp.validated_email:
                login(request, user)
                return JsonResponse({'status': 'ok'})
            return JsonResponse({'error': 'Credenciales invalidas'}, status=401)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

def capture_referral(request):
    ref = request.GET.get('ref')
    if ref:
        request.session['referral_code'] = ref.strip().upper()

class LandingView(View):
    def get(self, request):
        capture_referral(request)
        if request.user.is_authenticated:
            return redirect('index')
        return render(request, 'landing.html')
class LandingNom035View(View):
	def get(self, request, *args, **kwargs):
		return render(request, 'landing_nom035.html')
class LandingPsicometriaView(View):
	def get(self, request, *args, **kwargs):
		return render(request, 'landing_psicometria.html')
class LandingClimaView(View):
	def get(self, request, *args, **kwargs):
		return render(request, 'landing_clima.html')
class LandingContactoView(View):
	def get(self, request, *args, **kwargs):
		return render(request, 'landing_contacto.html')
class NewUserView(View):
	def get(self, request, *args, **kwargs):
		capture_referral(request)
		ctx = {'recaptcha_site_key': settings.RECAPTCHA_SITE_KEY, 'referral_code': request.session.get('referral_code', '')}
		return render(request, 'auth-register.html', ctx)
class WorkplaceView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		from django.test import RequestFactory
		factory = RequestFactory()
		NIVEL_NOMBRE = {0:"Nulo",1:"Bajo",2:"Medio",3:"Alto",4:"Muy alto"}

		workplaces_qs = Workplace.objects.filter(user_id=self.request.user.id)
		workplaces = []
		niveles_riesgo_presentes = []
		suma_cumplimiento = 0
		suma_evaluaciones = 0
		suma_empleados_registrados = 0

		for wk in workplaces_qs:
			eval_to_check = wk.last_completed_evaluation()
			evaluacion_finalizada = wk.evaluation_history.filter(numero_evaluacion=eval_to_check).exists()

			fake_req = factory.get('/get_riesgo_general/', {'workplace_id': str(wk.id), 'evaluation': str(eval_to_check)})
			fake_req.user = request.user
			riesgo_data = json.loads(get_riesgo_general(fake_req).content)
			if riesgo_data.get('status') == 'ok':
				nivel = riesgo_data['riesgo_general']['nivel']
				nivel_nombre = riesgo_data['riesgo_general']['nivel_nombre']
				niveles_riesgo_presentes.append(nivel)
			else:
				nivel = None
				nivel_nombre = 'Sin datos'

			fake_req2 = factory.get('/get_portafolio_status/', {'workplace_id': str(wk.id)})
			fake_req2.user = request.user
			portafolio_data = json.loads(get_portafolio_status(fake_req2).content)
			cumplimiento_pct = portafolio_data.get('porcentaje_cumplimiento', 0)

			evaluaciones_aplicadas = max(0, wk.evaluation - 1)
			empleados_registrados = wk.employees.count()

			employee_ids = list(wk.employees.values_list('id', flat=True))
			if wk.survey_type() != 3:
				survey_completed = RiskSurveyA.objects.filter(evaluation=eval_to_check, employee_id__in=employee_ids).count()
			else:
				survey_completed = RiskSurveyB.objects.filter(evaluation=eval_to_check, employee_id__in=employee_ids).count()
			survey_completion_pct = ceil((survey_completed / wk.employee_num) * 100) if wk.employee_num else 0
			if survey_completed == 0:
				evaluacion_estado = 'sin_aplicar'
			elif survey_completion_pct >= 100:
				evaluacion_estado = 'concluida'
			else:
				evaluacion_estado = 'en_curso'

			suma_cumplimiento += cumplimiento_pct
			suma_evaluaciones += evaluaciones_aplicadas
			suma_empleados_registrados += empleados_registrados

			workplaces.append({
				"id": wk.id,
				"name": wk.name,
				"address": wk.address,
				"employee_count": wk.employee_num,
				"empleados_registrados": empleados_registrados,
				"evaluaciones_aplicadas": evaluaciones_aplicadas,
				"cumplimiento_pct": cumplimiento_pct,
				"riesgo_nivel": nivel,
				"riesgo_nivel_nombre": nivel_nombre,
				"survey_completion_pct": survey_completion_pct,
				"evaluacion_estado": evaluacion_estado,
				"eval_to_check": eval_to_check,
				"evaluacion_finalizada": evaluacion_finalizada,
			})

		total_centros = len(workplaces)
		riesgo_predominante_nivel = max(niveles_riesgo_presentes) if niveles_riesgo_presentes else None
		riesgo_predominante_nombre = NIVEL_NOMBRE[riesgo_predominante_nivel] if riesgo_predominante_nivel is not None else 'Sin datos'

		ctx = {
			"workplaces": workplaces,
			"kpi_total_centros": total_centros,
			"kpi_empleados_totales": suma_empleados_registrados,
			"kpi_evaluaciones_totales": suma_evaluaciones,
			"kpi_cumplimiento_promedio": round(suma_cumplimiento / total_centros) if total_centros else 0,
			"kpi_riesgo_predominante_nivel": riesgo_predominante_nivel,
			"kpi_riesgo_predominante_nombre": riesgo_predominante_nombre,
		}
		return render(request, 'workplace.html', ctx)
class SupportView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	AREAS = {
		'soporte': 'Soporte técnico',
		'compras': 'Compras y facturación',
		'administracion': 'Administración',
	}
	def get(self, request, *args, **kwargs):
		nombre = ''
		telefono = ''
		if hasattr(request.user, 'userapp'):
			nombre = request.user.userapp.name
			telefono = request.user.userapp.phone
		ctx = {"nombre": nombre, "telefono": telefono, "correo": request.user.email}
		return render(request, 'soporte.html', ctx)
	def post(self, request, *args, **kwargs):
		area_key = request.POST.get('area', '')
		area = self.AREAS.get(area_key, area_key)
		message = request.POST.get('message', '')
		name = request.POST.get('name', '')
		email = request.POST.get('email', '')
		phone = request.POST.get('phone', '')
		if message == '' or area_key not in self.AREAS:
			return JsonResponse({'status': 'error'}, status=400)
		ctx = {"cname": name, "phone": phone, "email": email, "name": name, "area": area, "message": message,
			"type": 2, "title": "Soporte", "date_today": datetime.now().strftime('%d/%m/%Y')}
		send_mail(["normaia.sistemas@gmail.com"], ctx=ctx, subject=f'Soporte ({area}) — {request.user.email}')
		return JsonResponse({'status': 'ok'})
class PartnerDashboardView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		partner = Partner.objects.filter(user=request.user).first()
		if not partner:
			return HttpResponseRedirect(reverse_lazy('index'))
		from django.db.models import Sum
		from surveys.stripe_plans import PLANS as STRIPE_PLANS
		referidos = partner.referred_clients.select_related('user').order_by('-record_create')
		clientes = []
		activos = 0
		for u in referidos:
			plan_key = u.stripe_plan_key or u.psico_plan_key
			tuvo_compra = PlanPurchaseEvent.objects.filter(user=u.user).exists()
			if plan_key:
				estado = 'activo'
				activos += 1
				plan_nombre = STRIPE_PLANS.get(plan_key, {}).get('name', plan_key)
				valor_plan = STRIPE_PLANS.get(plan_key, {}).get('precio', 0)
			elif tuvo_compra:
				estado = 'cancelado'
				plan_nombre = 'Sin plan activo'
				valor_plan = None
			else:
				estado = 'prueba'
				plan_nombre = 'Sin plan / prueba'
				valor_plan = None
			comision = PartnerCommission.objects.filter(partner=partner, referred_userapp=u).first()
			clientes.append({
				'nombre': u.name,
				'correo': u.user.email,
				'plan_nombre': plan_nombre,
				'valor_plan': valor_plan,
				'fecha_alta': u.record_create,
				'estado': estado,
				'comision': comision.monto_comision if comision else None,
			})
		commissions = PartnerCommission.objects.filter(partner=partner)
		total_generado = commissions.aggregate(total=Sum('monto_comision'))['total'] or 0
		total_pendiente = commissions.filter(estado='pendiente').aggregate(total=Sum('monto_comision'))['total'] or 0
		pagos = commissions.filter(estado='pagada').order_by('-fecha_pago')
		ref_link = request.build_absolute_uri(reverse_lazy('newuser')) + f'?ref={partner.code}'
		ctx = {
			'partner': partner,
			'ref_link': ref_link,
			'clientes': clientes,
			'total_referidos': len(clientes),
			'total_activos': activos,
			'total_generado': total_generado,
			'total_pendiente': total_pendiente,
			'pagos': pagos,
		}
		return render(request, 'partner_dashboard.html', ctx)
class EditProfileView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		ctx={"msg":'',"msgok":'',"workplaces":[{"id":item.id,"name":item.name,"employee_count":item.employee_num
		} for item in Workplace.objects.filter(user_id=self.request.user.id)]}
		ctx["cards"]=[{"last4":item.last4,"exp_month":item.exp_month,"exp_year":item.exp_year,
			"brand":item.brand,"name":item.name,"card_token":item.card_token
			} for item in PaymentCard.objects.filter(user_id=self.request.user.id)]
		ctx["purchases"]=[]
		for payment in Payment.objects.filter(user_id=self.request.user.id):
			method=payment.purchased_products.last().payment_method
			prods=[]
			amount=0
			for item in payment.purchased_products.all():
				amount=amount+(item.product.price*item.quantity)
				prods.append({"product":item.product.name,"name":item.workplace.name,"quantity":item.workplace.employee_num,})
			ctx["purchases"].append({
				"last4":method.last4,"exp_month":method.exp_month,"exp_year":method.exp_year,
				"brand":method.brand,"name":method.name,"products":prods,"amount":amount,
			})
		ctx['tab']=self.request.GET.get('d','')
		return render(request, 'edit_profile.html',ctx)
	def post(self,request):
		msg=""
		msgok=""
		if "new_password1" in request.POST:
			form = PasswordChangeForm(request.user, request.POST)
			if form.is_valid():
				user = form.save()
				update_session_auth_hash(request, user)
				msgok=f"Su nueva contraseña ha sido guardada."
			msg=f"{form.errors}"
		else:
			name=request.POST.get('name','')
			username=request.POST.get('email','')
			phone=request.POST.get('phone','')
			file=None
			if "file" in request.FILES:
				if request.FILES['file']!='':
					file=request.FILES['file']
			user=User.objects.get(id=request.user.id)
			userapp=user.userapp
			if username !='':
				user.username=username
				user.email=username
				user.save()
			if phone !='':
				userapp.phone=phone
				userapp.save()
			if name!='':
				userapp.name=name
				userapp.save()
			if file is not None:
				userapp.image.save(file.name,file)
		ctx={"msg":msg,"msgok":msgok,"workplaces":[{"id":item.id,"name":item.name,"employee_count":item.employee_num
		} for item in Workplace.objects.filter(user_id=self.request.user.id)]}
		return render(request, 'edit_profile.html',ctx)

class WorkplaceDetailView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		ctx={"workplaces":[{"id":item.id,"name":item.name,"employee_count":item.employee_num
		} for item in Workplace.objects.filter(user_id=self.request.user.id)]}
		if 'workplace_id' in kwargs:
			ctx['workplace_id']=kwargs['workplace_id']
			wk=Workplace.objects.filter(id=kwargs['workplace_id']).last()
			ctx['name']=wk.name
			ctx['access_code']=wk.access_code
			ctx['address']=wk.address
			ctx['main_activity']=wk.main_activity
			ctx['evaluation']=wk.evaluation
			ctx['last_evaluation']=wk.evaluation
			ctx['paid']=wk.paid
			ctx['employees_available']=True if wk.employee_num>wk.employees.all().count() else False
			ctx['employees']=wk.employees.all().count()
			ctx['employees_total']=wk.employee_num
			if not request.user.workplaces.filter(id=kwargs['workplace_id']).exists():
				return HttpResponseRedirect(reverse_lazy('workplaces'))
		else:
			return HttpResponseRedirect(reverse_lazy('workplaces'))
		if 'evaluation' in kwargs:
			wk=Workplace.objects.filter(id=kwargs['workplace_id']).last()
			ctx['last_evaluation']=wk.evaluation
			ctx['evaluation']=kwargs['evaluation']
		historial_nums=set(wk.evaluation_history.values_list('numero_evaluacion', flat=True))
		ctx['evaluations']=[{'numero':n, 'done':n in historial_nums} for n in range(1, wk.evaluation+1)]
		ctx['viewing_done']=ctx['evaluation'] in historial_nums
		ctx['viewing_is_current']=(ctx['evaluation']==wk.evaluation)
		ctx['areas']=wk.areas
		return render(request, 'workplace_detail.html',ctx)
class UpdateWorkplaceAreasView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def post(self, request, workplace_id):
		wk = Workplace.objects.filter(id=workplace_id, user_id=request.user.id).first()
		if not wk:
			return JsonResponse({'status':'error','error':'Centro de trabajo no encontrado.'}, status=404)
		areas = [a.strip() for a in request.POST.getlist('areas[]') if a.strip()]
		if not areas:
			return JsonResponse({'status':'error','error':'Debes tener al menos un área de trabajo.'}, status=400)
		wk.areas = areas
		wk.save()
		return JsonResponse({'status':'ok','areas':wk.areas})
class WorkplaceResultView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		ctx={"workplaces":[{"id":item.id,"name":item.name,"employee_count":item.employee_num
		} for item in Workplace.objects.filter(user_id=self.request.user.id)]}
		if 'workplace_id' in kwargs:
			ctx['workplace_id']=kwargs['workplace_id']
			wk=Workplace.objects.filter(id=kwargs['workplace_id']).last()
			ctx['paid']=wk.paid
			# if not wk.paid:
			#	return HttpResponseRedirect(reverse_lazy('payments'))
			ctx['name']=wk.name
			ctx['access_code']=wk.access_code
			ctx['address']=wk.address
			ctx['main_activity']=wk.main_activity
			ctx['evaluation']=wk.evaluation
			ctx['last_evaluation']=wk.evaluation
			h=200+(20*wk.employees.all().count())
			ctx["chart_height"]=h+300 if h>600 else 900
			ctx["chart_hb"]=h if h>600 else 600
			if not request.user.workplaces.filter(id=kwargs['workplace_id']).exists():
				return HttpResponseRedirect(reverse_lazy('workplaces'))
		else:
			return HttpResponseRedirect(reverse_lazy('workplaces'))
		if 'evaluation' in kwargs:
			wk=Workplace.objects.filter(id=kwargs['workplace_id']).last()
			ctx['last_evaluation']=wk.evaluation
			ctx['evaluation']=kwargs['evaluation']
		return render(request, 'workplace_results.html',ctx)

def encript(text):
	key=settings.AES_ENCRYPTION_KEY
	enc=AES.new(key.encode(),AES.MODE_CBC)
	cipher=enc.encrypt(pad(text.encode(),AES.block_size))
	return (cipher.hex(),enc.iv.hex())

def decript(text,iv):
	key=settings.AES_ENCRYPTION_KEY
	decrypt=AES.new(key.encode(),AES.MODE_CBC,iv=bytes.fromhex(iv))
	# ddata=decrypt.decrypt(text2)
	return unpad(decrypt.decrypt(bytes.fromhex(text)),AES.block_size)

class EmailVerification(View):
	def get(self, request, *args, **kwargs):
		if request.user.is_authenticated:
			logout(request)
			return HttpResponseRedirect(reverse_lazy('login'))
		ctx={"resend_verification":True}
		if 'code' in kwargs and 'iv' in kwargs:
			code=f"0x{kwargs['code']}"
			iv=f"0x{kwargs['iv']}"
			valid_code=False
			try:
				text=decript(hex(int(code, 16))[2:],hex(int(iv, 16))[2:])
				data=text.decode().split('<->')
				if datetime.now()<(datetime.fromtimestamp(float(data[1]))+timedelta(days=1)):
					user=Userapp.objects.filter(user_id=data[0],record_create=datetime.fromtimestamp(float(data[2]))).last()
					if user:
						valid_code=True
						ctx["email"]=user.user.email
						ctx["name"]=user.name
						user.validated_email=True
						user.save()
				else:
					ctx["expired"]=True
			except:
				valid_code=False
			ctx["valid_code"]=valid_code
			ctx["resend_verification"]=False
		return render(request, 'valid_email.html',ctx)
	def post(self, request, *args, **kwargs):
		ct={"resend_verification":True}
		email=request.POST.get('email','')
		if email!="":
			user=User.objects.filter(username=email).last()
			if user:
				if not user.userapp.validated_email:
					verification_code,iv=encript(f"{user.id}<->{datetime.now().timestamp()}<->{user.userapp.record_create.timestamp()}")
					ctx={"v_code":verification_code,"token":iv,"username":user.username,"phone":user.userapp.phone,"email":user.email,"name":user.userapp.name,"title":"Verificación de correo.","type":"email_verification","date_today":datetime.now().strftime('%d/%m/%Y')}
					try:
						send_mail([user],ctx=ctx,template="register-template.html",subject="Verificación de Correo IHES")
						messages.success(request, "email_sent")
					except Exception as e:
						p_.error(f"Error enviando correo de verificacion a {user.email}: {e}")
						messages.success(request, "email_error")
				else:
					messages.success(request, "already_valid")
			else:
				messages.success(request, "not_user_found")
		else:
			messages.success(request, "email_not_provided")
		return render(request, 'valid_email.html',ct)

class PasswordRecover(View):
	def get(self, request, *args, **kwargs):
		if request.user.is_authenticated:
			logout(request)
			return HttpResponseRedirect(reverse_lazy('login'))
		ctx={"send_mail":True}
		if 'code' in kwargs and 'iv' in kwargs:
			code=f"0x{kwargs['code']}"
			iv=f"0x{kwargs['iv']}"
			valid_code=False
			try:
				text=decript(hex(int(code, 16))[2:],hex(int(iv, 16))[2:])
				data=text.decode().split('<->')
				if datetime.now()<(datetime.fromtimestamp(float(data[1]))+timedelta(hours=1)):
					user=Userapp.objects.filter(user_id=data[0],record_create=datetime.fromtimestamp(float(data[2]))).last()
					if user:
						valid_code=True
						ctx["email"]=user.user.email
						ctx["name"]=user.name
						if not user.validated_email:
							user.validated_email=True
							user.save()
				else:
					ctx["expired"]=True
			except:
				valid_code=False
			# ctx["form"]=PasswordChangeForm()
			ctx["send_mail"]=False
			ctx["valid_code"]=valid_code
		# ctx["resend_verification"]=False
		# if 'email' in kwargs:
		# 	ctx["send_mail"]=False
			
		return render(request, 'password_recover.html',ctx)
	def post(self, request, *args, **kwargs):
		ct={"send_mail":True}
		email=request.POST.get('email','')
		print(request.POST)
		if email!="":
			user=User.objects.filter(username=email).last()
			if user:
				verification_code,iv=encript(f"{user.id}<->{datetime.now().timestamp()}<->{user.userapp.record_create.timestamp()}")
				ctx={"v_code":verification_code,"token":iv,"username":user.username,"phone":user.userapp.phone,"email":user.email,"name":user.userapp.name,"title":"Verificación de correo.","type":"password_recover","date_today":datetime.now().strftime('%d/%m/%Y')}
				try:
					# print(verification_code)
					# print(iv)
					send_mail([user],ctx=ctx,template="register-template.html",subject="Recuperación de contraseña IHES")
					messages.success(request, "email_sent")
				except Exception as e:
					p_.error(f"Error enviando correo de recuperacion a {user.email}: {e}")
					messages.success(request, "email_error")
			else:
				messages.success(request, "email_error")
				# messages.success(request, "not_user_found")
		elif "new_password1" in request.POST and "new_password2" in request.POST:
			valid_code=False
			user=None
			if 'code' in kwargs and 'iv' in kwargs:
				code=f"0x{kwargs['code']}"
				iv=f"0x{kwargs['iv']}"
				try:
					text=decript(hex(int(code, 16))[2:],hex(int(iv, 16))[2:])
					data=text.decode().split('<->')
					if datetime.now()<(datetime.fromtimestamp(float(data[1]))+timedelta(hours=1)):
						userapp=Userapp.objects.filter(user_id=data[0],record_create=datetime.fromtimestamp(float(data[2]))).last()
						if userapp:
							valid_code=True
							user=userapp.user
				except:
					valid_code=False
			if valid_code and user:
				form =SetPasswordForm(user,request.POST)
				if form.is_valid():
					f=form.save()
					messages.success(request, "password_recovered")
				else:
					print(form.errors)
					messages.success(request, "form_error")
			else:
				messages.success(request, "error_changing_pass")
			ct["send_mail"]=False
			ct["valid_code"]=valid_code
		else:
			messages.success(request, "email_not_provided")
		return render(request, 'password_recover.html',ct)

class WorkplaceFormView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		ctx={"workplaces":[{"id":item.id,"name":item.name,"employee_count":item.employee_num
		} for item in Workplace.objects.filter(user_id=self.request.user.id)]}
		ctx['workplaces_available']=request.user.userapp.workplaces_available

		return render(request, 'workplaceform.html',ctx)
class GenerarPoliticaView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		workplace_id = kwargs.get('workplace_id')
		workplace = Workplace.objects.filter(id=workplace_id, user_id=request.user.id).first()
		if not workplace:
			return HttpResponseRedirect(reverse_lazy('workplaces'))
		portafolio = PortafolioEvidencias.objects.filter(workplace=workplace).first()
		if request.GET.get('print') == '1' and portafolio:
			ctx = {'workplace': workplace, 'portafolio': portafolio, 'print_mode': True}
			return render(request, 'pdf/politica_prevencion.html', ctx)
		form = PoliticaPrevencionForm(initial={
			'responsable_nombre': portafolio.responsable_nombre if portafolio else '',
			'responsable_puesto': portafolio.responsable_puesto if portafolio else '',
			'representante_legal_nombre': portafolio.representante_legal_nombre if portafolio else '',
			'representante_legal_cargo': portafolio.representante_legal_cargo if portafolio else '',
			'canal_quejas': portafolio.canal_quejas if portafolio else '',
			'responsable_quejas': portafolio.responsable_quejas if portafolio else '',
			'correo_quejas': portafolio.correo_quejas if portafolio else '',
			'tiempo_respuesta_quejas': portafolio.tiempo_respuesta_quejas if portafolio else '',
			'periodicidad_revision': portafolio.periodicidad_revision if portafolio else 'Anual',
		})
		politica_existe = bool(portafolio and portafolio.responsable_nombre)
		if politica_existe:
			for campo in form.fields.values():
				campo.widget.attrs['disabled'] = 'disabled'
		ctx = {'workplace': workplace, 'form': form, 'portafolio': portafolio, 'politica_existe': politica_existe}
		return render(request, 'politica_prevencion_form.html', ctx)
	def post(self, request, *args, **kwargs):
		workplace_id = kwargs.get('workplace_id')
		workplace = Workplace.objects.filter(id=workplace_id, user_id=request.user.id).first()
		if not workplace:
			return HttpResponseRedirect(reverse_lazy('workplaces'))
		form = PoliticaPrevencionForm(request.POST)
		if form.is_valid():
			portafolio, created = PortafolioEvidencias.objects.get_or_create(workplace=workplace, defaults={'periodo_evaluacion': str(timezone.now().year)})
			for field in form.cleaned_data:
				setattr(portafolio, field, form.cleaned_data[field])
			if not created:
				try:
					nueva_version = round(float(portafolio.version_politica) + 0.1, 1)
				except (TypeError, ValueError):
					nueva_version = 1.1
				portafolio.version_politica = f'{nueva_version:.1f}'
				portafolio.fecha_emision = timezone.now().date()
				portafolio.fecha_proxima_revision = portafolio.fecha_emision + timedelta(days=365)
			portafolio.save()
			return HttpResponseRedirect(reverse_lazy('generar_politica', kwargs={'workplace_id': workplace.id}) + '?print=1')
		ctx = {'workplace': workplace, 'form': form}
		return render(request, 'politica_prevencion_form.html', ctx)
class GenerarInformeResultadosView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		workplace_id = kwargs.get('workplace_id')
		workplace = Workplace.objects.filter(id=workplace_id, user_id=request.user.id).first()
		if not workplace:
			return HttpResponseRedirect(reverse_lazy('workplaces'))
		evaluation_param = request.GET.get('evaluation')
		if evaluation_param and evaluation_param.isdigit() and 1 <= int(evaluation_param) <= workplace.evaluation:
			eval_to_check = int(evaluation_param)
		else:
			eval_to_check = workplace.last_completed_evaluation()
		from django.test import RequestFactory
		factory = RequestFactory()
		fake_request = factory.get('/get_chart_data/', {'workplace_id': str(workplace_id), 'evaluation': str(eval_to_check)})
		fake_request.user = request.user
		chart_response = get_chart_data(fake_request)
		chart_data = json.loads(chart_response.content)
		dimensiones_resultado = []
		if chart_data.get('status') == 'ok':
			dim_names = chart_data['dimensions']
			total_dim = chart_data['total_dim']
			niveles_texto = ['Nulo', 'Bajo', 'Medio', 'Alto', 'Muy alto']
			conteos_por_dim = {}
			for item in total_dim:
				idx, nivel, val = item['value']
				conteos_por_dim.setdefault(idx, {})[nivel] = val
			for idx, nombre in enumerate(dim_names):
				conteos = conteos_por_dim.get(idx, {})
				if conteos:
					nivel_predominante = max(conteos, key=conteos.get)
				else:
					nivel_predominante = 0
				dimensiones_resultado.append({
					'nombre': nombre.replace(chr(10), ' '),
					'nivel': niveles_texto[nivel_predominante],
				})
		portafolio = PortafolioEvidencias.objects.filter(workplace=workplace).first()
		ctx = {
			'workplace': workplace,
			'portafolio': portafolio,
			'dimensiones_resultado': dimensiones_resultado,
			'print_mode': True,
		}
		return render(request, 'pdf/informe_resultados.html', ctx)
class CuestionariosAplicadosView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		workplace_id = kwargs.get('workplace_id')
		workplace = Workplace.objects.filter(id=workplace_id, user_id=request.user.id).first()
		if not workplace:
			return HttpResponseRedirect(reverse_lazy('workplaces'))
		evaluation_param = request.GET.get('evaluation')
		if evaluation_param and evaluation_param.isdigit() and 1 <= int(evaluation_param) <= workplace.evaluation:
			evaluation = int(evaluation_param)
		else:
			evaluation = workplace.last_completed_evaluation()
		empleados = Employee.objects.filter(workplace=workplace)
		convocados = empleados.count()
		survey_type = workplace.survey_type()
		fechas = []
		respondieron = 0
		for emp in empleados:
			survey = None
			if survey_type == 3:
				survey = emp.surveyB.filter(evaluation=evaluation).last()
			elif survey_type in (1, 2):
				survey = emp.surveyA.filter(evaluation=evaluation).last()
			if survey:
				respondieron += 1
				fechas.append(survey.record_create)
		fecha_inicio = min(fechas) if fechas else None
		fecha_cierre = max(fechas) if fechas else None
		porcentaje = round((respondieron / convocados) * 100) if convocados else 0
		guia = 'III' if survey_type == 3 else 'II'
		portafolio = PortafolioEvidencias.objects.filter(workplace=workplace).first()
		ctx = {
			'workplace': workplace,
			'portafolio': portafolio,
			'evaluation': evaluation,
			'convocados': convocados,
			'respondieron': respondieron,
			'porcentaje': porcentaje,
			'fecha_inicio': fecha_inicio,
			'fecha_cierre': fecha_cierre,
			'guia': guia,
			'print_mode': True,
		}
		return render(request, 'pdf/cuestionarios_aplicados.html', ctx)
class SubirEvidenciaFaseCView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	TIPOS_PERMITE_NO_APLICA = ('canalizacion',)
	def get(self, request, *args, **kwargs):
		workplace_id = kwargs.get('workplace_id')
		tipo = kwargs.get('tipo')
		workplace = Workplace.objects.filter(id=workplace_id, user_id=request.user.id).first()
		if not workplace:
			return HttpResponseRedirect(reverse_lazy('workplaces'))
		permite_no_aplica = tipo in self.TIPOS_PERMITE_NO_APLICA
		instancia = EvidenciaFaseC.objects.filter(workplace=workplace, tipo=tipo).first()
		initial = {'estado': instancia.estado, 'notas': instancia.notas} if instancia else {}
		form = EvidenciaEstadoForm(initial=initial, permite_no_aplica=permite_no_aplica)
		tipo_display = dict(EvidenciaFaseC.TIPO_CHOICES).get(tipo, tipo)
		ctx = {'workplace': workplace, 'form': form, 'instancia': instancia, 'tipo': tipo, 'tipo_display': tipo_display}
		return render(request, 'evidencia_fase_c_form.html', ctx)
	def post(self, request, *args, **kwargs):
		workplace_id = kwargs.get('workplace_id')
		tipo = kwargs.get('tipo')
		workplace = Workplace.objects.filter(id=workplace_id, user_id=request.user.id).first()
		if not workplace:
			return HttpResponseRedirect(reverse_lazy('workplaces'))
		permite_no_aplica = tipo in self.TIPOS_PERMITE_NO_APLICA
		form = EvidenciaEstadoForm(request.POST, permite_no_aplica=permite_no_aplica)
		if form.is_valid():
			EvidenciaFaseC.objects.update_or_create(
				workplace=workplace, tipo=tipo,
				defaults={'estado': form.cleaned_data['estado'], 'notas': form.cleaned_data['notas']},
			)
			return HttpResponseRedirect(reverse_lazy('subir_evidencia_fase_c', kwargs={'workplace_id': workplace.id, 'tipo': tipo}))
		instancia = EvidenciaFaseC.objects.filter(workplace=workplace, tipo=tipo).first()
		tipo_display = dict(EvidenciaFaseC.TIPO_CHOICES).get(tipo, tipo)
		ctx = {'workplace': workplace, 'form': form, 'instancia': instancia, 'tipo': tipo, 'tipo_display': tipo_display}
		return render(request, 'evidencia_fase_c_form.html', ctx)

@login_required
def guardar_estado_evidencia(request, workplace_id, tipo):
	if request.method != 'POST':
		return JsonResponse({'error': 'method_not_allowed'}, status=405)
	workplace = Workplace.objects.filter(id=workplace_id, user_id=request.user.id).first()
	if not workplace:
		return JsonResponse({'error': 'not_found'}, status=404)
	estado = request.POST.get('estado')
	permite_no_aplica = tipo in SubirEvidenciaFaseCView.TIPOS_PERMITE_NO_APLICA
	choices_validas = [c[0] for c in EvidenciaFaseC.ESTADO_CHOICES if permite_no_aplica or c[0] != 'no_aplica']
	if estado not in choices_validas:
		return JsonResponse({'error': 'estado_invalido'}, status=400)
	if tipo not in dict(EvidenciaFaseC.TIPO_CHOICES):
		return JsonResponse({'error': 'tipo_invalido'}, status=400)
	EvidenciaFaseC.objects.update_or_create(
		workplace=workplace, tipo=tipo,
		defaults={'estado': estado},
	)
	return JsonResponse({'ok': True})
def get_portafolio_status(request):
	workplace_id = request.GET.get('workplace_id')
	workplace = Workplace.objects.filter(id=workplace_id, user_id=request.user.id).first()
	if not workplace:
		return JsonResponse({'items': []})
	portafolio = PortafolioEvidencias.objects.filter(workplace=workplace).first()
	evaluation_param = request.GET.get('evaluation')
	if evaluation_param and evaluation_param.isdigit() and 1 <= int(evaluation_param) <= workplace.evaluation:
		eval_to_check = int(evaluation_param)
	else:
		eval_to_check = workplace.last_completed_evaluation()
	historial_nums = set(workplace.evaluation_history.values_list('numero_evaluacion', flat=True))
	evaluations = [{'numero': n, 'done': n in historial_nums} for n in range(1, workplace.evaluation + 1)]
	items = []
	# 1. Politica de Prevencion
	politica_completa = bool(portafolio and portafolio.responsable_nombre)
	items.append({
		'nombre': 'Politica de Prevencion',
		'estado': 'completo' if politica_completa else 'pendiente',
		'detalle': 'Version ' + portafolio.version_politica if politica_completa else 'No generada aun — da clic en Generar para crear la primera version',
		'url': '/generar_politica/' + str(workplace.id) + '/',
		'control': 'sistema',
	})
	# 2. Informe de Resultados
	from django.test import RequestFactory
	factory = RequestFactory()
	fake_request = factory.get('/get_chart_data/', {'workplace_id': str(workplace.id), 'evaluation': str(eval_to_check)})
	fake_request.user = request.user
	chart_response = get_chart_data(fake_request)
	chart_data = json.loads(chart_response.content)
	informe_completo = chart_data.get('status') == 'ok'
	items.append({
		'nombre': 'Informe de Resultados',
		'estado': 'completo' if informe_completo else 'pendiente',
		'detalle': 'Datos disponibles' if informe_completo else 'Se necesita al menos 1 cuestionario respondido en esta evaluacion para poder generarlo',
		'url': '/generar_informe_resultados/' + str(workplace.id) + '/?evaluation=' + str(eval_to_check),
		'control': 'sistema',
	})
	# 3. Cuestionarios Aplicados
	evaluation = eval_to_check
	empleados = Employee.objects.filter(workplace=workplace)
	convocados = empleados.count()
	survey_type = workplace.survey_type()
	respondieron = 0
	for emp in empleados:
		survey = None
		if survey_type == 3:
			survey = emp.surveyB.filter(evaluation=evaluation).last()
		elif survey_type in (1, 2):
			survey = emp.surveyA.filter(evaluation=evaluation).last()
		if survey:
			respondieron += 1
	porcentaje = round((respondieron / convocados) * 100) if convocados else 0
	cuestionarios_completo = respondieron > 0
	items.append({
		'nombre': 'Cuestionarios Aplicados',
		'estado': 'completo' if cuestionarios_completo else 'pendiente',
		'detalle': str(respondieron) + ' de ' + str(convocados) + ' (' + str(porcentaje) + '%)',
		'url': '/cuestionarios_aplicados/' + str(workplace.id) + '/?evaluation=' + str(eval_to_check),
		'control': 'sistema',
	})

	def _item_checklist(tipo, nombre, permite_no_aplica=False):
		ev = EvidenciaFaseC.objects.filter(workplace=workplace, tipo=tipo).first()
		estado_valor = ev.estado if ev else 'en_proceso'
		if estado_valor == 'completado':
			estado_badge = 'completo'
		elif estado_valor == 'no_aplica':
			estado_badge = 'no_aplica'
		else:
			estado_badge = 'pendiente'
		return {
			'nombre': nombre,
			'estado': estado_badge,
			'estado_valor': estado_valor,
			'control': 'checklist',
			'tipo': tipo,
			'permite_no_aplica': permite_no_aplica,
			'detalle': ev.get_estado_display() if ev else 'Sin estado registrado',
			'url': '/subir_evidencia_fase_c/' + str(workplace.id) + '/' + tipo + '/',
		}

	# Grupo A: siempre obligatorios, sin "No aplica"
	items.append(_item_checklist('difusion', 'Evidencia de Difusion'))
	items.append(_item_checklist('registros', 'Registros de resultados y medidas de control'))
	items.append(_item_checklist('mecanismos_queja', 'Mecanismos de queja/denuncia de violencia laboral'))

	# Grupo B: condicionado a un evento real, con "No aplica"
	items.append(_item_checklist('canalizacion', 'Canalizaciones Guia I', permite_no_aplica=True))

	# Automaticos: solo aparecen si el diagnostico muestra nivel Medio/Alto/Muy alto en alguna dimension
	requiere_intervencion = False
	if chart_data.get('status') == 'ok':
		for item in chart_data.get('total_dim', []):
			idx, nivel, val = item['value']
			if nivel >= 2 and val > 0:
				requiere_intervencion = True
				break
	if requiere_intervencion:
		items.append(_item_checklist('examen_medico', 'Examenes Medicos/Evaluaciones Psicologicas'))
		items.append(_item_checklist('medida_control', 'Medidas de Control/Programa de Intervencion'))

	items_contables = [i for i in items if i['estado'] != 'no_aplica']
	completos = sum(1 for i in items_contables if i['estado'] == 'completo')
	porcentaje_cumplimiento = round((completos / len(items_contables)) * 100) if items_contables else 0
	return JsonResponse({
		'items': items,
		'porcentaje_cumplimiento': porcentaje_cumplimiento,
		'completos': completos,
		'total': len(items_contables),
		'evaluations': evaluations,
		'viewing_evaluation': eval_to_check,
		'workplace_name': workplace.name,
	})
class TestView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		return render(request, 'pdf/muestra_chart.html')

EMPLEADO_CAMPOS_CARGA_MASIVA = [
	("Nombre completo", "name", None),
	("Sexo", "gender", [("Masculino", 1), ("Femenino", 2)]),
	("Edad", "age", [
		("15–19 años", 0), ("20–24 años", 1), ("25–29 años", 2), ("30–34 años", 3),
		("35–39 años", 4), ("40–44 años", 5), ("45–49 años", 6), ("50–54 años", 7),
		("55–59 años", 8), ("60–64 años", 9), ("65–69 años", 10), ("70 o más años", 11),
	]),
	("Estado civil", "civil_state", [
		("Casado", 0), ("Divorciado", 1), ("Soltero", 2), ("Viudo", 3), ("Unión Libre", 4),
	]),
	("Nivel de estudios", "study_level", [
		("Sin formación", 0), ("Primaria", 1), ("Secundaria", 2),
		("Preparatoria o Bachillerato", 3), ("Técnico superior", 4),
		("Licenciatura", 5), ("Maestría", 6), ("Doctorado", 7),
	]),
	("Ocupación, Profesión o Puesto", "ocupation", None),
	("Departamento, Sección o Área", "department", None),
	("Tipo de puesto", "charge_type", [
		("Operativo", 0), ("Supervisor", 1), ("Profesional o Técnico", 2), ("Gerente", 3),
	]),
	("Tipo de contratación", "contract_type", [
		("Por obra o proyecto", 0), ("Por tiempo determinado (Temporal)", 1),
		("Por tiempo indeterminado", 2), ("Por Honorarios", 3),
	]),
	("Tipo de personal", "employee_type", [
		("Sindicalizado", 0), ("Confianza", 1), ("Ninguno", 2),
	]),
	("Tipo de jornada", "shift_type", [
		("Fijo Nocturno (20:00 – 06:00 hrs)", 0), ("Fijo Diurno (06:00 – 20:00 hrs)", 1),
		("Fijo Mixto (Combinación)", 2),
	]),
	("Rotación de turnos", "shift_rotation", [("Sí", 0), ("No", 1)]),
	("Tiempo en el puesto actual", "time_in_charge", [
		("Menos de 6 meses", 0), ("Entre 6 meses y un año", 1), ("Entre 1 a 4 años", 2),
		("Entre 5 a 9 años", 3), ("Entre 10 a 14 años", 4), ("Entre 15 a 19 años", 5),
		("Entre 20 a 24 años", 6), ("25 años o más", 7),
	]),
	("Tiempo de experiencia laboral", "exp", [
		("Menos de 6 meses", 0), ("Entre 6 meses y un año", 1), ("Entre 1 a 4 años", 2),
		("Entre 5 a 9 años", 3), ("Entre 10 a 14 años", 4), ("Entre 15 a 19 años", 5),
		("Entre 20 a 24 años", 6), ("25 años o más", 7),
	]),
]
CARGA_MASIVA_MAX_FILAS = 500

def download_employee_template(request, workplace_id):
	if not request.user.is_authenticated:
		return HttpResponseRedirect(reverse_lazy('login'))
	if not request.user.workplaces.filter(id=workplace_id).exists():
		return HttpResponseRedirect(reverse_lazy('workplaces'))
	wk = Workplace.objects.filter(id=workplace_id).last()

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Empleados"
	header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)
	example_fill = PatternFill(start_color="FFF9DB", end_color="FFF9DB", fill_type="solid")
	for col_idx, (label, field, opciones) in enumerate(EMPLEADO_CAMPOS_CARGA_MASIVA, start=1):
		cell = ws.cell(row=1, column=col_idx, value=label)
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = Alignment(wrap_text=True, vertical="center")
		ws.column_dimensions[cell.column_letter].width = 26

	ejemplos = [
		["EJEMPLO 1 — borra esta fila antes de subir", "Masculino", "30–34 años", "Casado", "Licenciatura", "Operador", "Producción", "Operativo", "Por tiempo indeterminado", "Sindicalizado", "Fijo Diurno (06:00 – 20:00 hrs)", "No", "Entre 1 a 4 años", "Entre 1 a 4 años"],
		["EJEMPLO 2 — borra esta fila antes de subir", "Femenino", "40–44 años", "Soltero", "Técnico superior", "Supervisora", "Logística", "Supervisor", "Por tiempo indeterminado", "Confianza", "Fijo Nocturno (20:00 – 06:00 hrs)", "Sí", "Entre 5 a 9 años", "Entre 10 a 14 años"],
	]
	for row_idx, fila in enumerate(ejemplos, start=2):
		for col_idx, valor in enumerate(fila, start=1):
			cell = ws.cell(row=row_idx, column=col_idx, value=valor)
			cell.fill = example_fill

	ws_listas = wb.create_sheet("Listas")
	ws_listas.sheet_state = "hidden"
	for col_idx, (label, field, opciones) in enumerate(EMPLEADO_CAMPOS_CARGA_MASIVA, start=1):
		if not opciones:
			continue
		for row_idx, (texto, _codigo) in enumerate(opciones, start=1):
			ws_listas.cell(row=row_idx, column=col_idx, value=texto)
		col_letter = ws_listas.cell(row=1, column=col_idx).column_letter
		rango = f"Listas!${col_letter}$1:${col_letter}${len(opciones)}"
		dv = DataValidation(type="list", formula1=f"={rango}", allow_blank=False, showErrorMessage=True)
		dv.error = "Selecciona una opción de la lista."
		dv.errorTitle = "Valor no válido"
		main_col_letter = ws.cell(row=1, column=col_idx).column_letter
		dv.add(f"{main_col_letter}4:{main_col_letter}{CARGA_MASIVA_MAX_FILAS + 3}")
		ws.add_data_validation(dv)
	ws.freeze_panes = "A2"

	buffer = BytesIO()
	wb.save(buffer)
	buffer.seek(0)
	response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	nombre_archivo = f"plantilla_empleados_{wk.name}.xlsx".replace(" ", "_")
	response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
	return response

def upload_employees_bulk(request):
	if request.method != 'POST':
		return JsonResponse({'status': 'error', 'error': 'Método no permitido'}, status=405)
	if not request.user.is_authenticated:
		return JsonResponse({'status': 'error', 'error': 'No autenticado'}, status=401)
	workplace_id = request.POST.get('workplace_id')
	if not request.user.workplaces.filter(id=workplace_id).exists():
		return JsonResponse({'status': 'error', 'error': 'Centro de trabajo no encontrado'}, status=403)
	wk = Workplace.objects.filter(id=workplace_id).last()
	archivo = request.FILES.get('file')
	if not archivo:
		return JsonResponse({'status': 'error', 'error': 'No se recibió ningún archivo'}, status=400)
	if not archivo.name.lower().endswith('.xlsx'):
		return JsonResponse({'status': 'error', 'error': 'El archivo debe ser .xlsx (el generado por "Descargar plantilla")'}, status=400)
	if archivo.size > 5 * 1024 * 1024:
		return JsonResponse({'status': 'error', 'error': 'El archivo excede el tamaño máximo permitido (5 MB)'}, status=400)
	try:
		wb = openpyxl.load_workbook(archivo, data_only=True)
		ws = wb["Empleados"] if "Empleados" in wb.sheetnames else wb.active
	except Exception:
		return JsonResponse({'status': 'error', 'error': 'No se pudo leer el archivo. Verifica que sea un .xlsx válido y no esté dañado.'}, status=400)
	encabezados_esperados = [campo[0] for campo in EMPLEADO_CAMPOS_CARGA_MASIVA]
	encabezados_archivo = [cell.value for cell in ws[1]][:len(encabezados_esperados)]
	if encabezados_archivo != encabezados_esperados:
		return JsonResponse({'status': 'error', 'error': 'Los encabezados del archivo no coinciden con la plantilla. Descarga la plantilla actual y no modifiques los nombres de columna.'}, status=400)
	filas_datos = list(ws.iter_rows(min_row=2, values_only=False))
	if len(filas_datos) > CARGA_MASIVA_MAX_FILAS:
		return JsonResponse({'status': 'error', 'error': f'El archivo tiene más de {CARGA_MASIVA_MAX_FILAS} filas. Divide la carga en varios archivos.'}, status=400)
	mapas_por_columna = []
	for label, field, opciones in EMPLEADO_CAMPOS_CARGA_MASIVA:
		mapas_por_columna.append({texto.strip().lower(): codigo for texto, codigo in opciones} if opciones else None)
	creados = 0
	errores = []
	cupo_restante = wk.employee_num - wk.employees.count()
	for fila in filas_datos:
		numero_fila_excel = fila[0].row
		valores = [c.value for c in fila]
		if all((v is None or str(v).strip() == '') for v in valores):
			continue
		nombre = str(valores[0]).strip() if valores[0] is not None else ''
		if nombre.upper().startswith('EJEMPLO '):
			continue
		datos_limpios = {}
		error_fila = None
		for idx, (label, field, opciones) in enumerate(EMPLEADO_CAMPOS_CARGA_MASIVA):
			valor_crudo = valores[idx]
			valor_txt = str(valor_crudo).strip() if valor_crudo is not None else ''
			if not valor_txt:
				error_fila = f'"{label}" no puede quedar vacío.'
				break
			if field == 'name':
				if len(valor_txt) < 6:
					error_fila = '"Nombre completo" debe tener al menos 6 caracteres.'
					break
				datos_limpios['name'] = valor_txt
			elif opciones is None:
				datos_limpios[field] = valor_txt
			else:
				codigo = mapas_por_columna[idx].get(valor_txt.lower())
				if codigo is None:
					opciones_validas = ', '.join(texto for texto, _c in opciones)
					error_fila = f'"{label}" — "{valor_txt}" no coincide con ninguna opción. Usa una del menú desplegable: {opciones_validas}.'
					break
				datos_limpios[field] = codigo
		if error_fila:
			errores.append({'fila': numero_fila_excel, 'empleado': nombre or '(fila sin nombre)', 'error': error_fila})
			continue
		if cupo_restante <= 0:
			errores.append({'fila': numero_fila_excel, 'empleado': nombre, 'error': f'Límite del centro — ya se alcanzó el máximo de {wk.employee_num} empleados para este centro de trabajo.'})
			continue
		Employee.objects.create(workplace=wk, **datos_limpios)
		creados += 1
		cupo_restante -= 1
	return JsonResponse({'status': 'ok', 'creados': creados, 'errores': errores})

PLAN_ACCION_CAMPOS = [
	("Área o trabajadores sujetos", "area_trabajadores", None),
	("Tipo de acción", "tipo_accion", None),
	("Fecha programada", "fecha_programada", "fecha"),
	("Responsable", "responsable", None),
	("Estado", "estado", [("Pendiente", "pendiente"), ("En proceso", "en_proceso"), ("Completado", "completado")]),
	("Evaluación posterior", "evaluacion_posterior", None),
]
PLAN_ACCION_MAX_FILAS = 200

def download_plan_accion_template(request, workplace_id):
	if not request.user.is_authenticated:
		return HttpResponseRedirect(reverse_lazy('login'))
	if not request.user.workplaces.filter(id=workplace_id).exists():
		return HttpResponseRedirect(reverse_lazy('workplaces'))
	wk = Workplace.objects.filter(id=workplace_id).last()
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "PlanAccion"
	header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)
	example_fill = PatternFill(start_color="FFF9DB", end_color="FFF9DB", fill_type="solid")
	for col_idx, (label, field, opciones) in enumerate(PLAN_ACCION_CAMPOS, start=1):
		cell = ws.cell(row=1, column=col_idx, value=label)
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = Alignment(wrap_text=True, vertical="center")
		ws.column_dimensions[cell.column_letter].width = 30
	ejemplos = [
		["EJEMPLO 1 — borra esta fila antes de subir", "Rediseño de cargas de trabajo en Producción", "2026-09-15", "Ing. Rosa Delgado", "En proceso", "Reaplicación de cuestionario en 6 meses"],
		["EJEMPLO 2 — borra esta fila antes de subir", "Campaña de sensibilización contra la violencia laboral", "2026-10-01", "Lic. Marco Reyes", "Pendiente", "Encuesta de percepción post-campaña"],
	]
	for row_idx, fila in enumerate(ejemplos, start=2):
		for col_idx, valor in enumerate(fila, start=1):
			cell = ws.cell(row=row_idx, column=col_idx, value=valor)
			cell.fill = example_fill
	ws_listas = wb.create_sheet("Listas")
	ws_listas.sheet_state = "hidden"
	for col_idx, (label, field, opciones) in enumerate(PLAN_ACCION_CAMPOS, start=1):
		if not opciones or opciones == "fecha":
			continue
		for row_idx, (texto, _codigo) in enumerate(opciones, start=1):
			ws_listas.cell(row=row_idx, column=col_idx, value=texto)
		col_letter = ws_listas.cell(row=1, column=col_idx).column_letter
		rango = f"Listas!${col_letter}$1:${col_letter}${len(opciones)}"
		dv = DataValidation(type="list", formula1=f"={rango}", allow_blank=False, showErrorMessage=True)
		dv.error = "Selecciona una opción de la lista."
		dv.errorTitle = "Valor no válido"
		main_col_letter = ws.cell(row=1, column=col_idx).column_letter
		dv.add(f"{main_col_letter}4:{main_col_letter}{PLAN_ACCION_MAX_FILAS + 3}")
		ws.add_data_validation(dv)
	# Nota: no se pre-formatean celdas vacias de la columna fecha (a diferencia de un
	# intento anterior) porque tocar Cell objects vacios con ws.cell(row=r, column=c)
	# extiende ws.max_row/dimensions aunque no tengan valor, haciendo que iter_rows()
	# en upload_plan_accion_bulk cuente cientos de filas fantasma como si tuvieran datos.
	ws.freeze_panes = "A2"
	buffer = BytesIO()
	wb.save(buffer)
	buffer.seek(0)
	response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	nombre_archivo = f"plan_accion_{wk.name}.xlsx".replace(" ", "_")
	response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
	return response

def upload_plan_accion_bulk(request):
	if request.method != 'POST':
		return JsonResponse({'status': 'error', 'error': 'Método no permitido'}, status=405)
	if not request.user.is_authenticated:
		return JsonResponse({'status': 'error', 'error': 'No autenticado'}, status=401)
	workplace_id = request.POST.get('workplace_id')
	if not request.user.workplaces.filter(id=workplace_id).exists():
		return JsonResponse({'status': 'error', 'error': 'Centro de trabajo no encontrado'}, status=403)
	wk = Workplace.objects.filter(id=workplace_id).last()
	archivo = request.FILES.get('file')
	if not archivo:
		return JsonResponse({'status': 'error', 'error': 'No se recibió ningún archivo'}, status=400)
	if not archivo.name.lower().endswith('.xlsx'):
		return JsonResponse({'status': 'error', 'error': 'El archivo debe ser .xlsx (el generado por "Descargar plantilla")'}, status=400)
	if archivo.size > 5 * 1024 * 1024:
		return JsonResponse({'status': 'error', 'error': 'El archivo excede el tamaño máximo permitido (5 MB)'}, status=400)
	try:
		wb = openpyxl.load_workbook(archivo, data_only=True)
		ws = wb["PlanAccion"] if "PlanAccion" in wb.sheetnames else wb.active
	except Exception:
		return JsonResponse({'status': 'error', 'error': 'No se pudo leer el archivo. Verifica que sea un .xlsx válido y no esté dañado.'}, status=400)
	encabezados_esperados = [campo[0] for campo in PLAN_ACCION_CAMPOS]
	encabezados_archivo = [cell.value for cell in ws[1]][:len(encabezados_esperados)]
	if encabezados_archivo != encabezados_esperados:
		return JsonResponse({'status': 'error', 'error': 'Los encabezados del archivo no coinciden con la plantilla. Descarga la plantilla actual y no modifiques los nombres de columna.'}, status=400)
	filas_crudas = list(ws.iter_rows(min_row=2, values_only=False))
	filas_datos = []
	for fila in filas_crudas:
		valores = [c.value for c in fila]
		if all((v is None or str(v).strip() == '') for v in valores):
			continue
		primer_valor = str(valores[0]).strip() if valores[0] is not None else ''
		if primer_valor.upper().startswith('EJEMPLO '):
			continue
		filas_datos.append(fila)
	if len(filas_datos) > PLAN_ACCION_MAX_FILAS:
		return JsonResponse({'status': 'error', 'error': f'El archivo tiene más de {PLAN_ACCION_MAX_FILAS} filas. Divide la carga en varios archivos.'}, status=400)
	mapas_por_columna = []
	for label, field, opciones in PLAN_ACCION_CAMPOS:
		mapas_por_columna.append({texto.strip().lower(): codigo for texto, codigo in opciones} if isinstance(opciones, list) else None)
	creados = 0
	errores = []
	for fila in filas_datos:
		numero_fila_excel = fila[0].row
		valores = [c.value for c in fila]
		primer_valor = str(valores[0]).strip() if valores[0] is not None else ''
		datos_limpios = {}
		error_fila = None
		for idx, (label, field, opciones) in enumerate(PLAN_ACCION_CAMPOS):
			valor_crudo = valores[idx]
			if field == 'fecha_programada':
				if valor_crudo is None or (isinstance(valor_crudo, str) and not valor_crudo.strip()):
					error_fila = f'"{label}" no puede quedar vacío.'
					break
				if hasattr(valor_crudo, 'year'):
					datos_limpios[field] = valor_crudo.date() if hasattr(valor_crudo, 'date') and callable(valor_crudo.date) else valor_crudo
				else:
					try:
						from datetime import datetime as _dt
						datos_limpios[field] = _dt.strptime(str(valor_crudo).strip(), '%Y-%m-%d').date()
					except ValueError:
						error_fila = f'"{label}" — "{valor_crudo}" no es una fecha válida. Usa el formato AAAA-MM-DD (ej. 2026-09-15).'
						break
				continue
			valor_txt = str(valor_crudo).strip() if valor_crudo is not None else ''
			if not valor_txt:
				error_fila = f'"{label}" no puede quedar vacío.'
				break
			if opciones is None:
				datos_limpios[field] = valor_txt
			else:
				codigo = mapas_por_columna[idx].get(valor_txt.lower())
				if codigo is None:
					opciones_validas = ', '.join(texto for texto, _c in opciones)
					error_fila = f'"{label}" — "{valor_txt}" no coincide con ninguna opción. Usa una del menú desplegable: {opciones_validas}.'
					break
				datos_limpios[field] = codigo
		if error_fila:
			errores.append({'fila': numero_fila_excel, 'accion': datos_limpios.get('tipo_accion', primer_valor) or '(fila sin tipo de acción)', 'error': error_fila})
			continue
		PlanAccionItem.objects.create(workplace=wk, **datos_limpios)
		creados += 1
	return JsonResponse({'status': 'ok', 'creados': creados, 'errores': errores})

def get_plan_accion(request):
	workplace_id = request.GET.get('workplace_id')
	if not request.user.workplaces.filter(id=workplace_id).exists():
		return JsonResponse({'items': []})
	items = PlanAccionItem.objects.filter(workplace_id=workplace_id).order_by('fecha_programada')
	data = [{'id': item.id, 'area_trabajadores': item.area_trabajadores, 'tipo_accion': item.tipo_accion, 'fecha_programada': item.fecha_programada.strftime('%d/%m/%Y'), 'responsable': item.responsable, 'evaluacion_posterior': item.evaluacion_posterior, 'estado': item.estado, 'estado_display': item.get_estado_display()} for item in items]
	return JsonResponse({'items': data})

def guardar_estado_accion(request, accion_id):
	if request.method != 'POST':
		return JsonResponse({'error': 'method_not_allowed'}, status=405)
	item = PlanAccionItem.objects.filter(id=accion_id, workplace__user=request.user).first()
	if not item:
		return JsonResponse({'error': 'not_found'}, status=404)
	estado = request.POST.get('estado')
	if estado not in dict(PlanAccionItem.ESTADO_CHOICES):
		return JsonResponse({'error': 'estado_invalido'}, status=400)
	item.estado = estado
	item.save()
	return JsonResponse({'ok': True})

class EmployeeFormView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		ctx={"workplaces":[{"id":item.id,"name":item.name,"employee_count":item.employee_num
		} for item in Workplace.objects.filter(user_id=self.request.user.id)]}
		if 'workplace_id' in kwargs:
			ctx['workplace_id']=kwargs['workplace_id']
			if not request.user.workplaces.filter(id=kwargs['workplace_id']).exists():
				return HttpResponseRedirect(reverse_lazy('workplaces'))
			wk = Workplace.objects.filter(id=kwargs['workplace_id']).last()
			ctx['employee_num'] = wk.employee_num
			ctx['employee_current_count'] = wk.employees.count()
		else:
			return HttpResponseRedirect(reverse_lazy('workplaces'))
		return render(request, 'employeeform.html',ctx)
class EvidenceView(LoginRequiredMixin,View):
	login_url = reverse_lazy('login')
	redirect_field_name = 'redirect_to'
	def get(self, request, *args, **kwargs):
		ctx={"workplaces":[{"id":item.id,"name":item.name,"employee_count":item.employee_num
		} for item in Workplace.objects.filter(user_id=self.request.user.id)]}
		users=[{"id":user.user.id,"text":user.name} for user in Userapp.objects.all()]#.filter(user__is_staff=False)]
		ctx['users']=users
		preselected_workplace_id = request.GET.get('workplace_id')
		if preselected_workplace_id and Workplace.objects.filter(id=preselected_workplace_id, user_id=request.user.id).exists():
			ctx['preselected_workplace_id'] = int(preselected_workplace_id)
		else:
			ctx['preselected_workplace_id'] = None
		return render(request, 'evidence.html',ctx)

def validate_code(code):
	workplace=Workplace.objects.filter(access_code=code).last()
	ctx=[{'employee_id':'','form_name':'Código Inválido',
		'help':"No se reconoce el código proporcionado revisa que el enlace esté completo, si el código es idéntico al que proporciona el sistema, busca asistencia con nosotros.","employee_name":'',"next_form":999,
		'workplaceid':"",'name':"-",'address':"-",'social_reason':"-"}]
	if workplace:
		help_txt=f'Para {workplace.user.userapp.name} es importante conocer la opinión de todos los colaboradores.\n\
		Por eso, le invitamos a contestar la siguiente encuesta:\n\
		La información proporcionada en este cuestionario y sus resultados\nserá exclusivamente para fines de mejora del\
		ambiente de trabajo.'
		if workplace.employee_num>workplace.employees.all().count():
			ctx=[{'employee_id':'','form_name':'Formulario de datos del trabajador',
			'help':help_txt,"employee_name":'',"next_form":-1,
			'workplaceid':workplace.id,'name':workplace.name,'address':workplace.address,'social_reason':workplace.user.userapp.name}]
			return ctx
		else:
			ctx=[{'employee_id':'','form_name':'No quedan usuarios por registrar',
			'help':"No puedes registrar más usuarios, has llegado al límite de empleados registrados.","employee_name":'',"next_form":999,
			'workplaceid':workplace.id,'name':workplace.name,'address':workplace.address,'social_reason':workplace.user.userapp.name}]
			return ctx
	else:
		print(code)
		spl=code.split('0|0')
		emp=Employee.objects.filter(id=spl[0]).last()
		if emp:
			try:
				ts=int(spl[1],16)	
			except:
				
				return ctx
			if int(emp.record_create.timestamp())==ts:
				workplace=emp.workplace
				if not emp.survey3.filter(Q(evaluation=workplace.evaluation)):
					form_name='Acontecimientos traumáticos severos'
					help_txt=f'Para {workplace.user.userapp.name} es importante conocer la opinión de todos los colaboradores.\n\
						Por eso, le invitamos a contestar la siguiente encuesta:\
						La información proporcionada en este cuestionario y sus \nresultados será exclusivamente para fines de mejora del\
						ambiente de trabajo.'
					next_form=0
				elif not emp.surveyA.filter(Q(evaluation=workplace.evaluation)) and (workplace.survey_type()==1 or workplace.survey_type()==2):
					form_name='Instrumento para identificar los factores de riesgo psicosocial en los centros de trabajo'
					help_txt="Conteste el cuestionario completamente; no existen respuestas correctas o incorrectas; \
						es necesaria su concentración;  considere para contestar las preguntas las condiciones de los dos últimos\
						 meses, y que su opinión es lo más importante por lo que se le pide que conteste con sinceridad."
					next_form=1
				elif not emp.surveyB.filter(Q(evaluation=workplace.evaluation)) and workplace.survey_type()==3:
					form_name='Instrumento para identificar los factores de riesgo psicosocial en los centros de trabajo'
					help_txt="Conteste el cuestionario completamente; no existen respuestas correctas o incorrectas; \
						es necesaria su concentración;  considere para contestar las preguntas las condiciones de los dos últimos\
						 meses, y que su opinión es lo más importante por lo que se le pide que conteste con sinceridad."
					next_form=2
				else:
					next_form=3
					form_name='Ya has completado todos los formularios'
					help_txt='Gracias por su participación, puede cerrar esta pestaña'
				return [{'employee_id':f"{emp.id}",'employee_name':emp.name,'form_name':form_name,"next_form":next_form,'help':help_txt,'workplaceid':workplace.id,'name':workplace.name,'address':workplace.address,'social_reason':workplace.user.userapp.name}]
	return ctx

class SurveyView(View):
	def get(self, request, *args, **kwargs):
		if 'd' in request.GET:
			ctx=validate_code(request.GET['d'])[0]
			return render(request, 'survey.html',ctx)
		return render(request, 'survey.html')

@login_required
def employees_dt(request,workplace_id,t,evaluation):
	if not Workplace.objects.filter(id=workplace_id, user=request.user).exists():
		return JsonResponse({'draw':0,'recordsTotal':0,'recordsFiltered':0,'data':[]}, status=403)
	try:
		draw = int(request.GET.get('draw'))
		start = int(request.GET.get('start'))
		length = int(request.GET.get('length'))
		order = int(request.GET.get('order[0][column]'))
	except (TypeError, ValueError):
		return JsonResponse({'draw':0,'recordsTotal':0,'recordsFiltered':0,'data':[]}, status=400)
	if length <= 0:
		return JsonResponse({'draw':0,'recordsTotal':0,'recordsFiltered':0,'data':[]}, status=400)
	search = request.GET.get('search[value]')
	order_direction=request.GET.get('order[0][dir]')
	query= Employee.objects.filter(workplace_id=workplace_id).order_by('-record_create')
	if not query:
		return JsonResponse({'draw':draw,'recordsTotal':0,'recordsFiltered':0,'data':[],})
	if evaluation==0:
		evaluation=query.last().workplace.evaluation
	records_total = query.count()
	records_filtered = records_total
	ordering={0:"name",1:"gender",2:"ocupation",3:"department"}
	if order not in ordering:
		order=0
	order_direction='' if order_direction=='asc' else '-'
	if search:
		query=query.annotate(search=SearchVector('name','department')).filter(search__icontains=search)
		records_total = query.count()
		records_filtered = records_total
	if not query.exists():
		return JsonResponse({'draw':draw,'recordsTotal':records_total,'recordsFiltered':records_filtered,'data':[],})
	page=(start//length)+1
	paginator = Paginator(query.order_by("{}{}".format(order_direction, ordering[order])), length)
	try:
		workplaces = paginator.page(page).object_list
	except PageNotAnInteger:
		workplaces = paginator.page(page).object_list
	except EmptyPage:
		workplaces = paginator.page(paginator.num_pages).object_list
	if t==0:
		arr=[{
			"name":item.name,
			"gender":item.get_gender_display(),
			"ocupation":item.ocupation,
			"department":item.department,
			"status":f"<div class='badge badge-pill badge-glow badge-{item.get_status(evaluation=evaluation)}</div>",
			"code":f"<a class='emp-wa-btn' data-toggle='tooltip' data-original-title='Compartir enlace por WhatsApp para responder encuestas'\
			 href='https://api.whatsapp.com/send?text=Hola *{item.name}*, para *{item.workplace.name}* es importante conocer la opinión de todos los colaboradores.\
			 Por eso te invitamos a contestar las encuestas para identificar factores de riesgo psicosocial en tu centro de trabajo.%0D%0A %0D%0A\
			*Si tuvieras algún problema o duda* _para responderlo, contáctanos por este medio o marca al número *(33) 1491 1819* en un horario de *9 a 18 horas*._%0D%0A %0D%0A\
			Agradecemos tu apoyo para la mejora de nuestra empresa.%0D%0A %0D%0A\
			*Inicia la encuesta en este enlace*: https://normaia.ihes.mx/app/access/?d={item.get_code()}'\
			  target='_blank' rel='noopener'><svg viewBox='0 0 24 24' fill='currentColor' xmlns='http://www.w3.org/2000/svg'><path d='M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413Z'/></svg>Compartir</a>",
			} for item in workplaces]
	else:
		col={"nulo":"#9be5f7","bajo":"#6bf56e","medio":"#ffff00","alto":"#ffc000","muy_alto":"#ff7070"}
		arr=[]
		total=288
		for item in workplaces:
			val=0
			if item.surveyB.filter(evaluation=evaluation).last() and item.workplace.survey_type()==3:
				_sum=0
				for field in RiskSurveyB._meta.fields:
					if field.name not in ['id','employee','evaluation','record_create','record_update','r3_a','r3_b']:
						_sum=_sum+(getattr(item.surveyB.filter(evaluation=evaluation).last(),field.attname) or 0)
				val=f"{_sum}/288"
				result_class,result_text=("#9be5f7","Riesgo nulo") if _sum<50 else (("#6bf56e","Riesgo bajo") if _sum<75 else (("#ffff00","Riesgo medio") if _sum<99 else (("#ffc000","Riesgo alto") if _sum<140 else ("#ff7070","Riesgo muy alto"))))
			elif item.surveyA.filter(evaluation=evaluation).last() and item.workplace.survey_type()!=3:
				_sum=0
				for field in RiskSurveyA._meta.fields:
					if field.name not in ['id','employee','evaluation','record_create','record_update','r2_p_a','r2_p_b','es_demo']:
						_sum=_sum+(getattr(item.surveyA.filter(evaluation=evaluation).last(),field.attname) or 0)
				val=f"{_sum}/184"
				result_class,result_text=("#9be5f7","Riesgo nulo") if _sum<20 else (("#6bf56e","Riesgo bajo") if _sum<45 else (("#ffff00","Riesgo medio") if _sum<70 else (("#ffc000","Riesgo alto") if _sum<90 else ("#ff7070","Riesgo muy alto"))))
			else:
				val=f""
				result_class,result_text=("##f0f0f0","Sin encuesta")
			traumado=""
			if hasattr(item,"survey3"):
				survey =item.survey3.filter(evaluation=evaluation).last()
				if survey:
					if survey.r1_p1==0:
						traumado=f"<div class='badge badge-pill badge-glow badge-danger ttip' data-toggle='tooltip' data-original-title='Persona identificada con indicios de haber sufrido algun trauma severo'><i class='feather icon-alert-triangle'></i></div> <button class='btn btn-sm btn-outline-secondary get_file' data-temp='muestra_chart' data-fileid='{item.id}' style='font-size:11px;padding:2px 8px;'><i class='feather icon-download'></i> Ver trauma</button>"
			if hasattr(item,"surveyA") or hasattr(item,"surveyB"):
				traumado=f"{traumado} <button class='btn btn-sm btn-outline-primary get_file' data-temp='riesgo_psicosocial' data-fileid='{item.id}' style='font-size:11px;padding:2px 8px;'><i class='feather icon-download'></i> Ver riesgo psicosocial</button>"
			arr.append({
				"name":item.name,
				"status":f"<div class='badge badge-pill badge-glow badge-{item.get_status(evaluation=evaluation)}</div>",
				"department":item.department,
				"trauma":traumado,
				"result":f"<div class='chip mr-1' style='background-color:{result_class}'><div class='chip-body'>\
					<div class='avatar'><i class='feather icon-alert-triangle'></i>\
					</div><span class='chip-text'>{result_text} {val}</span></div></div>",
				})
	return JsonResponse({'draw':draw,'recordsTotal':records_total,'recordsFiltered':records_filtered,'data':arr,})

@login_required
def get_results(request):
	workplace_id=request.GET.get('workplace_id',None)
	if not Workplace.objects.filter(id=workplace_id, user=request.user).exists():
		return JsonResponse({'results':[]}, status=403)
	results=ResultFiles.objects.filter(workplace_id=workplace_id)
	data=[{"evaluation":item.evaluation,
		"result_type":item.get_result_type_display(),
		"date":item.record_create.strftime('%d/%m/%Y %H:%M'),
		"image":reverse('download_result_image', args=[item.workplace_id, item.id]) if item.image else "#"}for item in results]
	return JsonResponse({'results':data})

@login_required
def get_workplaces(request):
	data=[{'id':wk.id,"text":wk.name} for wk in Workplace.objects.filter(user_id=request.user.id)]
	return JsonResponse({'results':data})

@login_required
def get_departments(request):
	workplace_id=request.GET.get('workplace_id',None)
	if not Workplace.objects.filter(id=workplace_id, user=request.user).exists():
		return JsonResponse({'results':[]}, status=403)
	emp_list = Employee.objects.filter(workplace_id=workplace_id).values_list('department', flat=True).distinct()
	data=[{'id':wk,"text":wk} for wk in emp_list]
	return JsonResponse({'results':[{'id':'', "text":''},*data]})

@csrf_exempt
def contact(request):
	cname=request.POST.get('cname','')
	phone=request.POST.get('phone','')
	email=request.POST.get('email','')
	name=request.POST.get('name','')
	message=request.POST.get('message','')
	if email =='' and name =='':
		return JsonResponse({'status':'error'},status=400),
	if phone =='' and cname =='':
		ctx={"cname":cname,"phone":phone,"email":email,"name":name,"message":message,"type":1,"title":"Contacto","date_today":datetime.now().strftime('%d/%m/%Y')}
	else:
		ctx={"cname":cname,"phone":phone,"email":email,"name":name,"message":message,"type":2,"title":"Contacto","date_today":datetime.now().strftime('%d/%m/%Y')}
	send_mail(["normaia.sistemas@gmail.com"],ctx=ctx,subject=f'{name} quiere ponerse en contacto con nosotros')
	return JsonResponse({'status':'ok'})

def add_evidence(request):
	form=ResultFilesForm(request.POST,request.FILES)
	if form.is_valid():
		new=form.save(commit=False)
		new.evaluation=new.workplace.evaluation
		new.save()
		return JsonResponse({'status':"ok"})
	else:
		return JsonResponse({'status':"ok","errors":form.errors}, status=400)
def get_questions(request):
	workplace_id=request.GET.get('workplace_id',None)
	employee_id=request.GET.get('employee_id',None)
	emp=Employee.objects.filter(id=employee_id).last() if employee_id else None
	if emp:
		workplace=emp.workplace
		if not emp.survey3.filter(Q(evaluation=workplace.evaluation)):
			questions=[{'field_name':item.name,'field_desc':item.verbose_name,
				'field_choices':[{'id':choice[0],'title':choice[1]}for choice in item.choices] if item.choices else None,
				'type':1 if item.choices else 0,
				} for item in TraumaSurvey._meta.fields if item.name not in [
				'id','employee','evaluation','record_create','record_update'] ]
			return JsonResponse({'data':{"questions":questions},'form_name':TraumaSurvey._meta.verbose_name,'form':TraumaSurvey._meta.model_name})
		elif not emp.surveyA.filter(Q(evaluation=workplace.evaluation)) and workplace.survey_type()==1 or workplace.survey_type()==2:
			questions=[{'field_name':item.name,'field_desc':item.verbose_name,
				'field_choices':[{'id':choice[0],'title':choice[1]}for choice in item.choices] if item.choices else None,
				'type':1 if item.choices else 0,
				} for item in RiskSurveyA._meta.fields if item.name not in [
				'id','employee','evaluation','record_create','record_update','es_demo'] ]
			return JsonResponse({'data':{"questions":questions},'form_name':RiskSurveyA._meta.verbose_name,'form':RiskSurveyA._meta.model_name})
		elif not emp.surveyB.filter(Q(evaluation=workplace.evaluation)) and workplace.survey_type()==3:
			questions=[{'field_name':item.name,'field_desc':item.verbose_name,
				'field_choices':[{'id':choice[0],'title':choice[1]}for choice in item.choices] if item.choices else None,
				'type':1 if item.choices else 0,
				} for item in RiskSurveyB._meta.fields if item.name not in [
				'id','employee','evaluation','record_create','record_update'] ]
			return JsonResponse({'data':{"questions":questions},'form_name':RiskSurveyB._meta.verbose_name,'form':RiskSurveyB._meta.model_name})
		else:
			return JsonResponse({'data':{"questions":[]},'form_name':"completed",'form':''})

	else:
		workplace=Workplace.objects.filter(id=workplace_id).last()
		if workplace:
			questions=[{'field_name':item.name,'field_desc':item.verbose_name,
				'field_choices':[{'id':choice[0],'title':choice[1]}for choice in item.choices] if item.choices else None,
				'type':1 if item.choices else 0,
				} for item in Employee._meta.fields if item.name not in [
				'id','workplace','record_create','record_update','es_demo'] ]
			if workplace.areas:
				for q in questions:
					if q['field_name']=='department':
						q['field_choices']=[{'id':area,'title':area} for area in workplace.areas]
						q['type']=1

			return JsonResponse({'data':{"questions":questions},'form_name':Employee._meta.verbose_name,'form':Employee._meta.model_name})
	return JsonResponse({'data':{"questions":[]},'form_name':'invalid form','form':''})
@login_required
def get_chart_data(request):
	workplace_id=request.GET.get('workplace_id',None)
	if not Workplace.objects.filter(id=workplace_id, user=request.user).exists():
		return JsonResponse({'error':'not_found'}, status=403)
	dept_id=request.GET.get('dept_id',None)
	evaluation=request.GET.get('evaluation',None)
	if evaluation is None:
		evaluation=Workplace.objects.filter(id=workplace_id).last().evaluation
	employees=Employee.objects.filter(workplace_id=workplace_id).order_by('-record_create')
	if dept_id:
		employees=employees.filter(department=dept_id)
	dimensionA={
		"Condiciones peligrosas e inseguras":["r2_p2"],
		"Condiciones deficientes e insalubres":["r2_p1"],
		"Trabajos peligrosas":["r2_p3"],
		"Cargas cuantitativas":["r2_p4","r2_p9"],
		"Ritmos de trabajo acelerado":["r2_p5","r2_p6"],
		"Carga mental":["r2_p7","r2_p8"],
		"Cargas psicológicas emocionales":["r2_p41","r2_p42","r2_p43"],
		"Cargas de alta responsabilidad":["r2_p10","r2_p11"],
		"Cargas contradictorias o inconsistentes":["r2_p12","r2_p13"],
		"Falta de control y autonomía sobre el trabajo":["r2_p20","r2_p21","r2_p22"],
		"Limitada o nula posibilidad de desarrollo":["r2_p18","r2_p19"],
		"Limitada o inexistente capacitación":["r2_p26","r2_p27",],
		"Jornadas de trabajo extensas":["r2_p14","r2_p15"],
		"Influencia del trabajo fuera del centro laboral":["r2_p16"],
		"Influencia de las responsabilidades familiares":["r2_p17"],
		"Escasa claridad de funciones":["r2_p23","r2_p24","r2_p25"],
		"Características del liderazgo":["r2_p28","r2_p29"],
		"Relaciones sociales en el trabajo":["r2_p30","r2_p31","r2_p32"],
		"Deficiente relación con los colaboradores que supervisa":["r2_p44","r2_p45","r2_p46"],
		"Violencia laboral":["r2_p33","r2_p34","r2_p35","r2_p36","r2_p37","r2_p38","r2_p39","r2_p40"],
	}
	dimensionB={
		"Condiciones peligrosas e inseguras":["r3_p1","r3_p3"],
		"Condiciones deficientes e insalubres":["r3_p2","r3_p4"],
		"Trabajos peligrosas":["r3_p5"],
		"Cargas cuantitativas":["r3_p6","r3_p12"],
		"Ritmos de trabajo acelerado":["r3_p7","r3_p8"],
		"Carga mental":["r3_p9","r3_p10","r3_p11"],
		"Cargas psicológicas emocionales":["r3_p65","r3_p66","r3_p67","r3_p68"],
		"Cargas de alta responsabilidad":["r3_p13","r3_p14"],
		"Cargas contradictorias o inconsistentes":["r3_p15","r3_p16"],
		"Falta de control y autonomía sobre el trabajo":["r3_p25","r3_p26","r3_p27","r3_p28"],
		"Limitada o nula posibilidad de desarrollo":["r3_p23","r3_p24"],
		"Insuficiente participación y manejo del cambio":["r3_p29","r3_p30"],
		"Limitada o inexistente capacitación":["r3_p35","r3_p36"],
		"Jornadas de trabajo extensas":["r3_p17","r3_p18"],
		"Influencia del trabajo fuera del centro laboral":["r3_p19","r3_p20"],
		"Influencia de las responsabilidades familiares":["r3_p21","r3_p22"],
		"Escasa claridad de funciones":["r3_p31","r3_p32","r3_p33","r3_p34"],
		"Características del liderazgo":["r3_p37","r3_p38","r3_p39","r3_p40","r3_p41"],
		"Relaciones sociales en el trabajo":["r3_p42","r3_p43","r3_p44","r3_p45","r3_p46"],
		"Deficiente relación con los colaboradores que supervisa":["r3_p69","r3_p70","r3_p71","r3_p72"],
		"Violencia laboral":["r3_p57","r3_p58","r3_p59","r3_p60","r3_p61","r3_p62","r3_p63","r3_p64"],
		"Escasa o nula retroalimentación del desempeño":["r3_p47","r3_p48"],
		"Escasa o nulo reconocimiento y compensación":["r3_p49","r3_p50","r3_p51","r3_p52"],
		"Limitado sentido de pertenencia":["r3_p53","r3_p54"],
		"Inestabilidad laboral":["r3_p55","r3_p56"],
	}	
	domainsB={"Condiciones en el ambiente de trabajo":["r3_p1","r3_p2","r3_p3","r3_p4","r3_p5"],
		"Carga de trabajo":["r3_p6","r3_p7","r3_p8","r3_p9","r3_p10","r3_p11","r3_p12","r3_p13","r3_p14","r3_p15","r3_p16","r3_p65","r3_p66","r3_p67","r3_p68"],
		"Falta de control sobre el trabajo":["r3_p23","r3_p24","r3_p25","r3_p26","r3_p27","r3_p28","r3_p29","r3_p30","r3_p35","r3_p36",],
		"Jornada de trabajo":["r3_p17","r3_p18",],
		"Interferencia en la relación trabajo-familia":["r3_p19","r3_p20","r3_p21","r3_p22"],
		"Liderazgo":["r3_p31","r3_p32","r3_p33","r3_p34","r3_p37","r3_p38","r3_p39","r3_p40","r3_p41"],
		"Relaciones en el trabajo":["r3_p42","r3_p43","r3_p44","r3_p45","r3_p46","r3_p69","r3_p70","r3_p71","r3_p72",],
		"Violencia":["r3_p57","r3_p58","r3_p59","r3_p60","r3_p61","r3_p62","r3_p63","r3_p64",],
		"Reconocimiento del desempeño":["r3_p47","r3_p48","r3_p49","r3_p50","r3_p51","r3_p52",],
		"Insuficiente sentido de pertenencia e inestabilidad":["r3_p53","r3_p54","r3_p55","r3_p56",]
	}	
	domainsA={"Condiciones en el ambiente de trabajo":["r2_p1","r2_p2","r2_p3"],
		"Carga de trabajo":["r2_p4","r2_p5","r2_p6","r2_p7","r2_p8","r2_p9","r2_p10","r2_p11","r2_p12","r2_p13","r2_p41","r2_p42","r2_p43"],
		"Falta de control sobre el trabajo":["r2_p18","r2_p19","r2_p20","r2_p21","r2_p22","r2_p26","r2_p27",],
		"Jornada de trabajo":["r2_p14","r2_p15"],
		"Interferencia en la relación trabajo-familia":["r2_p16","r2_p17"],
		"Liderazgo":["r2_p23","r2_p24","r2_p25","r2_p28","r2_p29"],
		"Relaciones en el trabajo":["r2_p30","r2_p31","r2_p32","r2_p44","r2_p45","r2_p46"],
		"Violencia":["r2_p33","r2_p34","r2_p35","r2_p36","r2_p37","r2_p38","r2_p39","r2_p40"],
	}
	catA={"Ambiente de trabajo":["r2_p1","r2_p2","r2_p3"],
		"Factores propios de la actividad":["r2_p18","r2_p19","r2_p20","r2_p21","r2_p22","r2_p26","r2_p27","r2_p4","r2_p5","r2_p6","r2_p7","r2_p8","r2_p9","r2_p10","r2_p11","r2_p12","r2_p13","r2_p41","r2_p42","r2_p43"],
		"Organización del tiempo de trabajo":["r2_p14","r2_p15","r2_p16","r2_p17"],
		"Liderazgo y relaciones en el trabajo":["r2_p23","r2_p24","r2_p25","r2_p28","r2_p29","r2_p30","r2_p31","r2_p32","r2_p44","r2_p45","r2_p46","r2_p33","r2_p34","r2_p35","r2_p36","r2_p37","r2_p38","r2_p39","r2_p40"],
	}
	catB={"Ambiente de trabajo":["r3_p1","r3_p2","r3_p3","r3_p4","r3_p5"],
		"Factores propios de la actividad":["r3_p23","r3_p24","r3_p25","r3_p26","r3_p27","r3_p28","r3_p29","r3_p30","r3_p35","r3_p36","r3_p6","r3_p7","r3_p8","r3_p9","r3_p10","r3_p11","r3_p12","r3_p13","r3_p14","r3_p15","r3_p16","r3_p65","r3_p66","r3_p67","r3_p68"],
		"Organización del tiempo de trabajo":["r3_p17","r3_p18","r3_p19","r3_p20","r3_p21","r3_p22"],
		"Liderazgo y relaciones en el trabajo":["r3_p57","r3_p58","r3_p59","r3_p60","r3_p61","r3_p62","r3_p63","r3_p64","r3_p31","r3_p32","r3_p33","r3_p34","r3_p37","r3_p38","r3_p39","r3_p40","r3_p41","r3_p42","r3_p43","r3_p44","r3_p45","r3_p46","r3_p69","r3_p70","r3_p71","r3_p72",],
		"Entorno organizacional":["r3_p47","r3_p48","r3_p49","r3_p50","r3_p51","r3_p52","r3_p53","r3_p54","r3_p55","r3_p56",]}
	col={0:"#9be5f7",1:"#6bf56e",2:"#ffff00",3:"#ffc000",4:"#ff7070"}
	data=[]
	data2=[]
	data3=[]
	data4=[]
	data5=[]
	data6=[]
	total_level={0:{},1:{},2:{},3:{},4:{}}
	total_cat=  {0:{},1:{},2:{},3:{},4:{}}
	total_dim=  {0:{},1:{},2:{},3:{},4:{}}
	employee_names=[]
	e_count=0
	cat=[]
	domains=[]
	dimensions=[]
	if not employees:
		status="no_data"
		return JsonResponse({"status":status,'count':e_count,'data':data,'total_level':data3,'total_cat':data4,'data2':data2,'total_dim':data6,'data5':data5,'employees':employee_names,'cat':cat,'dimensions':dimensions,'domains':domains})
	for emp in employees:
		if emp.surveyB.filter(evaluation=evaluation).last() and emp.workplace.survey_type()==3:
			survey=emp.surveyB.filter(evaluation=evaluation).last()
			index=0
			for domain in domainsB:
				_sum=0
				for question in domainsB[domain]:
					_sum=_sum+(getattr(survey,RiskSurveyB._meta.get_field(question).attname) or 0)
				if domain=="Condiciones en el ambiente de trabajo":
					color=0 if _sum<5 else (1 if _sum<9 else (2 if _sum<11 else (3 if _sum<14 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Carga de trabajo":
					color=0 if _sum<15 else (1 if _sum<21 else (2 if _sum<27 else (3 if _sum<37 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Falta de control sobre el trabajo":
					color=0 if _sum<11 else (1 if _sum<16 else (2 if _sum<21 else (3 if _sum<25 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Jornada de trabajo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<4 else (3 if _sum<6 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Interferencia en la relación trabajo-familia":
					color=0 if _sum<4 else (1 if _sum<6 else (2 if _sum<8 else (3 if _sum<10 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Liderazgo":
					color=0 if _sum<9 else (1 if _sum<12 else (2 if _sum<16 else (3 if _sum<20 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Relaciones en el trabajo":
					color=0 if _sum<10 else (1 if _sum<13 else (2 if _sum<17 else (3 if _sum<21 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Violencia":
					color=0 if _sum<7 else (1 if _sum<10 else (2 if _sum<13 else (3 if _sum<16 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Reconocimiento del desempeño":
					color=0 if _sum<6 else (1 if _sum<10 else (2 if _sum<14 else (3 if _sum<18 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Insuficiente sentido de pertenencia e inestabilidad":
					color=0 if _sum<4 else (1 if _sum<6 else (2 if _sum<8 else (3 if _sum<10 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				data.append({"value":[index,e_count,_sum],'itemStyle':{'color':color}})
				index=index+1
			idx=0
			for item in catB:
				_sum=0
				for question in catB[item]:
					_sum=_sum+(getattr(survey,RiskSurveyB._meta.get_field(question).attname) or 0)
				if item=="Ambiente de trabajo":
					color=0 if _sum<5 else (1 if _sum<9 else (2 if _sum<11 else (3 if _sum<14 else 4)))
					total_cat[color][item]=total_cat[color][item]+1 if item in total_cat[color] else 1
					color=col[color]
				elif item=="Factores propios de la actividad":
					color=0 if _sum<15 else (1 if _sum<30 else (2 if _sum<45 else (3 if _sum<60 else 4)))
					total_cat[color][item]=total_cat[color][item]+1 if item in total_cat[color] else 1
					color=col[color]
				elif item=="Organización del tiempo de trabajo":
					color=0 if _sum<5 else (1 if _sum<7 else (2 if _sum<10 else (3 if _sum<13 else 4)))
					total_cat[color][item]=total_cat[color][item]+1 if item in total_cat[color] else 1
					color=col[color]
				elif item=="Liderazgo y relaciones en el trabajo":
					color=0 if _sum<14 else (1 if _sum<29 else (2 if _sum<42 else (3 if _sum<58 else 4)))
					total_cat[color][item]=total_cat[color][item]+1 if item in total_cat[color] else 1
					color=col[color]
				elif item=="Entorno organizacional":
					color=0 if _sum<10 else (1 if _sum<14 else (2 if _sum<18 else (3 if _sum<23 else 4)))
					total_cat[color][item]=total_cat[color][item]+1 if item in total_cat[color] else 1
					color=col[color]
				
				data2.append({"value":[idx,e_count,_sum],'itemStyle':{'color':color}})
				idx=idx+1
			index_dim=0
			for item in dimensionB:
				_sum=0.0
				for question in dimensionB[item]:
					_sum=_sum+(getattr(survey,RiskSurveyB._meta.get_field(question).attname) or 0.0)
				_sum=_sum/len(dimensionB[item])
				if item=="Entorno organizacional":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Condiciones peligrosas e inseguras":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Condiciones deficientes e insalubres":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Trabajos peligrosas":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Cargas cuantitativas":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Ritmos de trabajo acelerado":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Carga mental":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Cargas psicológicas emocionales":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Cargas de alta responsabilidad":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Cargas contradictorias o inconsistentes":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Falta de control y autonomía sobre el trabajo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Limitada o nula posibilidad de desarrollo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Limitada o inexistente capacitación":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Jornadas de trabajo extensas":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Influencia del trabajo fuera del centro laboral":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Influencia de las responsabilidades familiares":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Escasa claridad de funciones":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Características del liderazgo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Relaciones sociales en el trabajo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Deficiente relación con los colaboradores que supervisa":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item =="Violencia laboral":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Escasa o nula retroalimentación del desempeño":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Escasa o nulo reconocimiento y compensación":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Limitado sentido de pertenencia":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Inestabilidad laboral":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				data5.append({"value":[index_dim,e_count,_sum],'itemStyle':{'color':color}})
				index_dim=index_dim+1
			e_count=e_count+1
			employee_names.append(emp.name)
			dimensions=["Condiciones peligrosas e \ninseguras","Condiciones deficientes \ne insalubres","Trabajos peligrosas","Cargas cuantitativas","Ritmos de \ntrabajo acelerado","Carga mental","Cargas psicológicas \nemocionales","Cargas de \nalta responsabilidad","Cargas contradictorias \no inconsistentes","Falta de control y \nautonomía sobre el trabajo","Limitada o nula \nposibilidad de desarrollo","Insuficiente participación \ny manejo del cambio","Limitada o inexistente \ncapacitación","Jornadas de \ntrabajo extensas","Influencia del trabajo fuera \ndel centro laboral","Influencia de \nlas responsabilidades familiares","Escasa claridad \nde funciones","Características del liderazgo","Relaciones sociales \nen el trabajo","Deficiente relación con \nlos colaboradores que supervisa","Violencia laboral",
				"Escasa o nula retroalimentación\n del desempeño","Escasa o nulo \nreconocimiento y compensación","Limitado sentido \nde pertenencia","Inestabilidad laboral"]
			domains = ["Condiciones en el \nambiente de trabajo","Carga de trabajo","Falta de control \nsobre el trabajo","Jornada de trabajo","Interferencia en la relación \ntrabajo-familia","Liderazgo","Relaciones en el trabajo","Violencia","Reconocimiento \ndel desempeño","Insuficiente sentido de \npertenencia e inestabilidad",]
			cat = ["Ambiente de trabajo","Factores propios de \nla actividad","Organización del \ntiempo de trabajo","Liderazgo y relaciones \nen el trabajo","Entorno organizacional"]
		elif emp.surveyA.filter(evaluation=evaluation).last() and emp.workplace.survey_type()!=3:
			survey=emp.surveyA.filter(evaluation=evaluation).last()
			index=0
			for domain in domainsA:
				_sum=0
				for question in domainsA[domain]:
					_sum=_sum+(getattr(survey,RiskSurveyA._meta.get_field(question).attname) or 0)
				if domain=="Condiciones en el ambiente de trabajo":
					color=0 if _sum<3 else (1 if _sum<5 else (2 if _sum<7 else (3 if _sum<9 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Carga de trabajo":
					color=0 if _sum<12 else (1 if _sum<16 else (2 if _sum<20 else (3 if _sum<24 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Falta de control sobre el trabajo":
					color=0 if _sum<5 else (1 if _sum<8 else (2 if _sum<11 else (3 if _sum<14 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Jornada de trabajo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<4 else (3 if _sum<6 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Interferencia en la relación trabajo-familia":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<4 else (3 if _sum<6 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Liderazgo":
					color=0 if _sum<3 else (1 if _sum<5 else (2 if _sum<8 else (3 if _sum<11 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Relaciones en el trabajo":
					color=0 if _sum<5 else (1 if _sum<8 else (2 if _sum<11 else (3 if _sum<14 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				elif domain=="Violencia":
					color=0 if _sum<7 else (1 if _sum<10 else (2 if _sum<13 else (3 if _sum<16 else 4)))
					total_level[color][domain]=total_level[color][domain]+1 if domain in total_level[color] else 1
					color=col[color]
				data.append({"value":[index,e_count,_sum],'itemStyle':{'color':color}})
				index=index+1
			idx=0
			for item in catA:
				_sum=0
				for question in catA[item]:
					_sum=_sum+(getattr(survey,RiskSurveyA._meta.get_field(question).attname) or 0)
				if item=="Ambiente de trabajo":
					color=0 if _sum<3 else (1 if _sum<5 else (2 if _sum<7 else (3 if _sum<9 else 4)))
					total_cat[color][item]=total_cat[color][item]+1 if item in total_cat[color] else 1
					color=col[color]
				elif item=="Factores propios de la actividad":
					color=0 if _sum<10 else (1 if _sum<20 else (2 if _sum<30 else (3 if _sum<40 else 4)))
					total_cat[color][item]=total_cat[color][item]+1 if item in total_cat[color] else 1
					color=col[color]
				elif item=="Organización del tiempo de trabajo":
					color=0 if _sum<4 else (1 if _sum<6 else (2 if _sum<9 else (3 if _sum<12 else 4)))
					total_cat[color][item]=total_cat[color][item]+1 if item in total_cat[color] else 1
					color=col[color]
				elif item=="Liderazgo y relaciones en el trabajo":
					color=0 if _sum<10 else (1 if _sum<18 else (2 if _sum<28 else (3 if _sum<38 else 4)))
					total_cat[color][item]=total_cat[color][item]+1 if item in total_cat[color] else 1
					color=col[color]
				data2.append({"value":[idx,e_count,_sum],'itemStyle':{'color':color}})
				idx=idx+1
			index_dim=0
			for item in dimensionA:
				_sum=0.0
				for question in dimensionA[item]:
					_sum=_sum+(getattr(survey,RiskSurveyA._meta.get_field(question).attname) or 0.0)
				_sum=_sum/len(dimensionA[item])
				if item=="Condiciones peligrosas e inseguras":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Condiciones deficientes e insalubres":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Trabajos peligrosas":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Cargas cuantitativas":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Ritmos de trabajo acelerado":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Carga mental":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Cargas psicológicas emocionales":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Cargas de alta responsabilidad":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Cargas contradictorias o inconsistentes":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Falta de control y autonomía sobre el trabajo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Limitada o nula posibilidad de desarrollo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Insuficiente participación y manejo del cambio":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Limitada o inexistente capacitación":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Jornadas de trabajo extensas":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Influencia del trabajo fuera del centro laboral":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Influencia de las responsabilidades familiares":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Escasa claridad de funciones":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Características del liderazgo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Relaciones sociales en el trabajo":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Deficiente relación con los colaboradores que supervisa":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				elif item=="Violencia laboral":
					color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<3 else (3 if _sum<4 else 4)))
					total_dim[color][item]=total_dim[color][item]+1 if item in total_dim[color] else 1
					color=col[color]
				data5.append({"value":[index_dim,e_count,_sum],'itemStyle':{'color':color}})
				index_dim=index_dim+1
			e_count=e_count+1
			employee_names.append(emp.name)
			dimensions=["Condiciones peligrosas e \ninseguras","Condiciones deficientes \ne insalubres","Trabajos peligrosas","Cargas cuantitativas","Ritmos de \ntrabajo acelerado","Carga mental","Cargas psicológicas \nemocionales","Cargas de \nalta responsabilidad","Cargas contradictorias \no inconsistentes","Falta de control y \nautonomía sobre el trabajo","Limitada o nula posibilidad \nde desarrollo","Limitada o inexistente \ncapacitación","Jornadas de \ntrabajo extensas","Influencia del trabajo \nfuera del centro laboral","Influencia de las \nresponsabilidades familiares","Escasa claridad \nde funciones","Características del \nliderazgo","Relaciones sociales \nen el trabajo","Deficiente relación con \nlos colaboradores que supervisa","Violencia laboral"]
			domains = ["Condiciones en el \nambiente de trabajo","Carga de trabajo","Falta de control \nsobre el trabajo","Jornada de trabajo","Interferencia en la relación \ntrabajo-familia","Liderazgo","Relaciones en el trabajo","Violencia"]
			cat = ["Ambiente de trabajo","Factores propios de \nla actividad","Organización del \ntiempo de trabajo","Liderazgo y relaciones \nen el trabajo"]
	if employees.last().workplace.survey_type()==3:
		index=0
		for domain in domainsB:
			for categ in range(5):
				val=total_level[categ][domain] if domain in total_level[categ] else 0
				data3.append({"value":[index,categ,val],'itemStyle':{'color':col[categ]}})
			index=index+1
		index=0
		for item in catB:
			for categ in range(5):
				val=total_cat[categ][item] if item in total_cat[categ] else 0
				data4.append({"value":[index,categ,val],'itemStyle':{'color':col[categ]}})
			index=index+1
		index=0
		for item in dimensionB:
			for categ in range(5):
				val=total_dim[categ][item] if item in total_dim[categ] else 0
				data6.append({"value":[index,categ,val],'itemStyle':{'color':col[categ]}})
			index=index+1
	elif employees.last().workplace.survey_type()!=3:
		index=0
		for domain in domainsA:
			for categ in range(5):
				val=total_level[categ][domain] if domain in total_level[categ] else 0
				data3.append({"value":[index,categ,val],'itemStyle':{'color':col[categ]}})
			index=index+1
		index=0
		for item in catA:
			for categ in range(5):
				val=total_cat[categ][item] if item in total_cat[categ] else 0
				data4.append({"value":[index,categ,val],'itemStyle':{'color':col[categ]}})
			index=index+1
		index=0
		for item in dimensionA:
			for categ in range(5):
				val=total_dim[categ][item] if item in total_dim[categ] else 0
				data6.append({"value":[index,categ,val],'itemStyle':{'color':col[categ]}})
			index=index+1
	status="ok"
	if not cat:
		status="no_data"
	return JsonResponse({"status":status,'count':e_count,'data':data,'total_level':data3,'total_cat':data4,'data2':data2,'total_dim':data6,'data5':data5,'employees':employee_names,'cat':cat,'dimensions':dimensions,'domains':domains})

def get_riesgo_general(request):
	workplace_id=request.GET.get('workplace_id',None)
	if not Workplace.objects.filter(id=workplace_id, user=request.user).exists():
		return JsonResponse({'error':'not_found'}, status=403)
	evaluation=request.GET.get('evaluation',None)
	wk=Workplace.objects.filter(id=workplace_id).last()
	if evaluation is None:
		evaluation=wk.evaluation
	employees=Employee.objects.filter(workplace_id=workplace_id)
	guia3 = wk.survey_type()==3

	domainsB={"Condiciones en el ambiente de trabajo":["r3_p1","r3_p2","r3_p3","r3_p4","r3_p5"],
		"Carga de trabajo":["r3_p6","r3_p7","r3_p8","r3_p9","r3_p10","r3_p11","r3_p12","r3_p13","r3_p14","r3_p15","r3_p16","r3_p65","r3_p66","r3_p67","r3_p68"],
		"Falta de control sobre el trabajo":["r3_p23","r3_p24","r3_p25","r3_p26","r3_p27","r3_p28","r3_p29","r3_p30","r3_p35","r3_p36",],
		"Jornada de trabajo":["r3_p17","r3_p18",],
		"Interferencia en la relación trabajo-familia":["r3_p19","r3_p20","r3_p21","r3_p22"],
		"Liderazgo":["r3_p31","r3_p32","r3_p33","r3_p34","r3_p37","r3_p38","r3_p39","r3_p40","r3_p41"],
		"Relaciones en el trabajo":["r3_p42","r3_p43","r3_p44","r3_p45","r3_p46","r3_p69","r3_p70","r3_p71","r3_p72",],
		"Violencia":["r3_p57","r3_p58","r3_p59","r3_p60","r3_p61","r3_p62","r3_p63","r3_p64",],
		"Reconocimiento del desempeño":["r3_p47","r3_p48","r3_p49","r3_p50","r3_p51","r3_p52",],
		"Insuficiente sentido de pertenencia e inestabilidad":["r3_p53","r3_p54","r3_p55","r3_p56",]
	}
	domainsA={"Condiciones en el ambiente de trabajo":["r2_p1","r2_p2","r2_p3"],
		"Carga de trabajo":["r2_p4","r2_p5","r2_p6","r2_p7","r2_p8","r2_p9","r2_p10","r2_p11","r2_p12","r2_p13","r2_p41","r2_p42","r2_p43"],
		"Falta de control sobre el trabajo":["r2_p18","r2_p19","r2_p20","r2_p21","r2_p22","r2_p26","r2_p27",],
		"Jornada de trabajo":["r2_p14","r2_p15"],
		"Interferencia en la relación trabajo-familia":["r2_p16","r2_p17"],
		"Liderazgo":["r2_p23","r2_p24","r2_p25","r2_p28","r2_p29"],
		"Relaciones en el trabajo":["r2_p30","r2_p31","r2_p32","r2_p44","r2_p45","r2_p46"],
		"Violencia":["r2_p33","r2_p34","r2_p35","r2_p36","r2_p37","r2_p38","r2_p39","r2_p40"],
	}

	dimensionA={
		"Condiciones peligrosas e inseguras":["r2_p2"],
		"Condiciones deficientes e insalubres":["r2_p1"],
		"Trabajos peligrosas":["r2_p3"],
		"Cargas cuantitativas":["r2_p4","r2_p9"],
		"Ritmos de trabajo acelerado":["r2_p5","r2_p6"],
		"Carga mental":["r2_p7","r2_p8"],
		"Cargas psicológicas emocionales":["r2_p41","r2_p42","r2_p43"],
		"Cargas de alta responsabilidad":["r2_p10","r2_p11"],
		"Cargas contradictorias o inconsistentes":["r2_p12","r2_p13"],
		"Falta de control y autonomía sobre el trabajo":["r2_p20","r2_p21","r2_p22"],
		"Limitada o nula posibilidad de desarrollo":["r2_p18","r2_p19"],
		"Limitada o inexistente capacitación":["r2_p26","r2_p27",],
		"Jornadas de trabajo extensas":["r2_p14","r2_p15"],
		"Influencia del trabajo fuera del centro laboral":["r2_p16"],
		"Influencia de las responsabilidades familiares":["r2_p17"],
		"Escasa claridad de funciones":["r2_p23","r2_p24","r2_p25"],
		"Características del liderazgo":["r2_p28","r2_p29"],
		"Relaciones sociales en el trabajo":["r2_p30","r2_p31","r2_p32"],
		"Deficiente relación con los colaboradores que supervisa":["r2_p44","r2_p45","r2_p46"],
		"Violencia laboral":["r2_p33","r2_p34","r2_p35","r2_p36","r2_p37","r2_p38","r2_p39","r2_p40"],
	}
	dimensionB={
		"Condiciones peligrosas e inseguras":["r3_p1","r3_p3"],
		"Condiciones deficientes e insalubres":["r3_p2","r3_p4"],
		"Trabajos peligrosas":["r3_p5"],
		"Cargas cuantitativas":["r3_p6","r3_p12"],
		"Ritmos de trabajo acelerado":["r3_p7","r3_p8"],
		"Carga mental":["r3_p9","r3_p10","r3_p11"],
		"Cargas psicológicas emocionales":["r3_p65","r3_p66","r3_p67","r3_p68"],
		"Cargas de alta responsabilidad":["r3_p13","r3_p14"],
		"Cargas contradictorias o inconsistentes":["r3_p15","r3_p16"],
		"Falta de control y autonomía sobre el trabajo":["r3_p25","r3_p26","r3_p27","r3_p28"],
		"Limitada o nula posibilidad de desarrollo":["r3_p23","r3_p24"],
		"Insuficiente participación y manejo del cambio":["r3_p29","r3_p30"],
		"Limitada o inexistente capacitación":["r3_p35","r3_p36"],
		"Jornadas de trabajo extensas":["r3_p17","r3_p18"],
		"Influencia del trabajo fuera del centro laboral":["r3_p19","r3_p20"],
		"Influencia de las responsabilidades familiares":["r3_p21","r3_p22"],
		"Escasa claridad de funciones":["r3_p31","r3_p32","r3_p33","r3_p34"],
		"Características del liderazgo":["r3_p37","r3_p38","r3_p39","r3_p40","r3_p41"],
		"Relaciones sociales en el trabajo":["r3_p42","r3_p43","r3_p44","r3_p45","r3_p46"],
		"Deficiente relación con los colaboradores que supervisa":["r3_p69","r3_p70","r3_p71","r3_p72"],
		"Violencia laboral":["r3_p57","r3_p58","r3_p59","r3_p60","r3_p61","r3_p62","r3_p63","r3_p64"],
		"Escasa o nula retroalimentación del desempeño":["r3_p47","r3_p48"],
		"Escasa o nulo reconocimiento y compensación":["r3_p49","r3_p50","r3_p51","r3_p52"],
		"Limitado sentido de pertenencia":["r3_p53","r3_p54"],
		"Inestabilidad laboral":["r3_p55","r3_p56"],
	}
	catA={"Ambiente de trabajo":["r2_p1","r2_p2","r2_p3"],
		"Factores propios de la actividad":["r2_p18","r2_p19","r2_p20","r2_p21","r2_p22","r2_p26","r2_p27","r2_p4","r2_p5","r2_p6","r2_p7","r2_p8","r2_p9","r2_p10","r2_p11","r2_p12","r2_p13","r2_p41","r2_p42","r2_p43"],
		"Organización del tiempo de trabajo":["r2_p14","r2_p15","r2_p16","r2_p17"],
		"Liderazgo y relaciones en el trabajo":["r2_p23","r2_p24","r2_p25","r2_p28","r2_p29","r2_p30","r2_p31","r2_p32","r2_p44","r2_p45","r2_p46","r2_p33","r2_p34","r2_p35","r2_p36","r2_p37","r2_p38","r2_p39","r2_p40"],
	}
	catB={"Ambiente de trabajo":["r3_p1","r3_p2","r3_p3","r3_p4","r3_p5"],
		"Factores propios de la actividad":["r3_p23","r3_p24","r3_p25","r3_p26","r3_p27","r3_p28","r3_p29","r3_p30","r3_p35","r3_p36","r3_p6","r3_p7","r3_p8","r3_p9","r3_p10","r3_p11","r3_p12","r3_p13","r3_p14","r3_p15","r3_p16","r3_p65","r3_p66","r3_p67","r3_p68"],
		"Organización del tiempo de trabajo":["r3_p17","r3_p18","r3_p19","r3_p20","r3_p21","r3_p22"],
		"Liderazgo y relaciones en el trabajo":["r3_p57","r3_p58","r3_p59","r3_p60","r3_p61","r3_p62","r3_p63","r3_p64","r3_p31","r3_p32","r3_p33","r3_p34","r3_p37","r3_p38","r3_p39","r3_p40","r3_p41","r3_p42","r3_p43","r3_p44","r3_p45","r3_p46","r3_p69","r3_p70","r3_p71","r3_p72",],
		"Entorno organizacional":["r3_p47","r3_p48","r3_p49","r3_p50","r3_p51","r3_p52","r3_p53","r3_p54","r3_p55","r3_p56",]}

	CATEGORIA_DOMINIOS_A = {
		"Ambiente de trabajo": ["Condiciones en el ambiente de trabajo"],
		"Factores propios de la actividad": ["Carga de trabajo", "Falta de control sobre el trabajo"],
		"Organización del tiempo de trabajo": ["Jornada de trabajo", "Interferencia en la relación trabajo-familia"],
		"Liderazgo y relaciones en el trabajo": ["Liderazgo", "Relaciones en el trabajo", "Violencia"],
	}
	CATEGORIA_DOMINIOS_B = {
		"Ambiente de trabajo": ["Condiciones en el ambiente de trabajo"],
		"Factores propios de la actividad": ["Carga de trabajo", "Falta de control sobre el trabajo"],
		"Organización del tiempo de trabajo": ["Jornada de trabajo", "Interferencia en la relación trabajo-familia"],
		"Liderazgo y relaciones en el trabajo": ["Liderazgo", "Relaciones en el trabajo", "Violencia"],
		"Entorno organizacional": ["Reconocimiento del desempeño", "Insuficiente sentido de pertenencia e inestabilidad"],
	}
	DOMINIO_DIMENSIONES_A = {
		"Condiciones en el ambiente de trabajo": ["Condiciones peligrosas e inseguras", "Condiciones deficientes e insalubres", "Trabajos peligrosas"],
		"Carga de trabajo": ["Cargas cuantitativas", "Ritmos de trabajo acelerado", "Carga mental", "Cargas psicológicas emocionales", "Cargas de alta responsabilidad", "Cargas contradictorias o inconsistentes"],
		"Falta de control sobre el trabajo": ["Falta de control y autonomía sobre el trabajo", "Limitada o nula posibilidad de desarrollo", "Limitada o inexistente capacitación"],
		"Jornada de trabajo": ["Jornadas de trabajo extensas"],
		"Interferencia en la relación trabajo-familia": ["Influencia del trabajo fuera del centro laboral", "Influencia de las responsabilidades familiares"],
		"Liderazgo": ["Escasa claridad de funciones", "Características del liderazgo"],
		"Relaciones en el trabajo": ["Relaciones sociales en el trabajo", "Deficiente relación con los colaboradores que supervisa"],
		"Violencia": ["Violencia laboral"],
	}
	DOMINIO_DIMENSIONES_B = {
		"Condiciones en el ambiente de trabajo": ["Condiciones peligrosas e inseguras", "Condiciones deficientes e insalubres", "Trabajos peligrosas"],
		"Carga de trabajo": ["Cargas cuantitativas", "Ritmos de trabajo acelerado", "Carga mental", "Cargas psicológicas emocionales", "Cargas de alta responsabilidad", "Cargas contradictorias o inconsistentes"],
		"Falta de control sobre el trabajo": ["Falta de control y autonomía sobre el trabajo", "Limitada o nula posibilidad de desarrollo", "Insuficiente participación y manejo del cambio", "Limitada o inexistente capacitación"],
		"Jornada de trabajo": ["Jornadas de trabajo extensas"],
		"Interferencia en la relación trabajo-familia": ["Influencia del trabajo fuera del centro laboral", "Influencia de las responsabilidades familiares"],
		"Liderazgo": ["Escasa claridad de funciones", "Características del liderazgo"],
		"Relaciones en el trabajo": ["Relaciones sociales en el trabajo", "Deficiente relación con los colaboradores que supervisa"],
		"Violencia": ["Violencia laboral"],
		"Reconocimiento del desempeño": ["Escasa o nula retroalimentación del desempeño", "Escasa o nulo reconocimiento y compensación"],
		"Insuficiente sentido de pertenencia e inestabilidad": ["Limitado sentido de pertenencia", "Inestabilidad laboral"],
	}

	UMBRALES_CATEGORIA_GUIA_II = {
		"Ambiente de trabajo": [3,5,7,9],
		"Factores propios de la actividad": [10,20,30,40],
		"Organización del tiempo de trabajo": [4,6,9,12],
		"Liderazgo y relaciones en el trabajo": [10,18,28,38],
	}
	UMBRALES_CATEGORIA_GUIA_III = {
		"Ambiente de trabajo": [5,9,11,14],
		"Factores propios de la actividad": [15,30,45,60],
		"Organización del tiempo de trabajo": [5,7,10,13],
		"Liderazgo y relaciones en el trabajo": [14,29,42,58],
		"Entorno organizacional": [10,14,18,23],
	}

	UMBRALES_DOMINIO_GUIA_II = {
		"Condiciones en el ambiente de trabajo": [3,5,7,9],
		"Carga de trabajo": [12,16,20,24],
		"Falta de control sobre el trabajo": [5,8,11,14],
		"Jornada de trabajo": [1,2,4,6],
		"Interferencia en la relación trabajo-familia": [1,2,4,6],
		"Liderazgo": [3,5,8,11],
		"Relaciones en el trabajo": [5,8,11,14],
		"Violencia": [7,10,13,16],
	}
	UMBRALES_DOMINIO_GUIA_III = {
		"Condiciones en el ambiente de trabajo": [5,9,11,14],
		"Carga de trabajo": [15,21,27,37],
		"Falta de control sobre el trabajo": [11,16,21,25],
		"Jornada de trabajo": [1,2,4,6],
		"Interferencia en la relación trabajo-familia": [4,6,8,10],
		"Liderazgo": [9,12,16,20],
		"Relaciones en el trabajo": [10,13,17,21],
		"Violencia": [7,10,13,16],
		"Reconocimiento del desempeño": [6,10,14,18],
		"Insuficiente sentido de pertenencia e inestabilidad": [4,6,8,10],
	}
	UMBRALES_CFINAL_GUIA_II = [20,45,70,90]
	UMBRALES_CFINAL_GUIA_III = [50,75,99,140]
	NIVEL_NOMBRE = {0:"Nulo",1:"Bajo",2:"Medio",3:"Alto",4:"Muy alto"}

	def clasificar_nivel(valor, umbrales):
		if valor < umbrales[0]: return 0
		if valor < umbrales[1]: return 1
		if valor < umbrales[2]: return 2
		if valor < umbrales[3]: return 3
		return 4

	RECOMENDACION_NIVEL_GENERAL = {
		4: "Se requiere realizar el análisis de cada categoría y dominio para establecer las acciones de intervención apropiadas, mediante un Programa de intervención que deberá incluir evaluaciones específicas, y contemplar campañas de sensibilización, revisar la política de prevención de riesgos psicosociales y programas para la prevención de los factores de riesgo psicosocial, la promoción de un entorno organizacional favorable y la prevención de la violencia laboral, así como reforzar su aplicación y difusión.",
		3: "Se requiere realizar un análisis de cada categoría y dominio, de manera que se puedan determinar las acciones de intervención apropiadas a través de un Programa de intervención, que podrá incluir una evaluación específica y deberá incluir una campaña de sensibilización, revisar la política de prevención de riesgos psicosociales y programas para la prevención de los factores de riesgo psicosocial, la promoción de un entorno organizacional favorable y la prevención de la violencia laboral, así como reforzar su aplicación y difusión.",
		2: "Se requiere revisar la política de prevención de riesgos psicosociales y programas para la prevención de los factores de riesgo psicosocial, la promoción de un entorno organizacional favorable y la prevención de la violencia laboral, así como reforzar su aplicación y difusión, mediante un Programa de intervención.",
		1: "Es necesario una mayor difusión de la política de prevención de riesgos psicosociales y programas para: la prevención de los factores de riesgo psicosocial, la promoción de un entorno organizacional favorable y la prevención de la violencia laboral.",
		0: "El riesgo resulta despreciable por lo que no se requiere medidas adicionales.",
	}
	RECOMENDACION_DOMINIO_82 = {
		"Condiciones en el ambiente de trabajo": "El numeral 8.2 de la norma no contempla un catálogo específico para este dominio (regula prevención psicosocial y violencia, no condiciones físicas del lugar de trabajo). Se recomienda revisar las condiciones de seguridad e higiene del centro conforme a la normatividad aplicable (ej. NOM-030-STPS).",
		"Carga de trabajo": "Numeral 8.2 b): revisión y supervisión de que la distribución de la carga de trabajo se realice de forma equitativa, considerando el número de trabajadores, actividades a desarrollar, alcance de la actividad y su capacitación; planificar el trabajo con las pausas o periodos necesarios de descanso y rotación de tareas para evitar ritmos acelerados; e instructivos o procedimientos que definan claramente las tareas y responsabilidades.",
		"Falta de control sobre el trabajo": "Numeral 8.2 c): involucrar a los trabajadores en la toma de decisiones sobre la organización de su trabajo y en la mejora de las condiciones de trabajo y la productividad; acordar y mejorar el margen de libertad y control sobre su trabajo, impulsando el desarrollo de nuevas competencias; y reuniones para abordar áreas de oportunidad y determinar soluciones.",
		"Jornada de trabajo": "Numeral 8.2 e), numeral 2): establecer lineamientos con medidas y límites que eviten las jornadas de trabajo superiores a las previstas en la Ley Federal del Trabajo.",
		"Interferencia en la relación trabajo-familia": "Numeral 8.2 e): involucrar a los trabajadores en la definición de los horarios de trabajo cuando las condiciones lo permitan; apoyos para atender emergencias familiares comprobables; y promoción de actividades de integración familiar previo acuerdo con los trabajadores.",
		"Liderazgo": "Numeral 8.2 a): acciones para el manejo de conflictos, distribución de tiempos y determinación de prioridades en el trabajo; lineamientos contra la discriminación que fomenten equidad y respeto; mecanismos de comunicación entre supervisores/gerentes y trabajadores; instrucciones claras para atender problemas que limiten el trabajo; y capacitación/sensibilización de directivos, gerentes y supervisores.",
		"Relaciones en el trabajo": "Numerales 8.2 a) y d): además de lo señalado para Liderazgo, fomentar el apoyo social — relaciones entre trabajadores, supervisores, gerentes y patrones; reuniones periódicas (semestrales o anuales) de seguimiento; promoción de la ayuda mutua e intercambio de experiencias; y fomento de actividades culturales y deportivas.",
		"Violencia": "Numeral 8.2 g): difundir información para sensibilizar sobre violencia laboral a trabajadores, directivos, gerentes y supervisores; establecer procedimientos de actuación y seguimiento capacitando al responsable de su implementación; e informar sobre cómo denunciar actos de violencia laboral.",
		"Reconocimiento del desempeño": "Numeral 8.2 f): reconocer el desempeño sobresaliente de los trabajadores, difundir sus logros y, en su caso, expresarles sus posibilidades de desarrollo.",
		"Insuficiente sentido de pertenencia e inestabilidad": "Numerales 8.2 h) y f): promover comunicación directa y frecuente sobre problemas que afecten el trabajo, difundir cambios en la organización, dar oportunidad a los trabajadores de expresar opiniones sobre mejoras, y reforzar el reconocimiento del desempeño (ver dominio Reconocimiento).",
	}

	domains_dict = domainsB if guia3 else domainsA
	dimensions_dict = dimensionB if guia3 else dimensionA
	cat_dict = catB if guia3 else catA
	categoria_dominios = CATEGORIA_DOMINIOS_B if guia3 else CATEGORIA_DOMINIOS_A
	dominio_dimensiones = DOMINIO_DIMENSIONES_B if guia3 else DOMINIO_DIMENSIONES_A
	umbrales_categoria = UMBRALES_CATEGORIA_GUIA_III if guia3 else UMBRALES_CATEGORIA_GUIA_II
	umbrales_dominio = UMBRALES_DOMINIO_GUIA_III if guia3 else UMBRALES_DOMINIO_GUIA_II
	umbrales_cfinal = UMBRALES_CFINAL_GUIA_III if guia3 else UMBRALES_CFINAL_GUIA_II
	survey_model = RiskSurveyB if guia3 else RiskSurveyA

	cfinal_por_empleado = []
	sumas_por_dominio = {d: [] for d in domains_dict}
	sumas_por_categoria = {c: [] for c in cat_dict}
	sumas_por_dimension = {dm: [] for dm in dimensions_dict}

	for emp in employees:
		survey = emp.surveyB.filter(evaluation=evaluation).last() if guia3 else emp.surveyA.filter(evaluation=evaluation).last()
		if not survey:
			continue
		emp_cfinal = 0
		for domain, preguntas in domains_dict.items():
			_sum = 0
			for question in preguntas:
				_sum += getattr(survey, survey_model._meta.get_field(question).attname) or 0
			sumas_por_dominio[domain].append(_sum)
			emp_cfinal += _sum
		for categoria, preguntas in cat_dict.items():
			_sum = 0
			for question in preguntas:
				_sum += getattr(survey, survey_model._meta.get_field(question).attname) or 0
			sumas_por_categoria[categoria].append(_sum)
		for dimension, preguntas in dimensions_dict.items():
			_sum = 0
			for question in preguntas:
				_sum += getattr(survey, survey_model._meta.get_field(question).attname) or 0
			sumas_por_dimension[dimension].append(_sum)
		cfinal_por_empleado.append(emp_cfinal)

	if not cfinal_por_empleado:
		return JsonResponse({'status': 'no_data'})

	promedio_cfinal = sum(cfinal_por_empleado) / len(cfinal_por_empleado)
	nivel_general = clasificar_nivel(promedio_cfinal, umbrales_cfinal)
	distribucion = {0:0, 1:0, 2:0, 3:0, 4:0}
	for c in cfinal_por_empleado:
		distribucion[clasificar_nivel(c, umbrales_cfinal)] += 1

	dominios_detalle = []
	conteo_dominios_nivel = {0:0, 1:0, 2:0, 3:0, 4:0}
	for domain in domains_dict:
		valores = sumas_por_dominio[domain]
		promedio_dom = sum(valores) / len(valores) if valores else 0
		nivel_dom = clasificar_nivel(promedio_dom, umbrales_dominio[domain])
		conteo_dominios_nivel[nivel_dom] += 1
		dominios_detalle.append({'nombre': domain, 'nivel': nivel_dom, 'nivel_nombre': NIVEL_NOMBRE[nivel_dom]})

	for d in dominios_detalle:
		if d['nivel'] >= 2:
			d['recomendacion'] = RECOMENDACION_DOMINIO_82.get(d['nombre'], '')

	categorias_detalle = []
	nivel_por_categoria = {}
	for categoria in cat_dict:
		valores = sumas_por_categoria[categoria]
		promedio_cat = sum(valores) / len(valores) if valores else 0
		nivel_cat = clasificar_nivel(promedio_cat, umbrales_categoria[categoria])
		nivel_por_categoria[categoria] = nivel_cat
		categorias_detalle.append({'nombre': categoria, 'nivel': nivel_cat, 'nivel_nombre': NIVEL_NOMBRE[nivel_cat]})

	nivel_por_dominio = {d['nombre']: d['nivel'] for d in dominios_detalle}

	dimensiones_pct = {}
	for dimension, preguntas in dimensions_dict.items():
		valores = sumas_por_dimension[dimension]
		promedio_dim = sum(valores) / len(valores) if valores else 0
		maximo_dim = len(preguntas) * 4
		dimensiones_pct[dimension] = round((promedio_dim / maximo_dim) * 100, 1) if maximo_dim else 0

	jerarquia = []
	for categoria, dominios_de_categoria in categoria_dominios.items():
		dominios_json = []
		for dominio in dominios_de_categoria:
			dimensiones_json = [
				{'nombre': dim, 'porcentaje': dimensiones_pct.get(dim, 0)}
				for dim in dominio_dimensiones.get(dominio, [])
			]
			dominios_json.append({
				'nombre': dominio,
				'nivel': nivel_por_dominio.get(dominio),
				'nivel_nombre': NIVEL_NOMBRE.get(nivel_por_dominio.get(dominio)),
				'dimensiones': dimensiones_json,
			})
		jerarquia.append({
			'nombre': categoria,
			'nivel': nivel_por_categoria[categoria],
			'nivel_nombre': NIVEL_NOMBRE[nivel_por_categoria[categoria]],
			'dominios': dominios_json,
		})

	return JsonResponse({
		'status': 'ok',
		'guia': 3 if guia3 else 2,
		'riesgo_general': {
			'promedio': round(promedio_cfinal, 1),
			'nivel': nivel_general,
			'nivel_nombre': NIVEL_NOMBRE[nivel_general],
			'distribucion': distribucion,
			'total_empleados': len(cfinal_por_empleado),
		},
		'recomendacion_general': RECOMENDACION_NIVEL_GENERAL[nivel_general],
		'dominios': {
			'conteo_por_nivel': conteo_dominios_nivel,
			'detalle': dominios_detalle,
		},
		'jerarquia': jerarquia,
	})

def get_comparativo_evaluaciones(request):
	from django.test import RequestFactory
	workplace_id=request.GET.get('workplace_id',None)
	if not Workplace.objects.filter(id=workplace_id, user=request.user).exists():
		return JsonResponse({'error':'not_found'}, status=403)
	wk=Workplace.objects.filter(id=workplace_id).last()
	factory = RequestFactory()

	historial = {h.numero_evaluacion: h.fecha_finalizacion for h in wk.evaluation_history.all()}
	evaluaciones = []
	for numero in range(1, wk.evaluation + 1):
		fake_req = factory.get('/get_riesgo_general/', {'workplace_id': str(workplace_id), 'evaluation': str(numero)})
		fake_req.user = request.user
		data = json.loads(get_riesgo_general(fake_req).content)
		if data.get('status') != 'ok':
			continue
		es_actual = (numero == wk.evaluation)
		if es_actual:
			etiqueta = "Actual (en curso)"
		elif numero in historial:
			etiqueta = historial[numero].strftime('%d/%m/%Y')
		else:
			etiqueta = f"Evaluación #{numero} (sin fecha registrada)"
		evaluaciones.append({
			'numero_evaluacion': numero,
			'etiqueta': etiqueta,
			'es_actual': es_actual,
			'promedio': data['riesgo_general']['promedio'],
			'nivel': data['riesgo_general']['nivel'],
			'nivel_nombre': data['riesgo_general']['nivel_nombre'],
			'dominios': data['dominios']['detalle'],
		})

	if len(evaluaciones) < 1:
		return JsonResponse({'status': 'no_data'})

	cambio_dominios = []
	if len(evaluaciones) >= 2:
		anterior = evaluaciones[-2]['dominios']
		actual = evaluaciones[-1]['dominios']
		anterior_por_nombre = {d['nombre']: d for d in anterior}
		for dom_actual in actual:
			dom_anterior = anterior_por_nombre.get(dom_actual['nombre'])
			if not dom_anterior:
				continue
			if dom_actual['nivel'] > dom_anterior['nivel']:
				tendencia = 'empeoro'
			elif dom_actual['nivel'] < dom_anterior['nivel']:
				tendencia = 'mejoro'
			else:
				tendencia = 'sin_cambio'
			cambio_dominios.append({
				'nombre': dom_actual['nombre'],
				'nivel_anterior': dom_anterior['nivel_nombre'],
				'nivel_actual': dom_actual['nivel_nombre'],
				'tendencia': tendencia,
			})

	return JsonResponse({
		'status': 'ok',
		'evaluaciones': evaluaciones,
		'cambio_dominios': cambio_dominios,
		'etiqueta_anterior': evaluaciones[-2]['etiqueta'] if len(evaluaciones) >= 2 else None,
		'etiqueta_actual': evaluaciones[-1]['etiqueta'],
	})

class ValidateCodeList(generics.ListCreateAPIView):
	queryset = Workplace.objects.all()
	serializer_class = WorkplaceSerializer
	def get(self, request, *args, **kwargs):
		response = super(ValidateCodeList, self).get(request, args, kwargs)
		response.data=validate_code(self.request.query_params["workplace_id"])
		return response

class UserappList(generics.ListCreateAPIView):
	queryset = Userapp.objects.all()
	serializer_class = UserappSerializer
	def get(self, request, *args, **kwargs):
		response = super(UserappList, self).get(request, args, kwargs)
		for item in response.data:
			userapp=Userapp.objects.get(id=item['id'])
			item['email']=userapp.user.email
			item['name']=userapp.user.first_name
			# item['last_name']=userapp.user.last_name
			item['username']=userapp.user.username
		return response
	def post(self, request):
		print(request.data)
		dic=dict(request.data)
		dic2=dic
		print("clic1")

		if User.objects.filter(username=dic2['username']).last() is not None:
			return Response(f"Ya existe un usuario con el mismo correo {dic2['username']}", status=status.HTTP_401_UNAUTHORIZED)

		if 'webform' in dic:
			print(request.data['webform'])
			recaptcha_response =request.data['webform']
			secret_key=settings.RECAPTCHA_SECRET_KEY
			data = {
				'response': recaptcha_response,
				'secret': secret_key
			}
			resp = requests.post('https://www.google.com/recaptcha/api/siteverify', data=data)
			result_json = resp.json()
			print(result_json)
			#			 {
			#   "success": true|false,	  // whether this request was a valid reCAPTCHA token for your site
			#   "score": number			 // the score for this request (0.0 - 1.0)
			#   "action": string			// the action name for this request (important to verify)
			#   "challenge_ts": timestamp,  // timestamp of the challenge load (ISO format yyyy-MM-dd'T'HH:mm:ssZZ)
			#   "hostname": string,		 // the hostname of the site where the reCAPTCHA was solved
			#   "error-codes": [...]		// optional
			# }

			if result_json['success']:
				if result_json['score']>0.5:
					print("reCAPTCHA succeded")
				else:
					return Response("Solicitud denegada", status=status.HTTP_400_BAD_REQUEST)
			else:
				# return render(request, 'contact_sent.html', {'is_robot': True})
				return Response("El captcha no es válido, intente de nuevo", status=status.HTTP_400_BAD_REQUEST)
		if dic2['password1']!=dic2['password2']:
			return Response("Las contraseñas no coinciden", status=status.HTTP_400_BAD_REQUEST)
		user_data={
		'email':dic2['email'],
		# 'first_name':dic2['name'],
		# 'last_name':dic2['last_name'],
		'username':dic2['username'],
		'password':dic2['password1'],}
		user=None
		form=UserCreationForm(dic2)
		print("clic1")
		if form.is_valid():
			user=User.objects.create_user(**user_data)
		else:
			if user is not None:
				user.delete()
			p_.error("User Form error")
			jerror=form.errors.as_data()
			jsonerror=form.errors.get_json_data()
			p_.error(jerror)
			p_.error(jsonerror)
			if "password2" in jsonerror:
				if jsonerror["password2"][0]["code"]=="password_entirely_numeric":
					return Response("La contraseña debe tener letras", status=status.HTTP_411_LENGTH_REQUIRED)
				if jsonerror["password2"][0]["code"]=="password_mismatch":
					return Response("Las contraseñas no coinciden", status=status.HTTP_406_NOT_ACCEPTABLE)
				if jsonerror["password2"][0]["code"]=="password_too_common":
					return Response("Tu contraseña es muy facil de adivinar", status=status.HTTP_412_PRECONDITION_FAILED)
				if jsonerror["password2"][0]["code"]=="password_too_similar":
					return Response("Tu contraseña no puede ser similar a tu correo", status=status.HTTP_417_EXPECTATION_FAILED)
				if jsonerror["password2"][0]["code"]=="password_too_short":
					return Response("La contraseña es muy corta", status=status.HTTP_416_REQUESTED_RANGE_NOT_SATISFIABLE)
			error_text=""
			for error in jerror:
				if error=="password1" or error=="password2":
					v_name="Contraseña"
				else:
					v_name=User._meta.get_field(error).verbose_name
				error_text=error_text+"{}: {}, ".format(v_name,jerror[error][0].message)
			return Response(error_text, status=status.HTTP_400_BAD_REQUEST)
		print("clic2")
		if user is not None:
			dic['user']=user.id
			print("User created")
		else:
			print("User not created")
			return Response("User not created", status=status.HTTP_401_UNAUTHORIZED)
		print (dic)
		print("clic3")
		serializer = UserappSerializer(data=dic)
		if serializer.is_valid():
			print("serializer.errors1")
			try:
				new_user=serializer.save()
				new_user.validated_email=True
				ref_code = (dic2.get('referral_code') or '').strip() or request.session.pop('referral_code', None)
				if ref_code:
					partner = Partner.objects.filter(code__iexact=ref_code, active=True).first()
					if partner:
						new_user.referred_by = partner
				new_user.save()
				try:
					from django.core.management import call_command
					call_command("cargar_datos_demo", user.id)
				except Exception as demo_err:
					print(f"Error cargando datos demo: {demo_err}")
			except Exception as e:
				print(e)
				if user is not None:
					user.delete()
				return Response(e, status=status.HTTP_400_BAD_REQUEST)
			verification_code,iv=encript(f"{user.id}<->{datetime.now().timestamp()}<->{new_user.record_create.timestamp()}")
			try:
				ctx={"v_code":verification_code,"token":iv,"username":user.username,"phone":new_user.phone,"email":user.email,"name":new_user.name,"title":"Gracias por registrarte con nosotros.","type":"new_user","date_today":datetime.now().strftime('%d/%m/%Y')}
				send_mail([user.email],ctx=ctx,template="register-template.html")
				ctx={"v_code":verification_code,"token":iv,"username":user.username,"phone":new_user.phone,"email":user.email,"name":new_user.name,"title":"Nuevo usuario registrado.","type":"new_user","date_today":datetime.now().strftime('%d/%m/%Y')}
				send_mail(["normaia.sistemas@gmail.com"],ctx=ctx,template="register-template.html",subject="Registro de nuevo usuario")
			except Exception as mail_err:
				print(f"Error enviando correo de bienvenida: {mail_err}")
			print("serializer.errors")
			return Response(serializer.data, status=status.HTTP_201_CREATED)
		else:
			p_.error(serializer.errors)
			print(serializer.errors)
		print("clic4")
		if user is not None:
			user.delete()
		error_text=""
		for error in serializer.errors:
			s_errors=serializer.errors[error][0]
			v_name=Userapp._meta.get_field(error).verbose_name
			error_text=error_text+"{}: {}, ".format(v_name,s_errors)
		print("clic5")
		return Response(error_text, status=status.HTTP_400_BAD_REQUEST)

class ProductList(generics.ListCreateAPIView):
	serializer_class = ProductSerializer
	permission_classes = (IsAuthenticated,)
	authentication_classes = (TokenAuthentication,SessionAuthentication)
	def get_queryset(self):
		queryset = Product.objects.all()
		return queryset
	def get(self, request, *args, **kwargs):
		response = super(ProductList, self).get(request, args, kwargs)
		# for item in response.data:
		#	 pass
		return response
class ResultFilesList(generics.ListCreateAPIView):
	serializer_class = ResultFilesSerializer
	permission_classes = (IsAuthenticated,)
	authentication_classes = (TokenAuthentication,SessionAuthentication)
	def get_queryset(self):
		if "workplace" in self.request.query_params:
			queryset = ResultFiles.objects.filter(workplace=self.request.query_params["workplace"])
		else:
			queryset = ResultFiles.objects.none()
		return queryset
	def get(self, request, *args, **kwargs):
		response = super(ResultFilesList, self).get(request, args, kwargs)
		for item in response.data:
			res=ResultFiles.objects.filter(id=item["id"]).last()
			item["result_type"]=res.get_result_type_display()
			item["record_create"]=res.record_create.strftime('%d/%m/%Y')
		return response

class WorkplaceList(generics.ListCreateAPIView):
	serializer_class = WorkplaceSerializer
	permission_classes = (IsAuthenticated,)
	authentication_classes = (TokenAuthentication,SessionAuthentication)
	def get_queryset(self):
		queryset = Workplace.objects.filter(user_id=self.request.user.id)
		return queryset
	def get(self, request, *args, **kwargs):
		response = super(WorkplaceList, self).get(request, args, kwargs)
		# for item in response.data:
		#	 pass
		return response
	def post(self, request):
		data=request.data.copy()
		data['user']=self.request.user.id or request.user.id 
		print(request.user.id )
		print(self.request.user.id )
		# encription = AES.new("!K3y_toUnl0ck![]".encode(), AES.MODE_CBC)
		# encription = AES.new('!K3y_toUnl0ck![]', AES.MODE_CBC, '.S0m3Ran_dom-txt')
		# name=data['name'][:6]
		# to_encript = f"{name}{datetime.now()}"
		# print(to_encript)
		# ciphered_data = encription.encrypt(pad(to_encript.encode(), AES.block_size))
		# # ciphertext = encription.encrypt(to_encript)
		# print(ciphered_data)
		userapp=self.request.user.userapp
		# if userapp.workplaces_available>0 and int(data["employee_num"])<=15:
		# 	data['paid']=True
		# 	userapp.workplaces_available=userapp.workplaces_available-1
		# elif userapp.workplaces_availableB>0 and int(data["employee_num"])>15 and int(data["employee_num"])<=50: 
		# 	data['paid']=True
		# 	userapp.workplaces_availableB=userapp.workplaces_availableB-1
		# elif userapp.workplaces_availableC>0 and int(data["employee_num"])>50:
		data['paid']=True
		# 	userapp.workplaces_availableC=userapp.workplaces_availableC-1
		# else:
		# 	return Response("No cuentas con centros de trabajo disponibles.", status=status.HTTP_400_BAD_REQUEST)
		code=uuid.uuid4()
		data['access_code']=f"{code.int}-|{datetime.now().timestamp()}"
		# decrypt = AES.new("!K3y_toUnl0ck![]".encode(), AES.MODE_CBC, iv=cipher.iv)
		# original_data = unpad(decipher.decrypt(ciphered_data), AES.block_size)
		## ciphertext '\xd6\x83\x8dd!VT\x92\xaa`A\x05\xe0\x9b\x8b\xf1'
		## obj2 = AES.new('This is a key123', AES.MODE_CBC, 'This is an IV456')
		## obj2.decrypt(ciphertext)
		print(data)
		serializer = WorkplaceSerializer(data=data)
		if serializer.is_valid():
			se=serializer.save()
			userapp.save()
			return Response("ok", status=status.HTTP_201_CREATED)
		print(serializer.errors)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class StandardPagination(PageNumberPagination):
    page_size = 15
    page_query_param = 'page'

class EmployeeList(generics.ListCreateAPIView):
	serializer_class = EmployeeSerializer
	permission_classes = (IsAuthenticated,)
	authentication_classes = (TokenAuthentication,SessionAuthentication)
	pagination_class = StandardPagination
	def get_queryset(self):
		if 'workplace_id' in self.request.query_params:
			queryset = Employee.objects.filter(
				workplace_id=self.request.query_params['workplace_id'],
				workplace__user=self.request.user
			)
		else:
			queryset = Employee.objects.filter(workplace__user=self.request.user)
		return queryset
	def get(self, request, *args, **kwargs):
		response = super(EmployeeList, self).get(request, args, kwargs)
		val=0
		for item in response.data['results']:
			workplace=Workplace.objects.filter(id=item['workplace']).last()
			item['workplace_name']=workplace.name
			item['workplace_is_paid']=workplace.paid
			emp=Employee.objects.filter(id=item['id']).last()
			item['access_code']=emp.get_code()
			item['status']=emp.get_status(text="text")

			survey_b_actual = emp.surveyB.filter(evaluation=emp.workplace.evaluation).last()
			survey_a_actual = emp.surveyA.filter(evaluation=emp.workplace.evaluation).last()
			if survey_b_actual and emp.workplace.survey_type()==3:
				_sum=0
				for field in RiskSurveyB._meta.fields:
					if field.name not in ['id','employee','evaluation','record_create','record_update','r3_a','r3_b']:
						_sum=_sum+(getattr(survey_b_actual,field.attname) or 0)
				val=f"{_sum}/288"
				result_text="Riesgo nulo" if _sum<50 else ("Riesgo bajo" if _sum<75 else ("Riesgo medio" if _sum<99 else ("Riesgo alto" if _sum<140 else "Riesgo muy alto")))
			elif survey_a_actual and emp.workplace.survey_type()!=3:
				_sum=0
				for field in RiskSurveyA._meta.fields:
					if field.name not in ['id','employee','evaluation','record_create','record_update','r2_p_a','r2_p_b','es_demo']:
						_sum=_sum+(getattr(survey_a_actual,field.attname) or 0)
				val=f"{_sum}/184"
				result_text="Riesgo nulo" if _sum<20 else ("Riesgo bajo" if _sum<45 else ("Riesgo medio" if _sum<70 else ("Riesgo alto" if _sum<90 else "Riesgo muy alto")))
			else:
				val=f"0"
				result_text="Sin encuesta"
			traumado="Sin trauma"
			if hasattr(emp,"survey3"):
				survey =emp.survey3.filter(evaluation=emp.workplace.evaluation).last()
				if survey:
					if survey.r1_p1==0:
						traumado="Con trauma"
			item['results']=result_text
			item['results_value']=val
			item['results_traumado']=traumado
		response.data["detail"]="ok"
		return response
	def post(self, request):
		data=request.data.copy()
		data['user']=self.request.user.id
		serializer = EmployeeSerializer(data=data)
		if serializer.is_valid():
			workplace=Workplace.objects.filter(id=data["workplace"]).last()
			if workplace.employees.count()<workplace.employee_num:
				serializer.save()
			else:
				return Response("No puedes registrar más empleados", status=status.HTTP_400_BAD_REQUEST)
			return Response("ok", status=status.HTTP_201_CREATED)
		print(serializer.errors)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
class Employee2List(generics.ListCreateAPIView):
	serializer_class = EmployeeSerializer
	def get_queryset(self):
		if 'd' in self.request.query_params:
			code=self.request.query_params['d']
			queryset = Employee.objects.filter(workplace__access_code=code)
		else:
			queryset = Employee.objects.none()
		return queryset

	def get(self, request, *args, **kwargs):
		response = super(Employee2List, self).get(request, args, kwargs)
		return response
class PDFCreate(APIView):
	permission_classes = (IsAuthenticated,)
	authentication_classes = (TokenAuthentication,SessionAuthentication)
	def get(self, request, format=None):
		# serializer = self.serializer_class(self.get_queryset(), many=True)
		evaluation=None
		if "web" in self.request.query_params:
			web=self.request.query_params["web"]
		else:
			web=None
		if "template" in self.request.query_params:
			template=self.request.query_params["template"]
		else:
			template="muestra_chart"

		if "evaluation" in self.request.query_params:
			evaluation=self.request.query_params["evaluation"]
		if "employee_id" in self.request.query_params:
			employee_id=self.request.query_params["employee_id"]
			employee=Employee.objects.filter(id=employee_id).last()
			if not employee:
				return Response("Debes indicar un empleado válido",status=400)
			if template=="muestra_chart":
				if not hasattr(employee,'survey3'):
					return Response("Éste empleado aún no ha respondido la encuesta",status=400)
			elif template=="riesgo_psicosocial":
				if hasattr(employee,'surveyB') or hasattr(employee,'surveyA'):
					pass
				else:
					return Response("Éste empleado aún no ha respondido la encuesta",status=400)
			# workplace=self.request.user.workplaces.filter(id=employee.workplace.id).last()
			workplace=employee.workplace
			# if not workplace:
			# 	return Response("Éste empleado no pertenece a tu centro de trabajo",status=400)
		else:
			return Response("Debes indicar un empleado",status=400)
		if template=="muestra_chart":
			if evaluation:
				survey=employee.survey3.filter(evaluation=evaluation).last()
			else:
				survey=employee.survey3.last()
		domainsB={"Condiciones en el ambiente de trabajo":["r3_p1","r3_p2","r3_p3","r3_p4","r3_p5"],
			"Carga de trabajo":["r3_p6","r3_p7","r3_p8","r3_p9","r3_p10","r3_p11","r3_p12","r3_p13","r3_p14","r3_p15","r3_p16","r3_p65","r3_p66","r3_p67","r3_p68"],
			"Falta de control sobre el trabajo":["r3_p23","r3_p24","r3_p25","r3_p26","r3_p27","r3_p28","r3_p29","r3_p30","r3_p35","r3_p36",],
			"Jornada de trabajo":["r3_p17","r3_p18",],
			"Interferencia en la relación trabajo-familia":["r3_p19","r3_p20","r3_p21","r3_p22"],
			"Liderazgo":["r3_p31","r3_p32","r3_p33","r3_p34","r3_p37","r3_p38","r3_p39","r3_p40","r3_p41"],
			"Relaciones en el trabajo":["r3_p42","r3_p43","r3_p44","r3_p45","r3_p46","r3_p69","r3_p70","r3_p71","r3_p72",],
			"Violencia":["r3_p57","r3_p58","r3_p59","r3_p60","r3_p61","r3_p62","r3_p63","r3_p64",],
			"Reconocimiento del desempeño":["r3_p47","r3_p48","r3_p49","r3_p50","r3_p51","r3_p52",],
			"Insuficiente sentido de pertenencia e inestabilidad":["r3_p53","r3_p54","r3_p55","r3_p56",]
		}	
		domainsA={"Condiciones en el ambiente de trabajo":["r2_p1","r2_p2","r2_p3"],
			"Carga de trabajo":["r2_p4","r2_p5","r2_p6","r2_p7","r2_p8","r2_p9","r2_p10","r2_p11","r2_p12","r2_p13","r2_p41","r2_p42","r2_p43"],
			"Falta de control sobre el trabajo":["r2_p18","r2_p19","r2_p20","r2_p21","r2_p22","r2_p26","r2_p27",],
			"Jornada de trabajo":["r2_p14","r2_p15"],
			"Interferencia en la relación trabajo-familia":["r2_p16","r2_p17"],
			"Liderazgo":["r2_p23","r2_p24","r2_p25","r2_p28","r2_p29"],
			"Relaciones en el trabajo":["r2_p30","r2_p31","r2_p32","r2_p44","r2_p45","r2_p46"],
			"Violencia":["r2_p33","r2_p34","r2_p35","r2_p36","r2_p37","r2_p38","r2_p39","r2_p40"],
		}
		catA={"Ambiente de trabajo":["r2_p1","r2_p2","r2_p3"],
			"Factores propios de la actividad":["r2_p18","r2_p19","r2_p20","r2_p21","r2_p22","r2_p26","r2_p27","r2_p4","r2_p5","r2_p6","r2_p7","r2_p8","r2_p9","r2_p10","r2_p11","r2_p12","r2_p13","r2_p41","r2_p42","r2_p43"],
			"Organización del tiempo de trabajo":["r2_p14","r2_p15","r2_p16","r2_p17"],
			"Liderazgo y relaciones en el trabajo":["r2_p23","r2_p24","r2_p25","r2_p28","r2_p29","r2_p30","r2_p31","r2_p32","r2_p44","r2_p45","r2_p46","r2_p33","r2_p34","r2_p35","r2_p36","r2_p37","r2_p38","r2_p39","r2_p40"],
		}
		catB={"Ambiente de trabajo":["r3_p1","r3_p2","r3_p3","r3_p4","r3_p5"],
			"Factores propios de la actividad":["r3_p23","r3_p24","r3_p25","r3_p26","r3_p27","r3_p28","r3_p29","r3_p30","r3_p35","r3_p36","r3_p6","r3_p7","r3_p8","r3_p9","r3_p10","r3_p11","r3_p12","r3_p13","r3_p14","r3_p15","r3_p16","r3_p65","r3_p66","r3_p67","r3_p68"],
			"Organización del tiempo de trabajo":["r3_p17","r3_p18","r3_p19","r3_p20","r3_p21","r3_p22"],
			"Liderazgo y relaciones en el trabajo":["r3_p57","r3_p58","r3_p59","r3_p60","r3_p61","r3_p62","r3_p63","r3_p64","r3_p31","r3_p32","r3_p33","r3_p34","r3_p37","r3_p38","r3_p39","r3_p40","r3_p41","r3_p42","r3_p43","r3_p44","r3_p45","r3_p46","r3_p69","r3_p70","r3_p71","r3_p72",],
			"Entorno organizacional":["r3_p47","r3_p48","r3_p49","r3_p50","r3_p51","r3_p52","r3_p53","r3_p54","r3_p55","r3_p56",]}
		col={0:"#9be5f7",1:"#6bf56e",2:"#ffff00",3:"#ffc000",4:"#ff7070"}
		name={0:"Nulo",1:"Bajo",2:"Medio",3:"Alto",4:"Muy alto"}
		val=None
		result_class=None
		result_text=None
		promedio=None
		trauma_survey=[]
		survey_result=[]
		if template=="muestra_chart":
			for question in TraumaSurvey._meta.fields:
				if question.name in ['id','employee','evaluation','record_create','record_update']:
					continue
				response=getattr(survey,question.attname)
				answers=[]
				for choice in question.choices:
					char="\U00002B1B" if choice[0] == response else "\U00002B1C"
					answers.append(f"{char} {choice[1]}")
				trauma_survey.append({
					"title":question.verbose_name,
					"answers":answers,
				})
		elif template=="riesgo_psicosocial":
			if employee.surveyB.filter(evaluation=evaluation).last() and employee.workplace.survey_type()==3:
				res=0
				for field in RiskSurveyB._meta.fields:
					if field.name not in ['id','employee','evaluation','record_create','record_update','r3_a','r3_b']:
						res=res+(getattr(employee.surveyB.filter(evaluation=evaluation).last(),field.attname) or 0)
				val=f"{res}/288"
				promedio=round((res/288)*100)
				result_class,result_text=("#9be5f7","Riesgo nulo") if res<50 else (("#6bf56e","Riesgo bajo") if res<75 else (("#ffff00","Riesgo medio") if res<99 else (("#ffc000","Riesgo alto") if res<140 else ("#ff7070","Riesgo muy alto"))))
				survey=employee.surveyB.filter(evaluation=evaluation).last()
				for domain in domainsB:
					_sum=0
					for question in domainsB[domain]:
						_sum=_sum+(getattr(survey,RiskSurveyB._meta.get_field(question).attname) or 0)
					max_score=len(domainsB[domain])*4
					if domain=="Condiciones en el ambiente de trabajo":
						color=0 if _sum<5 else (1 if _sum<9 else (2 if _sum<11 else (3 if _sum<14 else 4)))
					elif domain=="Carga de trabajo":
						color=0 if _sum<15 else (1 if _sum<21 else (2 if _sum<27 else (3 if _sum<37 else 4)))
					elif domain=="Falta de control sobre el trabajo":
						color=0 if _sum<11 else (1 if _sum<16 else (2 if _sum<21 else (3 if _sum<25 else 4)))
					elif domain=="Jornada de trabajo":
						color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<4 else (3 if _sum<6 else 4)))
					elif domain=="Interferencia en la relación trabajo-familia":
						color=0 if _sum<4 else (1 if _sum<6 else (2 if _sum<8 else (3 if _sum<10 else 4)))
					elif domain=="Liderazgo":
						color=0 if _sum<9 else (1 if _sum<12 else (2 if _sum<16 else (3 if _sum<20 else 4)))
					elif domain=="Relaciones en el trabajo":
						color=0 if _sum<10 else (1 if _sum<13 else (2 if _sum<17 else (3 if _sum<21 else 4)))
					elif domain=="Violencia":
						color=0 if _sum<7 else (1 if _sum<10 else (2 if _sum<13 else (3 if _sum<16 else 4)))
					elif domain=="Reconocimiento del desempeño":
						color=0 if _sum<6 else (1 if _sum<10 else (2 if _sum<14 else (3 if _sum<18 else 4)))
					elif domain=="Insuficiente sentido de pertenencia e inestabilidad":
						color=0 if _sum<4 else (1 if _sum<6 else (2 if _sum<8 else (3 if _sum<10 else 4)))
					survey_result.append({"type":"Dominio","cat_dom":domain,"value":f"{_sum}/{max_score}","promedio":round((_sum/max_score)*100),"result_class":col[color],"result_text":name[color]})
				for item in catB:
					_sum=0
					for question in catB[item]:
						_sum=_sum+(getattr(survey,RiskSurveyB._meta.get_field(question).attname) or 0)
					max_score=len(catB[item])*4	
					if item=="Ambiente de trabajo":
						color=0 if _sum<5 else (1 if _sum<9 else (2 if _sum<11 else (3 if _sum<14 else 4)))
					elif item=="Factores propios de la actividad":
						color=0 if _sum<15 else (1 if _sum<30 else (2 if _sum<45 else (3 if _sum<60 else 4)))
					elif item=="Organización del tiempo de trabajo":
						color=0 if _sum<5 else (1 if _sum<7 else (2 if _sum<10 else (3 if _sum<13 else 4)))
					elif item=="Liderazgo y relaciones en el trabajo":
						color=0 if _sum<14 else (1 if _sum<29 else (2 if _sum<42 else (3 if _sum<58 else 4)))
					elif item=="Entorno organizacional":
						color=0 if _sum<10 else (1 if _sum<14 else (2 if _sum<18 else (3 if _sum<23 else 4)))
					survey_result.append({"type":"Categoría","cat_dom":item,"value":f"{_sum}/{max_score}","promedio":round((_sum/max_score)*100),"result_class":col[color],"result_text":name[color]})
			
			elif employee.surveyA.filter(evaluation=evaluation).last() and employee.workplace.survey_type()!=3:
				res=0
				for field in RiskSurveyA._meta.fields:
					if field.name not in ['id','employee','evaluation','record_create','record_update','r2_p_a','r2_p_b','es_demo']:
						res=res+(getattr(employee.surveyA.filter(evaluation=evaluation).last(),field.attname) or 0)
				val=f"{res}/184"
				promedio=round((res/184)*100)
				result_class,result_text=("#9be5f7","Riesgo nulo") if res<20 else (("#6bf56e","Riesgo bajo") if res<45 else (("#ffff00","Riesgo medio") if res<70 else (("#ffc000","Riesgo alto") if res<90 else ("#ff7070","Riesgo muy alto"))))
				survey=employee.surveyA.filter(evaluation=evaluation).last()
				for domain in domainsA:
					_sum=0
					for question in domainsA[domain]:
						_sum=_sum+(getattr(survey,RiskSurveyA._meta.get_field(question).attname) or 0)
					max_score=len(domainsA[domain])*4
					if domain=="Condiciones en el ambiente de trabajo":
						color=0 if _sum<3 else (1 if _sum<5 else (2 if _sum<7 else (3 if _sum<9 else 4)))
					elif domain=="Carga de trabajo":
						color=0 if _sum<12 else (1 if _sum<16 else (2 if _sum<20 else (3 if _sum<24 else 4)))
					elif domain=="Falta de control sobre el trabajo":
						color=0 if _sum<5 else (1 if _sum<8 else (2 if _sum<11 else (3 if _sum<14 else 4)))
					elif domain=="Jornada de trabajo":
						color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<4 else (3 if _sum<6 else 4)))
					elif domain=="Interferencia en la relación trabajo-familia":
						color=0 if _sum<1 else (1 if _sum<2 else (2 if _sum<4 else (3 if _sum<6 else 4)))
					elif domain=="Liderazgo":
						color=0 if _sum<3 else (1 if _sum<5 else (2 if _sum<8 else (3 if _sum<11 else 4)))
					elif domain=="Relaciones en el trabajo":
						color=0 if _sum<5 else (1 if _sum<8 else (2 if _sum<11 else (3 if _sum<14 else 4)))
					elif domain=="Violencia":
						color=0 if _sum<7 else (1 if _sum<10 else (2 if _sum<13 else (3 if _sum<16 else 4)))
					survey_result.append({"type":"Dominio","cat_dom":domain,"value":f"{_sum}/{max_score}","promedio":round((_sum/max_score)*100),"result_class":col[color],"result_text":name[color]})
				for item in catA:
					_sum=0
					for question in catA[item]:
						_sum=_sum+(getattr(survey,RiskSurveyA._meta.get_field(question).attname) or 0)
					max_score=len(catA[item])*4
					if item=="Ambiente de trabajo":
						color=0 if _sum<3 else (1 if _sum<5 else (2 if _sum<7 else (3 if _sum<9 else 4)))
					elif item=="Factores propios de la actividad":
						color=0 if _sum<10 else (1 if _sum<20 else (2 if _sum<30 else (3 if _sum<40 else 4)))
					elif item=="Organización del tiempo de trabajo":
						color=0 if _sum<4 else (1 if _sum<6 else (2 if _sum<9 else (3 if _sum<12 else 4)))
					elif item=="Liderazgo y relaciones en el trabajo":
						color=0 if _sum<10 else (1 if _sum<18 else (2 if _sum<28 else (3 if _sum<38 else 4)))
					survey_result.append({"type":"Categoría","cat_dom":item,"value":f"{_sum}/{max_score}","promedio":round((_sum/max_score)*100),"result_class":col[color],"result_text":name[color]})
			else:
				val=f""
				result_class,result_text=("#f0f0f0","Sin encuesta")
			
		ctx={"user":self.request.user.id,
			"name":workplace.name,
			"address":workplace.address,
			"address_locality":workplace.address_locality,
			"address_state":workplace.address_state,
			"address_postal_code":workplace.address_postal_code,
			"phone":self.request.user.userapp.phone,
			"emp_name":employee.name,
			"gender":employee.get_gender_display(),
			"age":employee.get_age_display(),
			"civil_state":employee.get_civil_state_display(),
			"study_level":employee.get_study_level_display(),
			"ocupation":employee.ocupation,
			"department":employee.department,
			"charge_type":employee.get_charge_type_display(),
			"contract_type":employee.get_contract_type_display(),
			"employee_type":employee.get_employee_type_display(),
			"shift_type":employee.get_shift_type_display(),
			"shift_rotation":employee.get_shift_rotation_display(),
			"time_in_charge":employee.get_time_in_charge_display(),
			"exp":employee.get_exp_display(),
			"result_val":val or "",
			"result_class":result_class or "",
			"result_text":result_text or "",
			"promedio":promedio or "",
			"trauma_survey":trauma_survey,
			"survey_result":survey_result,
		}
		# template="riesgo_psicosocial"
		# template="traumatics_events_p1"
		task=pdf_reports_task.delay(self.request.user.id,template,ctx)
		time.sleep(7)
		t=AsyncResult(task.id)
		if t.successful():
			p_.error("task succeded")
			# else:
			# 	response=f"https://normaia.ihes.mx/files/tmp/{self.request.user.id}/{template}.pdf"
		return Response({"pdf":f"https://normaia.ihes.mx/files/tmp/{self.request.user.id}/{template}.pdf"})

class EndEvaluation(APIView):
	http_method_names = ['get',]
	permission_classes = (IsAuthenticated,)
	authentication_classes = (TokenAuthentication,SessionAuthentication)
	def get(self, request, format=None):
		try:
			workplace_id=request.query_params.get('workplace_id') or request.data.get('workplace_id')
			if not request.user.workplaces.filter(id=workplace_id).exists():
				return Response({'status':'error', 'error':'Centro de trabajo no encontrado.'}, status=403)
			workplace=Workplace.objects.filter(id=workplace_id).last()
			if workplace.es_demo:
				return Response({'status':'error', 'error':'No es posible avanzar evaluacion en un centro de trabajo de demostracion.'})
			employee_ids = list(workplace.employees.values_list('id', flat=True))
			if not employee_ids:
				return Response({'status':'error', 'error':'No puedes finalizar la aplicación: no tienes empleados registrados en este centro de trabajo.'})
			if workplace.survey_type() != 3:
				respondidas = RiskSurveyA.objects.filter(evaluation=workplace.evaluation, employee_id__in=employee_ids).count()
			else:
				respondidas = RiskSurveyB.objects.filter(evaluation=workplace.evaluation, employee_id__in=employee_ids).count()
			if respondidas == 0:
				return Response({'status':'error', 'error':'No puedes finalizar la aplicación: ningún empleado ha respondido la encuesta todavía.'})
			EvaluationHistory.objects.create(
				workplace=workplace,
				numero_evaluacion=workplace.evaluation,
				guia=workplace.survey_type(),
			)
			workplace.evaluation=workplace.evaluation+1
			workplace.paid=True
			workplace.save()
			return Response({'status':'ok'})
		except Exception as e:
			return Response({'status':'error', 'error':f"error::{e}"})
# Views Web Page
class WebIndex(View):
	def get(self, request, *args, **kwargs):
		return render(request, 'default.html')

class BotView(View):
	def get(self, request, *args, **kwargs):
		print ("verify token")
		if self.request.GET['hub.verify_token'] == '122333444455555':
			return HttpResponse(self.request.GET['hub.challenge'])
		else:
			return HttpResponse('Error, invalid token')

	@csrf_exempt
	def dispatch(self, request, *args, **kwargs):
		return View.dispatch(self, request, *args, **kwargs)

	def post(self, request, *args, **kwargs):
		incoming_message = json.loads(self.request.body.decode('utf-8'))
		# TOKEN="EAAM7xv9LqRsBADjhZAdkzTZClkFtgqsLYrATjHZAt2da23rQXG6UyUudLFqajVrTNkUcRRS56Bbtm3LGIktk9pKVW2SlaHDqZAK7ILwiGmyma0dPGmC8C34NK8ZAu4hf2uRS0ofYnpFZCf7RZBJemiRVUFZCQKjSTUOo9Juxjli1XpsBzmXNAZBWY"##pruebabot
		TOKEN="EAAM7xv9LqRsBAJH0bFL5ZCS8dZAO01qkBJsqzMwvUjtgElqYypH7ZCEOiEuQZA7VMW7GVLGUdpyNcNdJSaR8MFX8NDdOXkKPn4tCO4MvZBakqjUCpYJUqZAYCXZCl0BGeegn9mVHtNjNsGyQEz8xfrUV8CEfPGZAuzGXKhXDJlq9LgZDZD"

		# {"greeting": [{"locale":"default","text":"Bienvenido {{user_first_name}}, NOM-035-STPS-2018 Servicio Profesional, Te ayudamos a dar cumplimiento a las obligaciones patronales en la identificación, análisis y prevención de factores de los riesgos psicosociales en el trabajo."}],"get_started": {"payload": "USER_START_CHAT"}}
		# import requests,json
		# >>> token="EAAM7xv9LqRsBAJH0bFL5ZCS8dZAO01qkBJsqzMwvUjtgElqYypH7ZCEOiEuQZA7VMW7GVLGUdpyNcNdJSaR8MFX8NDdOXkKPn4tCO4MvZBakqjUCpYJU
		# qZAYCXZCl0BGeegn9mVHtNjNsGyQEz8xfrUV8CEfPGZAuzGXKhXDJlq9LgZDZD"
		# >>> url="https://035.ihes.mx/static/web-assets/media/new_user.mp4"
		# '{"attachment_id":"2708755729338290"}'
		# >>> url2="https://035.ihes.mx/static/web-assets/media/new_employee.mp4"
		# '{"attachment_id":"657789728253254"}'
		# >>> url3="https://035.ihes.mx/static/web-assets/media/workplace_register.mp4"
		# '{"attachment_id":"345793446487018"}'
		# >>> urla=f"https://graph.facebook.com/v8.0/me/message_attachments?access_token={token}"
		# >>> data={"message":{"attachment":{"type":"video","payload":{"is_reusable":True,"url":url}}}}
		# >>> response=requests.post(urla,headers={"Content-Type":"application/json"},data=json.dumps(data))
		# >>> response.text

		for entry in incoming_message['entry']:
			if not 'messaging' in entry:
				print (" not a message %s"%entry)
				continue
			for message in entry['messaging']:
				fbid=message['sender']['id']
				page_id=entry['id']
				if not 'delivery' in message:
					print(entry)
				if 'postback' in message:
					if 'payload' in message['postback']:
						payload=message['postback']['payload']
						if payload=="USER_START_CHAT":
							# post_facebook_message_fbid(TOKEN,fbid)
							post_facebook_template_msg(TOKEN,fbid,f"msg",f"Bienvenido, a nuestro bot de atención al cliente")
						if payload=="USER_menu":
							buttons=[{"type":"postback","title":"Nuestros servicios","payload":"GET_service_desc"},
								{"type":"postback","title":"¿Cuánto cuesta?","payload":"GET_prices"},
								{"type":"postback","title":"Prueba sin costo","payload":"GET_free_trial"}]
							post_facebook_message_button(TOKEN,fbid,"Opciones",buttons)
						elif payload=="USER_support":
							buttons=[{"content_type":"text","title":"No puedo registrarme","payload":"USER_help_register"},
								# {"content_type":"text","title":"No puedo pagar","payload":"USER_help_payments"},
								{"content_type":"text","title":"Captura centros de trabajo","payload":"USER_help_workplaces"},
								{"content_type":"text","title":"Captura de empleados","payload":"USER_help_employees"},
								{"content_type":"text","title":"Otro problema","payload":"USER_more_help"}]
							post_facebook_message_quick(TOKEN,fbid,"¿Que problema tienes?",buttons)
						elif payload=="GET_prices":
							buttons=[{"type":"postback","title":"Hacer prueba sin costo","payload":"GET_free_trial"},
								{"type":"phone_number","title":"Llamanos","payload":"+1523314414070"},
								{"type":"postback","title":"Volver al menu","payload":"USER_START_CHAT"},]
							post_facebook_message_button(TOKEN,fbid,"Costos del servicio profesional\nNOM035\nMicro y pequeñas empresas (1-5 Trabajadores) $1750MXN\nEmpresas Categoria 1 (6-15 Tabajadores) $4500MXN(6 Trabajadores + $750 por trabajador adicional)\nEmpresas Categoria 2 (16-50 Tabajadores) $11200MXN(16 Trabajadores + $700 por trabajador adicional)\nEmpresas Categoria 3 (51-100 Tabajadores) $33150MXN(51 Trabajadores + $650 por trabajador adicional)",buttons)
						elif payload=="GET_service_desc":
							buttons=[{"type":"postback","title":"Hacer prueba sin costo","payload":"GET_free_trial"},
								{"type":"phone_number","title":"Llamanos","payload":"+1523314414070"},
								{"type":"postback","title":"Volver al menu","payload":"USER_START_CHAT"},]
							post_facebook_message_button(TOKEN,fbid,"NOM-035-STPS-2018 Servicio Profesional\nTe ayudamos a dar cumplimiento a las obligaciones patronales en la identificación, análisis y prevención de factores de los riesgos psicosociales en el trabajo.",buttons)
						elif payload=="GET_free_trial":
							buttons=[{"type":"postback","title":"Ir a la tienda de aplicaciones","payload":"GET_free_trial"},
								{"type":"postback","title":"Volver al menu","payload":"USER_START_CHAT"},]
							post_facebook_message_button(TOKEN,fbid,"Descarga nuestra Aplicación para iniciar tu prueba gratuita.",buttons)
						
				elif 'message' in message:
					if "quick_reply" in message['message']:
						payload=message['message']["quick_reply"]["payload"]
						if payload=="USER_help_register":
							post_facebook_template_media(TOKEN,fbid,2708755729338290)
							# elif payload=="USER_help_payments":
							# 	post_facebook_template_media(TOKEN,fbid,356832572428320)
						elif payload=="USER_help_workplaces":
							post_facebook_template_media(TOKEN,fbid,657789728253254)
						elif payload=="USER_help_employees":
							post_facebook_template_media(TOKEN,fbid,345793446487018)
						elif payload=="USER_more_help":
							buttons=[{"type":"web_url","title":"Abrir Caso","url":"https://035.ihes.mx/"},
								{"type":"postback","title":"Volver al menu","payload":"USER_support"},]
							post_facebook_message_button(TOKEN,fbid,"Ayuda adicional",buttons)
						# buttons=[{"type":"web_url","title":"Ver video tutorial","url":"https://035.ihes.mx/"},
						# 	{"type":"postback","title":"Volver al menu","payload":"USER_support"},]
						# post_facebook_message_button(TOKEN,fbid,"Ayuda con la aplicación de encuestas",buttons)
				# 	print(message['message'])
				# 	buttons=[{"type":"postback","title":"opcion 1","payload":"USER_START_CHAT"},
				# 			{"type":"postback","title":"opcion 2","payload":"GET_Clases_a_distancia"},
				# 			{"type":"postback","title":"opcion 3","payload":"GET_Etc"}]
				# 	post_facebook_message_button(TOKEN,fbid,"chat iniciado",buttons)
				# 	# post_facebook_template_msg(TOKEN,fbid,f"Hola,fbid {fbid}, page id {page_id}")
		return HttpResponse()

def post_facebook_message_fbid(TK,facebook_id,recevied_message):
	post_message_url=f'https://graph.facebook.com/me/messages?access_token={TK}'
	response_msg = json.dumps({"messaging_type": "RESPONSE","recipient":{"id":facebook_id}, "message":{"text":recevied_message}})
	status = requests.post(post_message_url,headers={"Content-Type":"application/json"},data=response_msg)
	print (status.json())

def post_facebook_template_msg(TK,facebook_id,message,title="Bienvenido a nuestro bot.",subtitle="Hola, en que podemos ayudarte?"):
	post_message_url=f'https://graph.facebook.com/me/messages?access_token={TK}'
	response_msg = json.dumps({"recipient":{"id":facebook_id},"message":{"attachment":{"type":"template",
		"payload":{"template_type":"generic","elements":[{"title":title,"subtitle":subtitle,
			"image_url":"https://035.ihes.mx/static/web-assets/img/logo-BOT2.jpg",
			"default_action": {"type": "web_url","url": "https://035.ihes.mx/","messenger_extensions": False,"webview_height_ratio": "FULL"
			},"buttons":[{"type":"postback","title":"Prueba sin costo","payload":"GET_free_trial"},
				{"type":"postback","title":"Más información","payload":"USER_menu"},
				{"type":"postback","title":"Soporte para usuarios","payload":"USER_support"},]}]}}}})
	status = requests.post(post_message_url,headers={"Content-Type":"application/json"},data=response_msg)
	print (status.json())
def post_facebook_template_media(TK,facebook_id,attachment):
	post_message_url=f'https://graph.facebook.com/me/messages?access_token={TK}'
	response_msg = json.dumps({"recipient":{"id":facebook_id},"message":{"attachment":{"type":"template",
		"payload":{"template_type":"media","elements":[{"media_type":"video","attachment_id":attachment,#"url":url,
			"buttons":[{"type":"postback","title":"Más información","payload":"USER_menu"},
				{"type":"postback","title":"Menú de ayuda","payload":"USER_support"},
				{"type":"phone_number","title":"Llamanos","payload":"+1523314414070"}]}]}}}})
	status = requests.post(post_message_url,headers={"Content-Type":"application/json"},data=response_msg)
	print (status.json())


def post_facebook_message_button(TK,user,button_text,buttons):
	post_message_url = 'https://graph.facebook.com/me/messages?access_token=%s'%TK
	if len(buttons)>0:
		response_msg = json.dumps(
			{"messaging_type": "RESPONSE","recipient":{"id":user},"message":{"attachment":{
			"type":"template","payload":{"template_type":"button","text":button_text,"buttons":buttons}}}})
	status = requests.post(post_message_url, headers={"Content-Type": "application/json"},data=response_msg)
	print (status.json())
def post_facebook_message_quick(TK,user,button_text,buttons):
	post_message_url = 'https://graph.facebook.com/me/messages?access_token=%s'%TK
	if len(buttons)>0:
		response_msg = json.dumps({"recipient":{"id":user},"messaging_type": "RESPONSE",
			"message":{"text": button_text,"quick_replies":buttons}})
	status = requests.post(post_message_url, headers={"Content-Type": "application/json"},data=response_msg)
	print (status.json())





from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.contrib.auth.models import User
import stripe
from django.conf import settings

@csrf_exempt
def stripe_webhook(request):
    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
    endpoint_secret = settings.STRIPE_WEBHOOK_SECRET
    try:
        event = stripe.Webhook.construct_event(payload, sig_header, endpoint_secret)
    except ValueError as e:
        p_.error(f"Stripe webhook: payload invalido: {e}")
        return HttpResponse(status=400)
    except stripe.error.SignatureVerificationError as e:
        p_.error(f"Stripe webhook: firma invalida: {e}")
        return HttpResponse(status=400)

    event_type = event['type']
    p_.info(f"Stripe webhook recibido: {event_type}")

    if event_type == 'checkout.session.completed':
        _stripe_handle_checkout_completed(event['data']['object'])
    elif event_type == 'invoice.paid':
        _stripe_handle_invoice_paid(event['data']['object'])
    elif event_type == 'invoice.payment_failed':
        _stripe_handle_payment_failed(event['data']['object'])
    elif event_type in ('customer.subscription.deleted', 'customer.subscription.updated'):
        _stripe_handle_subscription_change(event['data']['object'])

    return HttpResponse(status=200)


def _stripe_handle_checkout_completed(session):
    customer_email = (
        session.get('customer_email') or
        session.get('customer_details', {}).get('email')
    )
    user_id = session.get('metadata', {}).get('user_id')
    plan_key = session.get('metadata', {}).get('plan_key')
    try:
        if user_id:
            user = User.objects.get(id=user_id)
        else:
            user = User.objects.get(email=customer_email)
    except User.DoesNotExist:
        p_.error(
            f"Stripe webhook checkout.session.completed: no se encontro usuario "
            f"(user_id={user_id}, email={customer_email}), evento ignorado"
        )
        return

    product_name = plan_key or session.get('metadata', {}).get('product_type')
    if not product_name:
        p_.warning(f"Stripe webhook checkout.session.completed: sin plan_key/product_type, user_id={user.id}")
        return

    assign_nom035_credits(user, product_name)
    try:
        userapp = user.userapp
        from surveys.stripe_plans import PLANS as STRIPE_PLANS
        _plan = STRIPE_PLANS.get(plan_key, {})
        if _plan.get('modulo') == 'psicometria':
            userapp.psico_plan_key = plan_key
        else:
            userapp.stripe_plan_key = plan_key
        if session.get('customer'):
            userapp.stripe_customer_id = session['customer']
        if session.get('subscription'):
            userapp.stripe_subscription_id = session['subscription']
        userapp.save()
        PlanPurchaseEvent.objects.create(
            user=userapp.user,
            plan_key=plan_key,
            modulo=_plan.get('modulo', ''),
            precio=_plan.get('precio', 0),
            periodo=_plan.get('periodo', ''),
            stripe_customer_id=userapp.stripe_customer_id,
        )
        if userapp.referred_by and userapp.referred_by.active and not PartnerCommission.objects.filter(partner=userapp.referred_by, referred_userapp=userapp).exists():
            from surveys.services.partner_commissions import calcular_comision_partner
            monto_plan = _plan.get('precio', 0) or 0
            monto_comision = calcular_comision_partner(plan_key, monto_plan, es_renovacion=False)
            if monto_comision > 0:
                PartnerCommission.objects.create(
                    partner=userapp.referred_by,
                    referred_userapp=userapp,
                    plan_key=plan_key,
                    tipo='venta',
                    monto_plan=monto_plan,
                    monto_comision=monto_comision,
                )
                p_.info(f"Comision de venta generada para partner {userapp.referred_by.code}: ${monto_comision} por user_id={user.id}")
        p_.info(f"Stripe webhook: plan {plan_key} activado para user_id={user.id}")
    except Exception as e:
        p_.error(f"Stripe webhook checkout.session.completed: error guardando plan para user_id={user.id}: {e}")


def _stripe_handle_invoice_paid(invoice):
    # Solo re-acreditar en renovaciones de suscripcion, no en la factura inicial
    # (esa ya se acredito via checkout.session.completed, evita doble acreditacion).
    if invoice.get('billing_reason') != 'subscription_cycle':
        return
    customer_id = invoice.get('customer')
    try:
        userapp = Userapp.objects.get(stripe_customer_id=customer_id)
    except Userapp.DoesNotExist:
        p_.error(f"Stripe webhook invoice.paid: no se encontro Userapp para customer={customer_id}")
        return
    plan_key = userapp.stripe_plan_key or userapp.psico_plan_key
    if not plan_key:
        p_.warning(f"Stripe webhook invoice.paid: Userapp sin plan activo, customer={customer_id}")
        return
    assign_nom035_credits(userapp.user, plan_key)
    if userapp.referred_by and userapp.referred_by.active:
        invoice_id = invoice.get('id')
        ya_registrada = invoice_id and PartnerCommission.objects.filter(stripe_invoice_id=invoice_id).exists()
        if not ya_registrada:
            from surveys.services.partner_commissions import calcular_comision_partner
            from surveys.stripe_plans import PLANS as STRIPE_PLANS
            monto_plan = STRIPE_PLANS.get(plan_key, {}).get('precio', 0) or 0
            monto_comision = calcular_comision_partner(plan_key, monto_plan, es_renovacion=True)
            if monto_comision > 0:
                PartnerCommission.objects.create(
                    partner=userapp.referred_by,
                    referred_userapp=userapp,
                    plan_key=plan_key,
                    tipo='renovacion',
                    monto_plan=monto_plan,
                    monto_comision=monto_comision,
                    stripe_invoice_id=invoice_id,
                )
                p_.info(f"Comision de renovacion generada para partner {userapp.referred_by.code}: ${monto_comision}, customer={customer_id}")
    p_.info(f"Stripe webhook: renovacion acreditada, plan={plan_key}, customer={customer_id}")


def _stripe_handle_payment_failed(invoice):
    customer_id = invoice.get('customer')
    p_.warning(f"Stripe webhook invoice.payment_failed: customer={customer_id}")


def _stripe_handle_subscription_change(subscription):
    customer_id = subscription.get('customer')
    status = subscription.get('status')
    if status not in ('canceled', 'unpaid', 'past_due'):
        return
    try:
        userapp = Userapp.objects.get(stripe_customer_id=customer_id)
    except Userapp.DoesNotExist:
        p_.error(f"Stripe webhook subscription change: no se encontro Userapp para customer={customer_id}")
        return
    userapp.stripe_plan_key = ''
    userapp.psico_plan_key = ''
    userapp.save()
    p_.info(f"Stripe webhook: plan activo limpiado (status={status}), customer={customer_id}, creditos existentes intactos")


class ReporteHTMLView(LoginRequiredMixin, View):
    login_url = '/login/'
    def get(self, request, employee_id, evaluation):
        from surveys.models import Employee, RiskSurveyA, RiskSurveyB
        employee = Employee.objects.filter(id=employee_id).last()
        if not employee:
            return HttpResponse('Empleado no encontrado', status=404)
        workplace = employee.workplace
        col={0:'#9be5f7',1:'#6bf56e',2:'#ffff00',3:'#ffc000',4:'#ff7070'}
        name={0:'Nulo',1:'Bajo',2:'Medio',3:'Alto',4:'Muy alto'}
        val=None
        result_class=None
        result_text=None
        promedio=None
        survey_result=[]
        domainsA={
            'Condiciones en el ambiente de trabajo':['r2_p1','r2_p2','r2_p3'],
            'Carga de trabajo':['r2_p4','r2_p5','r2_p6','r2_p7','r2_p8','r2_p9','r2_p10','r2_p11','r2_p12','r2_p13','r2_p41','r2_p42','r2_p43'],
            'Falta de control sobre el trabajo':['r2_p18','r2_p19','r2_p20','r2_p21','r2_p22','r2_p26','r2_p27'],
            'Jornada de trabajo':['r2_p14','r2_p15'],
            'Interferencia en la relacion trabajo-familia':['r2_p16','r2_p17'],
            'Liderazgo':['r2_p23','r2_p24','r2_p25','r2_p28','r2_p29'],
            'Relaciones en el trabajo':['r2_p30','r2_p31','r2_p32','r2_p44','r2_p45','r2_p46'],
            'Violencia':['r2_p33','r2_p34','r2_p35','r2_p36','r2_p37','r2_p38','r2_p39','r2_p40'],
        }
        survey = employee.surveyA.filter(evaluation=evaluation).last()
        if survey:
            res = 0
            for field in RiskSurveyA._meta.fields:
                if field.name not in ['id','employee','evaluation','record_create','record_update','r2_p_a','r2_p_b','es_demo']:
                    res = res + (getattr(survey, field.attname) or 0)
            val = f'{res}/184'
            promedio = round((res/184)*100)
            result_class,result_text = ('#9be5f7','Riesgo nulo') if res<20 else (('#6bf56e','Riesgo bajo') if res<45 else (('#ffff00','Riesgo medio') if res<70 else (('#ffc000','Riesgo alto') if res<90 else ('#ff7070','Riesgo muy alto'))))
            for domain in domainsA:
                _sum = 0
                for question in domainsA[domain]:
                    try:
                        _sum = _sum + (getattr(survey, RiskSurveyA._meta.get_field(question).attname) or 0)
                    except Exception:
                        pass
                max_score = len(domainsA[domain])*4
                color = 0 if _sum < max_score*0.2 else (1 if _sum < max_score*0.4 else (2 if _sum < max_score*0.6 else (3 if _sum < max_score*0.8 else 4)))
                survey_result.append({'type':'Dominio','cat_dom':domain,'value':f'{_sum}/{max_score}','promedio':round((_sum/max_score)*100),'result_class':col[color],'result_text':name[color]})
        ctx = {
            'name': workplace.name,
            'address': workplace.address,
            'address_locality': workplace.address_locality,
            'address_state': workplace.address_state,
            'address_postal_code': workplace.address_postal_code,
            'phone': request.user.userapp.phone,
            'emp_name': employee.name,
            'gender': employee.get_gender_display(),
            'age': employee.get_age_display(),
            'civil_state': employee.get_civil_state_display(),
            'study_level': employee.get_study_level_display(),
            'ocupation': employee.ocupation,
            'department': employee.department,
            'charge_type': employee.get_charge_type_display(),
            'contract_type': employee.get_contract_type_display(),
            'employee_type': employee.get_employee_type_display(),
            'shift_type': employee.get_shift_type_display(),
            'shift_rotation': employee.get_shift_rotation_display(),
            'time_in_charge': employee.get_time_in_charge_display(),
            'exp': employee.get_exp_display(),
            'result_val': val or '',
            'result_class': result_class or '',
            'result_text': result_text or '',
            'promedio': promedio or '',
            'survey_result': survey_result,
            'print_mode': True,
        }
        return render(request, 'pdf/riesgo_psicosocial.html', ctx)


CLIMA_REACTIVOS = [
    (1,'liderazgo','Mi jefe inmediato comunica con claridad lo que espera de mi trabajo.'),
    (2,'liderazgo','Recibo apoyo de mi jefe inmediato cuando enfrento dificultades en mi trabajo.'),
    (3,'liderazgo','Mi jefe inmediato escucha mis opiniones antes de tomar decisiones que afectan mi trabajo.'),
    (4,'liderazgo','Recibo retroalimentacion util para mejorar mi desempeno.'),
    (5,'liderazgo','El trato que recibo de mi jefe inmediato es respetuoso.'),
    (6,'comunicacion','La informacion importante para realizar mi trabajo se comunica de manera oportuna.'),
    (7,'comunicacion','Los cambios relevantes dentro de la organizacion se comunican con claridad.'),
    (8,'comunicacion','Puedo expresar dudas o inquietudes sin temor a represalias.'),
    (9,'comunicacion','La comunicacion entre areas facilita el cumplimiento de los objetivos de trabajo.'),
    (10,'comunicacion','Conozco los canales adecuados para comunicar problemas, solicitudes o propuestas.'),
    (11,'trabajo_equipo','En mi area existe disposicion para colaborar entre companeros.'),
    (12,'trabajo_equipo','Mis companeros me apoyan cuando necesito ayuda para cumplir mis actividades.'),
    (13,'trabajo_equipo','El equipo de trabajo se coordina adecuadamente para alcanzar los objetivos.'),
    (14,'trabajo_equipo','Los conflictos entre companeros se manejan de manera respetuosa.'),
    (15,'trabajo_equipo','En mi equipo se comparten conocimientos y buenas practicas.'),
    (16,'reconocimiento','Mi esfuerzo y resultados son reconocidos por la organizacion o por mi jefe inmediato.'),
    (17,'reconocimiento','Cuando realizo bien mi trabajo, recibo algun tipo de reconocimiento o valoracion.'),
    (18,'reconocimiento','Me siento motivado para dar un buen desempeno en mi puesto.'),
    (19,'reconocimiento','La organizacion valora las aportaciones de sus colaboradores.'),
    (20,'reconocimiento','Considero que mi trabajo tiene importancia para los resultados de la organizacion.'),
    (21,'condiciones_trabajo','Cuento con las herramientas necesarias para realizar adecuadamente mi trabajo.'),
    (22,'condiciones_trabajo','El espacio fisico donde trabajo es adecuado para desempenar mis funciones.'),
    (23,'condiciones_trabajo','La tecnologia o sistemas que utilizo facilitan mi trabajo diario.'),
    (24,'condiciones_trabajo','Las condiciones de seguridad e higiene en mi lugar de trabajo son adecuadas.'),
    (25,'condiciones_trabajo','La organizacion proporciona los recursos necesarios para cumplir con mis responsabilidades.'),
    (26,'carga_equilibrio','La carga de trabajo que tengo es razonable para el tiempo disponible.'),
    (27,'carga_equilibrio','Las tareas se distribuyen de manera equilibrada dentro de mi equipo.'),
    (28,'carga_equilibrio','Puedo cumplir mis responsabilidades laborales sin afectar de forma constante mi vida personal.'),
    (29,'carga_equilibrio','Los tiempos de entrega de mis actividades suelen ser realistas.'),
    (30,'carga_equilibrio','La presion laboral que experimento es manejable.'),
    (31,'desarrollo_crecimiento','Tengo oportunidades para aprender nuevas habilidades dentro de la organizacion.'),
    (32,'desarrollo_crecimiento','La capacitacion que recibo contribuye a mejorar mi desempeno.'),
    (33,'desarrollo_crecimiento','Conozco las posibilidades de crecimiento o desarrollo que existen dentro de la organizacion.'),
    (34,'desarrollo_crecimiento','Mi puesto me permite desarrollar mis capacidades profesionales.'),
    (35,'desarrollo_crecimiento','La organizacion promueve el aprendizaje y la mejora continua.'),
    (36,'pertenencia','Me siento orgulloso de pertenecer a esta organizacion.'),
    (37,'pertenencia','Me identifico con los valores y objetivos de la organizacion.'),
    (38,'pertenencia','Recomendaria esta organizacion como un buen lugar para trabajar.'),
    (39,'pertenencia','Deseo continuar trabajando en esta organizacion en el futuro cercano.'),
    (40,'pertenencia','Siento que mi trabajo contribuye al proposito general de la organizacion.'),
]

CLIMA_DIMENSIONES = {
    'liderazgo': 'Liderazgo y supervision',
    'comunicacion': 'Comunicacion interna',
    'trabajo_equipo': 'Trabajo en equipo',
    'reconocimiento': 'Reconocimiento y motivacion',
    'condiciones_trabajo': 'Condiciones de trabajo',
    'carga_equilibrio': 'Carga laboral y equilibrio',
    'desarrollo_crecimiento': 'Desarrollo y crecimiento',
    'pertenencia': 'Sentido de pertenencia',
}

class ClimaLaboralView(View):
    def get(self, request, access_code):
        wk = Workplace.objects.filter(access_code=access_code).last()
        if not wk:
            return HttpResponse("Centro de trabajo no encontrado", status=404)
        if not getattr(wk.user.userapp, "stripe_plan_key", None) and not getattr(wk.user.userapp, "nom035_creditos", 0):
            return render(request, "clima_sin_plan.html", {"workplace": wk})
        if request.session.get(f'clima_submitted_{wk.id}'):
            return render(request, 'clima_gracias.html', {'workplace': wk, 'ya_enviado': True})
        departments = Employee.objects.filter(workplace=wk).values_list('department', flat=True).distinct()
        ctx = {
            'workplace': wk,
            'reactivos': CLIMA_REACTIVOS,
            'departments': [d for d in departments if d],
        }
        return render(request, 'clima_laboral.html', ctx)

    def post(self, request, access_code):
        wk = Workplace.objects.filter(access_code=access_code).last()
        if not wk:
            return HttpResponse("Centro de trabajo no encontrado", status=404)
        if not getattr(wk.user.userapp, "stripe_plan_key", None) and not getattr(wk.user.userapp, "nom035_creditos", 0):
            return HttpResponse("Sin plan activo", status=403)
        if request.session.get(f'clima_submitted_{wk.id}'):
            return render(request, 'clima_gracias.html', {'workplace': wk, 'ya_enviado': True})
        department = request.POST.get('department', '')
        data = {'workplace': wk, 'department': department}
        for i in range(1, 41):
            val = request.POST.get(f'cl_p{i}')
            data[f'cl_p{i}'] = int(val) if val and val.isdigit() else None
        WorkEnvironmentSurvey.objects.create(**data)
        request.session[f'clima_submitted_{wk.id}'] = True
        return render(request, 'clima_gracias.html', {'workplace': wk})

class ClimaResultadosView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')
    def get(self, request, workplace_id):
        wk = Workplace.objects.filter(id=workplace_id, user=request.user).last()
        if not wk:
            return HttpResponseRedirect(reverse_lazy('workplaces'))
        surveys = WorkEnvironmentSurvey.objects.filter(workplace=wk)
        total = surveys.count()
        dimensiones = {}
        for codigo, nombre in CLIMA_DIMENSIONES.items():
            reactivos_dim = [r[0] for r in CLIMA_REACTIVOS if r[1] == codigo]
            promedios = []
            for s in surveys:
                vals = [getattr(s, f'cl_p{i}') for i in reactivos_dim if getattr(s, f'cl_p{i}') is not None]
                if len(vals) >= 3:
                    promedios.append(sum(vals)/len(vals))
            promedio_dim = round(sum(promedios)/len(promedios), 2) if promedios else None
            if promedio_dim:
                if promedio_dim < 2.5: nivel = ('Critico', '#ff7070')
                elif promedio_dim < 3.5: nivel = ('En riesgo', '#ffc000')
                elif promedio_dim < 4.25: nivel = ('Adecuado', '#ffff00')
                else: nivel = ('Favorable', '#6bf56e')
            else:
                nivel = ('Sin datos', '#e2e8f0')
            dimensiones[codigo] = {'nombre': nombre, 'promedio': promedio_dim, 'nivel': nivel[0], 'color': nivel[1]}
        ctx = {
            'workplace': wk,
            'total': total,
            'dimensiones': dimensiones,
            'access_url': request.build_absolute_uri(f'/clima/{wk.access_code}/'),
            'workplaces': Workplace.objects.filter(user=request.user),
        }
        return render(request, 'clima_resultados.html', ctx)

class BorrarDemoView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')
    def post(self, request):
        try:
            from django.core.management import call_command
            call_command('borrar_datos_demo', request.user.id)
            return JsonResponse({'status': 'ok', 'msg': 'Datos de prueba eliminados correctamente.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'msg': str(e)}, status=500)
